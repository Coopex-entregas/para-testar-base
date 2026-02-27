import os
import io
import re
import json
import random
from flask_socketio import SocketIO
import unicodedata
from datetime import datetime, timedelta, time, date
from collections import Counter, defaultdict
from urllib.parse import urlparse, parse_qs
from functools import wraps
from decimal import Decimal

from flask import (
    Flask, render_template, render_template_string, request, redirect, url_for,
    flash, session, send_file, jsonify, abort, current_app, 
)
from flask_login import LoginManager, login_user, logout_user, current_user, login_required
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func
from sqlalchemy.orm import joinedload
from sqlalchemy.sql import text
from sqlalchemy.exc import IntegrityError
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from itsdangerous import URLSafeSerializer, BadSignature

import pandas as pd
import holidays
import pytz
from jinja2 import TemplateNotFound

# =========================================================
# CONFIGURAÇÃO BÁSICA
# =========================================================
app = Flask(__name__)

# Usa a mesma chave que você já tinha, só mudando para config
app.config['SECRET_KEY'] = os.environ.get(
    'SECRET_KEY',
    'COOPEX_ULTRA_SEGURA_2024_FIXA'
)

# 🔽 INSTÂNCIA DO SOCKETIO LIGADA NO APP
from flask_socketio import SocketIO

socketio = SocketIO(
    app,
    async_mode="threading",   # (opcional, mas bom deixar explícito)
    logger=False,
    engineio_logger=False
)


# --- Admins fixos (usuario: coopex, 2 senhas) ---
ADMIN_CREDENTIALS = {
    'coopex': {
        os.environ.get('ADMIN_PWD_COOPEX_MASTER', 'coopex05289'): {'is_master': True},
        os.environ.get('ADMIN_PWD_COOPEX',        '84253700'):     {'is_master': False},
    }
}

# ------------------------
# Configuração do Banco
# ------------------------
database_url = os.environ.get('DATABASE_URL', 'sqlite:///db.sqlite3')

# Render (e outros serviços) costumam vir com "postgres://"
# O SQLAlchemy precisa de "postgresql+psycopg2://"
if database_url.startswith("postgres://"):
    database_url = database_url.replace("postgres://", "postgresql+psycopg2://", 1)

app.config['SQLALCHEMY_DATABASE_URI'] = database_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    "pool_pre_ping": True,
    "pool_recycle": 300,  # 5 minutos
    "pool_size": 5,
    "max_overflow": 10,
}

db = SQLAlchemy(app)

# =========================================================
# COMPROVANTE DE ENTREGA (FOTO) — armazenado por 7 dias
# =========================================================
COMPROVANTE_DIR = os.path.join(app.instance_path, "comprovantes")
COMPROVANTE_INDEX = os.path.join(app.instance_path, "comprovantes_index.json")
COMPROVANTE_TTL_DAYS = 7

def _ensure_comprovante_dirs():
    try:
        os.makedirs(COMPROVANTE_DIR, exist_ok=True)
    except Exception:
        pass

def _load_comprovante_index():
    _ensure_comprovante_dirs()
    data = {}
    try:
        if os.path.exists(COMPROVANTE_INDEX):
            with open(COMPROVANTE_INDEX, "r", encoding="utf-8") as f:
                data = json.load(f) or {}
    except Exception:
        data = {}
    _cleanup_comprovantes(data)
    return data

def _save_comprovante_index(data: dict):
    _ensure_comprovante_dirs()
    try:
        tmp = COMPROVANTE_INDEX + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        os.replace(tmp, COMPROVANTE_INDEX)
    except Exception:
        pass

from typing import Optional, Dict, Any

def _cleanup_comprovantes(index_data: Optional[Dict[str, Any]] = None):
    # apaga arquivos com mais de 7 dias (por mtime) e limpa o index
    _ensure_comprovante_dirs()
    cutoff = datetime.utcnow() - timedelta(days=COMPROVANTE_TTL_DAYS)
    try:
        for name in os.listdir(COMPROVANTE_DIR):
            p = os.path.join(COMPROVANTE_DIR, name)
            try:
                mtime = datetime.utcfromtimestamp(os.path.getmtime(p))
                if mtime < cutoff:
                    os.remove(p)
            except Exception:
                pass
    except Exception:
        pass

    if index_data is None:
        return

    # remove entradas expiradas ou sem arquivo
    changed = False
    for k in list(index_data.keys()):
        fn = (index_data.get(k) or {}).get("filename")
        if not fn:
            index_data.pop(k, None); changed = True; continue
        fp = os.path.join(COMPROVANTE_DIR, fn)
        if not os.path.exists(fp):
            index_data.pop(k, None); changed = True; continue
        try:
            mtime = datetime.utcfromtimestamp(os.path.getmtime(fp))
            if mtime < cutoff:
                try: os.remove(fp)
                except Exception: pass
                index_data.pop(k, None); changed = True
        except Exception:
            pass
    if changed:
        _save_comprovante_index(index_data)

def comprovante_info(entrega_id: int):
    idx = _load_comprovante_index()
    return idx.get(str(entrega_id))

def comprovante_existe(entrega_id: int) -> bool:
    info = comprovante_info(entrega_id)
    if not info: 
        return False
    fn = info.get("filename")
    if not fn:
        return False
    return os.path.exists(os.path.join(COMPROVANTE_DIR, fn))

def _salvar_comprovante(entrega_id: int, file_storage):
    _ensure_comprovante_dirs()
    _cleanup_comprovantes()
    if not file_storage:
        return None

    filename = secure_filename(file_storage.filename or "")
    ext = os.path.splitext(filename)[1].lower()
    if ext not in [".jpg", ".jpeg", ".png", ".webp"]:
        # tenta salvar como jpg se vier sem extensão correta
        ext = ".jpg"

    ts = datetime.utcnow().strftime("%Y%m%d%H%M%S")
    out_name = f"entrega_{entrega_id}_{ts}{ext}"
    out_path = os.path.join(COMPROVANTE_DIR, out_name)
    file_storage.save(out_path)

    idx = _load_comprovante_index()
    idx[str(entrega_id)] = {"filename": out_name, "uploaded_at": datetime.utcnow().isoformat() + "Z"}
    _save_comprovante_index(idx)
    return out_name

# =========================================================
# RASTREIO POR LINK (por entrega)
# =========================================================
def _rastreio_serializer():
    return URLSafeSerializer(app.config["SECRET_KEY"], salt="rastreio_entrega_v1")

def gerar_token_rastreio(entrega_id: int):
    return _rastreio_serializer().dumps({"entrega_id": int(entrega_id)})

def ler_token_rastreio(token: str):
    return _rastreio_serializer().loads(token)


# =========================================================
# FLASK-LOGIN / LOGIN MANAGER
# =========================================================
login_manager = LoginManager()
login_manager.init_app(app)

# nome da view/rota que mostra a tela de login
# (ajuste se sua função de login tiver outro endpoint, tipo 'login_admin')
login_manager.login_view = 'login'


@login_manager.user_loader
def load_user(user_id):
    """
    Função usada pelo Flask-Login para carregar o usuário
    a partir do ID salvo na sessão.
    Importa o modelo aqui dentro para evitar problemas de import circular.
    """
    from models import Usuario  # importa só quando necessário
    try:
        return Usuario.query.get(int(user_id))
    except (ValueError, TypeError):
        return None

# Fuso Brasil
BRAZIL_TZ = pytz.timezone('America/Sao_Paulo')

# =========================================================
# MODELS
# =========================================================
class Cooperado(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(100), nullable=False)
    senha_hash = db.Column(db.String(128), nullable=False)
    ativo = db.Column(db.Boolean, nullable=False, default=True)

    # NOVOS CAMPOS PARA RASTREIO EM TEMPO REAL
    last_lat = db.Column(db.Float, nullable=True)
    last_lng = db.Column(db.Float, nullable=True)
    last_ping = db.Column(db.DateTime, nullable=True)
    online = db.Column(db.Boolean, nullable=False, default=False)

    # NOVOS CAMPOS (tempo real)
    last_speed_kmh = db.Column(db.Float, nullable=True)
    last_heading = db.Column(db.Float, nullable=True)
    last_accuracy_m = db.Column(db.Float, nullable=True)

    last_moving_at = db.Column(db.DateTime, nullable=True)  # última vez que estava se movendo

    def set_senha(self, senha):
        self.senha_hash = generate_password_hash(senha)

    def check_senha(self, senha):
        return check_password_hash(self.senha_hash, senha)


class Cliente(db.Model):
    __tablename__ = 'cliente'

    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(100), nullable=False)
    telefone = db.Column(db.String(20), nullable=True)
    bairro_origem = db.Column(db.String(80), nullable=True)
    endereco = db.Column(db.String(255), nullable=True)

    saldo_atual = db.Column(db.Float, default=0.0)

    username = db.Column(db.String(80), unique=True, nullable=True)
    senha_hash = db.Column(db.String(128), nullable=True)

    email = db.Column(db.String(120), unique=True, nullable=True)
    reset_code = db.Column(db.String(10), nullable=True)
    reset_expires_at = db.Column(db.DateTime, nullable=True)

    def set_senha(self, senha: str) -> None:
        self.senha_hash = generate_password_hash(senha)

    def check_senha(self, senha: str) -> bool:
        if not self.senha_hash:
            return False
        return check_password_hash(self.senha_hash, senha)


class Entrega(db.Model):
    __tablename__ = 'entrega'

    id = db.Column(db.Integer, primary_key=True)

    # Informações básicas da entrega
    cliente = db.Column(db.String(100), nullable=False)
    bairro = db.Column(db.String(50), nullable=False)   # bairro principal da corrida (pode ser o final)
    valor = db.Column(db.Float, nullable=False)

    # Datas/horas (UTC no banco; você converte para America/Sao_Paulo na view)
    data_envio = db.Column(
        db.DateTime,
        nullable=False,
        default=datetime.utcnow,  # UTC naive; converter na view
    )
    data_atribuida = db.Column(db.DateTime, nullable=True)

    # Relação com cooperado
    cooperado_id = db.Column(
        db.Integer,
        db.ForeignKey('cooperado.id'),
        nullable=True
    )
    cooperado = db.relationship('Cooperado', backref='entregas')

    # Pagamento / status geral
    status_pagamento = db.Column(db.String(20), nullable=True)  # pago / pendente
    status = db.Column(db.String(20), nullable=True)            # entregue / pendente / etc.
    pagamento = db.Column(db.String(50), nullable=False)        # PIX, dinheiro, etc.
    recebido_por = db.Column(db.String(100), nullable=True)

    # Controle de crédito usado nesta entrega
    credito_usado = db.Column(db.Float, nullable=False, default=0.0)
    credito_mov_id = db.Column(db.Integer, nullable=True)

    # Link explícito com Cliente (tabela cliente)
    cliente_id = db.Column(
        db.Integer,
        db.ForeignKey('cliente.id'),
        nullable=True
    )

    # JSON com dados de ORIGEM (coleta)
    # Pode conter endereço completo OU apenas o bairro.
    # Exemplos:
    #   {"endereco": "Rua X, 123", "bairro": "Lagoa Nova", "ref": "Portaria azul",
    #    "lat": -5.79, "lng": -35.21}
    #   {"bairro": "Lagoa Nova"}
    origem_json = db.Column(db.Text, nullable=True)

    # JSON com dados de DESTINO FINAL (entrega)
    # Pode conter endereço de entrega completo OU só o bairro de destino final.
    # Exemplos:
    #   {"endereco": "Av. Y, 999", "bairro": "Tirol", "ref": "Em frente ao hospital",
    #    "lat": -5.80, "lng": -35.20}
    #   {"bairro": "Tirol"}
    destino_json = db.Column(db.Text, nullable=True)

    # JSON com PARADAS INTERMEDIÁRIAS
    # Pode ser usado para:
    #   - 1 parada somente
    #   - várias paradas
    # Cada parada pode ter endereço completo ou só bairro.
    # Exemplo:
    #   {
    #     "stops": [
    #       {"endereco": "Rua A, 12", "bairro": "Barro Vermelho"},
    #       {"endereco": "Av. B, 456", "bairro": "Petrópolis", "ref": "Padaria tal"}
    #     ]
    #   }
    # Ou apenas bairros:
    #   {
    #     "stops": [
    #       {"bairro": "Barro Vermelho"},
    #       {"bairro": "Petrópolis"}
    #     ]
    #   }
    paradas_json = db.Column(db.Text, nullable=True)

    # Status da corrida na visão do cooperado
    # pendente  -> tocando "chamada.mp3", aguardando aceite
    # aceita    -> cooperado aceitou, em andamento
    # recusada  -> cooperado recusou
    status_corrida = db.Column(
        db.String(20),
        nullable=False,
        default='pendente'
    )

    # =========================
    #   HELPERS DE ORIGEM
    # =========================

    def set_origem(self, endereco=None, bairro=None, ref=None, lat=None, lng=None, extra=None):
        """
        Seta o JSON de origem.

        Pode chamar só com bairro:
            set_origem(bairro="Lagoa Nova")

        Ou com endereço completo:
            set_origem(
                endereco="Rua X, 123",
                bairro="Lagoa Nova",
                ref="Portaria azul",
                lat=-5.79,
                lng=-35.21
            )
        """
        data = {}
        if endereco:
            data["endereco"] = endereco
        if bairro:
            data["bairro"] = bairro
        if ref:
            data["ref"] = ref
        if lat is not None:
            data["lat"] = lat
        if lng is not None:
            data["lng"] = lng
        if extra and isinstance(extra, dict):
            data.update(extra)

        self.origem_json = json.dumps(data, ensure_ascii=False) if data else None

    def get_origem(self):
        if not self.origem_json:
            return {}
        try:
            return json.loads(self.origem_json)
        except Exception:
            return {}

    # =========================
    #   HELPERS DE DESTINO
    # =========================

    def set_destino(self, endereco=None, bairro=None, ref=None, lat=None, lng=None, extra=None):
        """
        Seta o JSON de destino final.

        Pode chamar só com bairro:
            set_destino(bairro="Tirol")

        Ou com endereço completo:
            set_destino(
                endereco="Av. Y, 999",
                bairro="Tirol",
                ref="Em frente ao hospital",
                lat=-5.80,
                lng=-35.20
            )
        """
        data = {}
        if endereco:
            data["endereco"] = endereco
        if bairro:
            data["bairro"] = bairro
        if ref:
            data["ref"] = ref
        if lat is not None:
            data["lat"] = lat
        if lng is not None:
            data["lng"] = lng
        if extra and isinstance(extra, dict):
            data.update(extra)

        self.destino_json = json.dumps(data, ensure_ascii=False) if data else None

    def get_destino(self):
        if not self.destino_json:
            return {}
        try:
            return json.loads(self.destino_json)
        except Exception:
            return {}

    # =========================
    #   HELPERS DE PARADAS
    # =========================

    def _get_paradas_dict(self):
        if not self.paradas_json:
            return {"stops": []}
        try:
            data = json.loads(self.paradas_json)
            if "stops" not in data or not isinstance(data["stops"], list):
                data["stops"] = []
            return data
        except Exception:
            return {"stops": []}

    def get_paradas(self):
        """
        Retorna uma lista de paradas.

        Cada item é um dict, ex:
          {"endereco": "...", "bairro": "...", "ref": "..."}
        ou
          {"bairro": "Tirol"}
        """
        data = self._get_paradas_dict()
        return data.get("stops", [])

    def set_paradas(self, lista_paradas):
        """
        Seta TODAS as paradas de uma vez.

        Exemplo de lista_paradas:
          [
            {"endereco": "Rua A, 12", "bairro": "Barro Vermelho"},
            {"bairro": "Petrópolis"}
          ]
        """
        if not lista_paradas:
            self.paradas_json = None
            return

        # Garante que seja sempre lista de dicts
        stops = []
        for parada in lista_paradas:
            if isinstance(parada, dict):
                stops.append(parada)

        data = {"stops": stops}
        self.paradas_json = json.dumps(data, ensure_ascii=False)

    def add_parada(self, endereco=None, bairro=None, ref=None, lat=None, lng=None, extra=None):
        """
        Adiciona UMA parada à lista de paradas.

        Pode ser só bairro:
            add_parada(bairro="Petrópolis")

        Ou com endereço completo:
            add_parada(
                endereco="Rua A, 12",
                bairro="Barro Vermelho",
                ref="Ao lado do mercado"
            )
        """
        data = self._get_paradas_dict()
        parada = {}
        if endereco:
            parada["endereco"] = endereco
        if bairro:
            parada["bairro"] = bairro
        if ref:
            parada["ref"] = ref
        if lat is not None:
            parada["lat"] = lat
        if lng is not None:
            parada["lng"] = lng
        if extra and isinstance(extra, dict):
            parada.update(extra)

        if parada:
            data["stops"].append(parada)

        self.paradas_json = json.dumps(data, ensure_ascii=False)

    # =========================
    #   UTIL
    # =========================

    def __repr__(self):
        return f'<Entrega {self.id} - {self.cliente} - {self.bairro} - R${self.valor:.2f}>'


# =========================================================
# HELPER: EMITIR ATUALIZAÇÃO EM TEMPO REAL
# =========================================================
def emitir_atualizacao_entrega(entrega: Entrega, acao: str):
    """
    Emite para todos os painéis (admin, cooperado, rastreamento) que
    uma entrega foi criada / editada / excluída / status alterado.

    Evento Socket.IO: 'entrega_atualizada'
    """
    if not entrega:
        return

    try:
        payload = {
            "id": entrega.id,
            "acao": acao,  # 'criada', 'editada', 'excluida', etc.
            "cliente": entrega.cliente,
            "bairro": entrega.bairro,
            "valor": float(entrega.valor or 0),
            "status": entrega.status,
            "status_pagamento": entrega.status_pagamento,
            "pagamento": entrega.pagamento,
            "cooperado_id": entrega.cooperado_id,
            "cooperado_nome": entrega.cooperado.nome if entrega.cooperado else None,
            "data_envio": (
                to_brasilia(entrega.data_envio).strftime('%Y-%m-%d %H:%M')
                if entrega.data_envio else None
            ),
            "data_atribuida": (
                to_brasilia(entrega.data_atribuida).strftime('%Y-%m-%d %H:%M')
                if entrega.data_atribuida else None
            ),
        }

        # Evento específico para os painéis de entregas
        socketio.emit(
            "entrega_atualizada",
            payload)

    except Exception as e:
        # não quebra o fluxo se der problema no websocket
        try:
            current_app.logger.warning(f'Falha ao emitir entrega_atualizada: {e}')
        except Exception:
            pass

def emitir_posicao_motoboy(cooperado: Cooperado, lat: float, lng: float, velocidade=None):
    try:
        ultima_str = ""
        if cooperado.last_ping:
            ultima_str = to_brasilia(cooperado.last_ping).strftime('%d/%m %H:%M:%S')

        is_online, idle_s, status_str = calc_status_cooperado(cooperado)

        payload = {
            'id': cooperado.id,
            'nome': cooperado.nome,
            'lat': float(lat),
            'lng': float(lng),

            'online': bool(is_online),
            'status': status_str,                 # offline | ocioso | livre | em_corrida
            'idle_seconds': idle_s,               # tempo ocioso em segundos (se online)

            'velocidade_kmh': float(velocidade) if velocidade is not None else None,
            'heading': cooperado.last_heading,
            'accuracy_m': cooperado.last_accuracy_m,

            'ultima_atualizacao': ultima_str,
        }

        socketio.emit('posicao_motoboy_atualizada', payload, broadcast=True)

    except Exception as e:
        try:
            current_app.logger.warning(f'Falha ao emitir posicao_motoboy_atualizada: {e}')
        except Exception:
            pass


class Credito(db.Model):
    __tablename__ = 'credito'

    id = db.Column(db.Integer, primary_key=True)
    cliente_id = db.Column(db.Integer, db.ForeignKey('cliente.id'), nullable=False)

    # CAMPOS ANTIGOS (mantidos pra compatibilidade)
    valor = db.Column(db.Float, default=0.0)          # pode ser o mesmo que valor_final
    saldo_atual = db.Column(db.Float, default=0.0)    # saldo do cliente após este crédito (opcional)

    # CAMPOS NOVOS – são exatamente os usados no creditos.html
    valor_bruto = db.Column(db.Float, default=0.0)    # valor original do crédito
    desconto_tipo = db.Column(db.String(20))          # 'nenhum', 'percentual', 'real'
    desconto_valor = db.Column(db.Float, default=0.0) # número usado no desconto
    valor_final = db.Column(db.Float, default=0.0)    # valor_bruto - desconto aplicado

    saldo_antes = db.Column(db.Float, default=0.0)    # saldo do cliente ANTES deste crédito
    saldo_depois = db.Column(db.Float, default=0.0)   # saldo do cliente DEPOIS deste crédito

    motivo = db.Column(db.String(180))                # observação
    criado_em = db.Column(db.DateTime, default=datetime.utcnow)
    criado_por = db.Column(db.String(80))

    movimentos = db.relationship(
        'CreditoMovimento',
        backref='credito',
        lazy=True,
        cascade='all, delete-orphan'
    )

    def __repr__(self):
        return f'<Credito {self.id} - Cliente {self.cliente_id} - R${self.valor_final:.2f}>'


class CreditoMovimento(db.Model):
    __tablename__ = 'credito_movimento'

    id = db.Column(db.Integer, primary_key=True)

    credito_id = db.Column(
        db.Integer,
        db.ForeignKey('credito.id', ondelete='CASCADE'),
        nullable=True
    )

    # NOVO CAMPO: para não dar mais erro no cliente_id=...
    cliente_id = db.Column(
        db.Integer,
        db.ForeignKey('cliente.id'),
        nullable=True
    )

    # NOVO CAMPO: ligação opcional com entrega (para rastrear consumo)
    entrega_id = db.Column(
        db.Integer,
        db.ForeignKey('entrega.id'),
        nullable=True
    )

    tipo = db.Column(db.String(20), nullable=False)   # 'credito' ou 'debito'
    valor = db.Column(db.Float, nullable=False)

    # CAMPO ANTIGO (pode ficar por compatibilidade)
    data = db.Column(db.DateTime, default=datetime.utcnow)

    # CAMPO NOVO USADO EM VÁRIAS TELAS
    criado_em = db.Column(db.DateTime, default=datetime.utcnow)

    descricao = db.Column(db.String(255))
    referencia = db.Column(db.String(255))


class ListaEspera(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(100), nullable=False)  # legado
    cooperado_id = db.Column(db.Integer, db.ForeignKey('cooperado.id'), nullable=True)
    pos = db.Column(db.Integer, nullable=True)
    created_at = db.Column(db.DateTime, nullable=True, default=datetime.utcnow)
    cooperado = db.relationship('Cooperado', lazy='joined')

def emitir_lista_espera():
    """
    Emite para todos os painéis a situação atual da fila de espera.
    """
    try:
        itens = (
            ListaEspera.query
            .order_by(ListaEspera.pos.asc(), ListaEspera.created_at.asc())
            .all()
        )
        payload = []
        for item in itens:
            payload.append({
                "id": item.id,
                "cooperado_id": item.cooperado_id,
                "nome": item.cooperado.nome if item.cooperado else item.nome,
                "pos": item.pos,
                "created_at": to_brasilia(item.created_at).strftime('%d/%m %H:%M')
                              if item.created_at else "",
            })

        socketio.emit(
            "fila_espera_atualizada",
            {"itens": payload},
            broadcast=True
        )
    except Exception as e:
        try:
            current_app.logger.warning(f"Falha ao emitir fila_espera_atualizada: {e}")
        except Exception:
            pass

class Trajeto(db.Model):
    __tablename__ = 'trajeto'

    id = db.Column(db.Integer, primary_key=True)

    # Quem fez o trajeto
    cooperado_id = db.Column(db.Integer, db.ForeignKey('cooperado.id'), nullable=False)
    cooperado = db.relationship('Cooperado', backref='trajetos')

    # Horários em UTC naive (igual ao resto do sistema)
    inicio = db.Column(db.DateTime, nullable=False)   # quando começou o trajeto
    fim = db.Column(db.DateTime, nullable=True)       # quando terminou (se tiver)

    # Métricas principais
    distancia_m = db.Column(db.Float, nullable=True)          # em metros
    duracao_s = db.Column(db.Integer, nullable=True)          # em segundos
    velocidade_media_kmh = db.Column(db.Float, nullable=True) # km/h

    # Coordenadas (opcionais)
    origem_lat = db.Column(db.Float, nullable=True)
    origem_lng = db.Column(db.Float, nullable=True)
    destino_lat = db.Column(db.Float, nullable=True)
    destino_lng = db.Column(db.Float, nullable=True)

    # JSON com pontos do trajeto (lista de lat/lng/hora) – para futuro "vídeo" no mapa
    pontos_json = db.Column(db.Text, nullable=True)

    # Quando foi gravado no sistema
    criado_em = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)


# =========================================================
# HELPERS DE DATA / FUSO
# =========================================================
def to_brasilia(dt):
    if not dt:
        return None
    if dt.tzinfo is None:
        dt = pytz.utc.localize(dt)
    return dt.astimezone(BRAZIL_TZ)

from datetime import datetime, timezone
import os

# -------------------------------------------------------------------
# CONFIG DE TEMPOS (para NÃO dar NameError)
# Ajuste os valores como quiser. Também aceita env vars no Render.
# -------------------------------------------------------------------
OFFLINE_AFTER_SEC = int(os.getenv("OFFLINE_AFTER_SEC", "120"))  # 2 min sem ping => offline
IDLE_AFTER_SEC    = int(os.getenv("IDLE_AFTER_SEC", "300"))     # 5 min sem movimento => ocioso
# Se velocidade >= isso, considera "em movimento"
MOVING_SPEED_KMH = float(os.getenv("MOVING_SPEED_KMH", "3.0"))


def _to_utc_aware(dt):
    """
    Garante datetime timezone-aware em UTC.
    - None -> None
    - naive -> assume que está em UTC e adiciona tzinfo
    - aware -> converte para UTC
    """
    if dt is None:
        return None
    if dt.tzinfo is None:
        return dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc)


def calc_status_cooperado(c):
    """
    Retorna: (is_online, idle_seconds, status_str)
    status_str: offline | ocioso | livre | em_corrida
    """
    # Agora é UTC-aware (não dá erro com TIMESTAMPTZ)
    now_utc = datetime.now(timezone.utc)

    last_ping = _to_utc_aware(getattr(c, "last_ping", None))
    last_moving_at = _to_utc_aware(getattr(c, "last_moving_at", None))

    # ONLINE “REAL” = ping recente
    is_online = bool(getattr(c, "online", False)) and (last_ping is not None)
    if is_online:
        delta = (now_utc - last_ping).total_seconds()
        if delta > OFFLINE_AFTER_SEC:
            is_online = False

    if not is_online:
        return (False, None, "offline")

    # Ocioso = online, mas sem movimento por tempo
    if last_moving_at:
        idle_seconds = int((now_utc - last_moving_at).total_seconds())
    else:
        # se nunca marcou movimento, usa last_ping como referência
        idle_seconds = int((now_utc - last_ping).total_seconds()) if last_ping else 0

    # Se você tem “em_corrida/ocupado” no cooperado, priorize isso:
    em_corrida = bool(getattr(c, "em_corrida", False) or getattr(c, "ocupado", False))
    if em_corrida:
        return (True, idle_seconds, "em_corrida")

    if idle_seconds >= IDLE_AFTER_SEC:
        return (True, idle_seconds, "ocioso")

    return (True, idle_seconds, "livre")


def local_date_window_to_utc_range(local_date: date):
    inicio_brasil = BRAZIL_TZ.localize(datetime.combine(local_date, time.min))
    fim_brasil = BRAZIL_TZ.localize(datetime.combine(local_date, time.max))
    return (
        inicio_brasil.astimezone(pytz.utc).replace(tzinfo=None),
        fim_brasil.astimezone(pytz.utc).replace(tzinfo=None),
    )


def month_range_utc(local_date: date):
    first = local_date.replace(day=1)
    next_first = (
        first.replace(year=first.year + 1, month=1, day=1)
        if first.month == 12
        else first.replace(month=first.month + 1, day=1)
    )
    return (
        local_date_window_to_utc_range(first)[0],
        local_date_window_to_utc_range(next_first - timedelta(days=1))[1],
    )


def year_range_utc(local_date: date):
    first = local_date.replace(month=1, day=1)
    next_first = first.replace(year=first.year + 1)
    return (
        local_date_window_to_utc_range(first)[0],
        local_date_window_to_utc_range(next_first - timedelta(days=1))[1],
    )


def parse_local_datetime_to_utc_naive(data_str: str):
    dt_local_naive = datetime.strptime(data_str, '%Y-%m-%dT%H:%M')
    dt_local = BRAZIL_TZ.localize(dt_local_naive)
    return dt_local.astimezone(pytz.utc).replace(tzinfo=None)


def diasemana(data):
    dias = ['Seg', 'Ter', 'Qua', 'Qui', 'Sex', 'Sáb', 'Dom']
    return dias[data.weekday()]

app.jinja_env.filters['diasemana'] = diasemana
app.jinja_env.globals['tem_comprovante'] = comprovante_existe
app.jinja_env.globals['token_rastreio'] = gerar_token_rastreio

# =========================================================
# RASTREAMENTO - HELPER DE LINHA DO TEMPO
# =========================================================
def montar_eventos_rastreamento(entrega: Entrega):
    """
    Gera uma listinha de eventos para exibir a linha do tempo do rastreio.
    Usa os dados que já existem na tabela entrega (data_envio, data_atribuida,
    status, recebido_por, cooperado etc.).
    """
    eventos = []

    # 1) Pedido criado
    dt_criacao = to_brasilia(entrega.data_envio)
    if dt_criacao:
        eventos.append({
            "titulo": "Pedido criado",
            "descricao": f"Entrega registrada para o cliente {entrega.cliente or '---'}",
            "quando": dt_criacao,
            "icone": "📦"
        })

    # 2) Motoboy atribuído
    if entrega.cooperado_id and entrega.cooperado:
        dt_att = to_brasilia(entrega.data_atribuida or entrega.data_envio)
        eventos.append({
            "titulo": "Motoboy atribuído",
            "descricao": f"Cooperado: {entrega.cooperado.nome}",
            "quando": dt_att,
            "icone": "🏍️"
        })

    # 3) Saiu para entrega (se tiver cooperado e status diferente de pendente)
    st = (entrega.status or '').strip().lower()
    if entrega.cooperado_id and st not in ('', 'pendente', 'aguardando'):
        dt_envio = to_brasilia(entrega.data_atribuida or entrega.data_envio)
        eventos.append({
            "titulo": "Saiu para entrega",
            "descricao": "Pedido está em rota de entrega.",
            "quando": dt_envio,
            "icone": "🚚"
        })

    # 4) Entrega concluída
    if st in ('entregue', 'recebido'):
        eventos.append({
            "titulo": "Entrega concluída",
            "descricao": f"Recebido por: {entrega.recebido_por or 'destinatário'}",
            # não temos hora exata do recebimento, então reaproveito data_atribuida/envio
            "quando": to_brasilia(entrega.data_atribuida or entrega.data_envio),
            "icone": "✅"
        })

    # Garante ordenação por data (quando não for None)
    eventos.sort(key=lambda ev: ev["quando"] or datetime.min)

    return eventos


# =========================================================
# NORMALIZAÇÃO DE TEXTO / NOME / PAGAMENTO
# =========================================================
def _strip_accents(s: str) -> str:
    return ''.join(
        c for c in unicodedata.normalize('NFD', s or '')
        if unicodedata.category(c) != 'Mn'
    )


def normalize_letters_key(s: str) -> str:
    s = _strip_accents(s).lower()
    s = re.sub(r'[^a-z\u00c0-\u024f\s]', ' ', s)
    return re.sub(r'\s+', ' ', s).strip()


def normalize_first_token(s: str) -> str:
    k = normalize_letters_key(s)
    return (k.split(' ')[0] if k else '')


def pagamento_usa_credito(pagamento: str) -> bool:
    """
    True se a forma de pagamento usar crédito.
    Ex:
      - "Crédito"
      - "Credito"
      - "Crédito automático"
      - "Crédito + Pix"
    """
    txt = _strip_accents((pagamento or '').strip().lower())
    txt = re.sub(r'\s+', ' ', txt)
    return txt.startswith('credito')



# =========================================================
# CRÉDITO: HELPERS E REGRAS
# =========================================================
def _as_decimal(x) -> Decimal:
    if x is None:
        return Decimal("0.00")
    if isinstance(x, Decimal):
        return x
    return Decimal(str(x)).quantize(Decimal("0.01"))


def calcular_valor_final(valor_bruto, desconto_tipo, desconto_valor) -> Decimal:
    """
    Mantido por compatibilidade, mas na prática vamos usar SEM desconto.
    No formulário novo, desconto_tipo='nenhum' e desconto_valor=0.
    """
    bruto = _as_decimal(valor_bruto)
    d = _as_decimal(desconto_valor)
    if desconto_tipo == "percentual":
        desc = (bruto * d) / Decimal("100")
    elif desconto_tipo == "real":
        desc = d
    else:
        desc = Decimal("0.00")
    if desc > bruto:
        desc = bruto
    return (bruto - desc).quantize(Decimal("0.01"))


def _find_cliente_by_nome(nome: str):
    if not nome:
        return None

    # busca exata (lower)
    cli = Cliente.query.filter(func.lower(Cliente.nome) == (nome or '').lower()).first()
    if cli:
        return cli

    # fallback por normalização forte
    target = normalize_letters_key(nome or '')
    for c in Cliente.query.all():
        if normalize_letters_key(c.nome or '') == target:
            return c

    # último recurso: 1º token
    tok = normalize_first_token(nome or '')
    for c in Cliente.query.all():
        if normalize_first_token(c.nome or '') == tok:
            return c
    return None


def atualizar_saldo_credito_cliente(cliente_id):
    """
    Recalcula o saldo do cliente SOMENTE pelos movimentos em CreditoMovimento.

    Isso garante:
      - Se você excluir um crédito, remover seus movimentos e chamar esta função,
        o saldo volta a ser o que sobrar dos outros movimentos.
    """
    total_creditos = (
        db.session.query(func.coalesce(func.sum(CreditoMovimento.valor), 0.0))
        .filter(
            CreditoMovimento.cliente_id == cliente_id,
            CreditoMovimento.tipo == 'credito'
        )
        .scalar()
        or 0.0
    )

    total_debitos = (
        db.session.query(func.coalesce(func.sum(CreditoMovimento.valor), 0.0))
        .filter(
            CreditoMovimento.cliente_id == cliente_id,
            CreditoMovimento.tipo == 'debito'
        )
        .scalar()
        or 0.0
    )

    saldo = float(total_creditos - total_debitos)

    cliente = Cliente.query.get(cliente_id)
    if cliente:
        cliente.saldo_atual = saldo
        db.session.add(cliente)
        db.session.commit()

    return Decimal(str(saldo)).quantize(Decimal("0.01"))


def registrar_credito(cliente_id: int, valor_bruto, desconto_tipo: str,
                      desconto_valor, motivo: str = "", criado_por: str = ""):
    """
    Cria um crédito, registra movimento 'credito' e recalcula saldo do cliente.

    No novo design:
      - desconto_tipo virá sempre como 'nenhum'
      - desconto_valor = 0
    """
    cli = Cliente.query.get(cliente_id)
    if not cli:
        raise ValueError("Cliente não encontrado")

    valor_final = calcular_valor_final(valor_bruto, desconto_tipo, desconto_valor)

    # saldo_antes vai ser o saldo recalculado pelos movimentos atuais
    saldo_antes = atualizar_saldo_credito_cliente(cli.id)

    c = Credito(
        cliente_id=cli.id,
        valor_bruto=float(_as_decimal(valor_bruto)),
        desconto_tipo=desconto_tipo or "nenhum",
        desconto_valor=float(_as_decimal(desconto_valor or 0)),
        valor_final=float(valor_final),
        motivo=motivo or "",
        saldo_antes=float(saldo_antes),
        criado_por=criado_por or "Supervisor"
    )
    db.session.add(c)
    db.session.flush()  # garante c.id

    # 👇 AQUI SIM: movimento de CRÉDITO correspondente a esse lançamento
    mov = CreditoMovimento(
        credito_id=c.id,
        cliente_id=cli.id,
        tipo="credito",
        valor=float(valor_final),
        referencia=f"Crédito #{c.id}",
    )
    db.session.add(mov)
    db.session.commit()

    # Recalcula saldo a partir de TODOS os movimentos (incluindo este crédito)
    novo_saldo = atualizar_saldo_credito_cliente(cli.id)

    c.saldo_depois = float(novo_saldo)
    db.session.add(c)
    db.session.commit()
    return c

def editar_credito(credito_id: int, valor_bruto, desconto_tipo: str,
                   desconto_valor, motivo: str = ""):
    """
    Ajusta um crédito EXISTENTE, atualiza o movimento de crédito correspondente
    e recalcula o saldo do cliente.
    """
    c = Credito.query.get_or_404(credito_id)
    cli = Cliente.query.get(c.cliente_id)
    if not cli:
        raise ValueError("Cliente não encontrado para esse crédito")

    valor_final = calcular_valor_final(valor_bruto, desconto_tipo, desconto_valor)

    c.valor_bruto = float(_as_decimal(valor_bruto))
    c.desconto_tipo = desconto_tipo or "nenhum"
    c.desconto_valor = float(_as_decimal(desconto_valor or 0))
    c.valor_final = float(valor_final)
    if motivo is not None:
        c.motivo = motivo

    # Atualiza o movimento principal desse crédito
    mov = (
        CreditoMovimento.query
        .filter_by(credito_id=c.id, tipo='credito')
        .order_by(CreditoMovimento.id.asc())
        .first()
    )
    if mov:
        mov.valor = float(valor_final)
        mov.referencia = f"Crédito #{c.id} (ajustado)"

    db.session.commit()

    # Recalcula saldo do cliente com base em TODOS os movimentos
    novo_saldo = atualizar_saldo_credito_cliente(cli.id)
    c.saldo_depois = float(novo_saldo)

    db.session.add(c)
    db.session.commit()
    return c


def consumir_credito_em_entrega(entrega_id: int, exigir_saldo_total: bool = True) -> Decimal:
    """
    Consome crédito na entrega.

    - Se exigir_saldo_total=True (default):
        * Só consome se o saldo do cliente cobrir TODO o valor que falta pagar.
        * Se o saldo for menor que o valor da entrega, NÃO consome nada
          e retorna Decimal("0.00") -> a rota deve pedir outra forma de pagamento.

    - Atualiza:
        * saldo_atual do cliente (via movimentos + recálculo)
        * entrega.credito_usado
        * cria CreditoMovimento tipo='debito'
        * marca status_pagamento='pago' se cobrir o valor total.
    """
    e = Entrega.query.get(entrega_id)
    if not e:
        return Decimal("0.00")

    cli = None
    # 1) tenta pelo cliente_id
    if getattr(e, "cliente_id", None):
        cli = Cliente.query.get(e.cliente_id)

    # 2) tenta pelo nome e JÁ VINCULA o cliente_id se achar
    if not cli:
        cli = _find_cliente_by_nome(e.cliente)
        if cli and not getattr(e, "cliente_id", None):
            e.cliente_id = cli.id  # garante vínculo
            db.session.add(e)

    if not cli:
        return Decimal("0.00")

    valor = _as_decimal(e.valor or 0)
    usado_antes = _as_decimal(e.credito_usado or 0)
    faltante = valor - usado_antes
    if faltante <= 0:
        return Decimal("0.00")

    # saldo atual sempre recalculado pelos movimentos
    saldo_atual = atualizar_saldo_credito_cliente(cli.id)
    saldo = _as_decimal(saldo_atual)

    # Se exigimos saldo total e o saldo é menor que o valor faltante,
    # NÃO consome nada. A rota deve tratar isso como "crédito insuficiente".
    if exigir_saldo_total and saldo < faltante:
        return Decimal("0.00")

    consumir_val = min(saldo, faltante)
    if consumir_val <= 0:
        return Decimal("0.00")

    novo_usado = usado_antes + consumir_val
    e.credito_usado = float(novo_usado)

    mov = CreditoMovimento(
      cliente_id=cli.id,
      tipo="debito",
      valor=float(consumir_val),
      referencia=f"Entrega #{e.id}",
      entrega_id=e.id,   # 👈 AQUI SIM: vínculo da movimentação com a entrega
    )
    db.session.add(mov)
    db.session.commit()

    # Atualiza saldo do cliente DEPOIS do débito
    atualizar_saldo_credito_cliente(cli.id)

    if novo_usado >= valor:
        e.status_pagamento = "pago"
        if not (e.pagamento or "").strip():
            e.pagamento = "Crédito"
        if not (e.recebido_por or "").strip():
            e.recebido_por = "Crédito automático"
    else:
        if not (e.status_pagamento or "").strip():
            e.status_pagamento = "pendente"

    db.session.add(e)
    db.session.commit()
    return consumir_val


def desfazer_consumo_credito_da_entrega(entrega_id: int) -> Decimal:
    """
    Estorna TODO crédito usado nesta entrega, devolvendo para o saldo do cliente
    e zerando entrega.credito_usado.
    NÃO mexe em pagamento/status_pagamento.
    """
    e = Entrega.query.get(entrega_id)
    if not e:
        return Decimal("0.00")

    usado = _as_decimal(e.credito_usado or 0)
    if usado <= 0:
        return Decimal("0.00")

    cli = None
    if getattr(e, "cliente_id", None):
        cli = Cliente.query.get(e.cliente_id)
    if not cli:
        cli = _find_cliente_by_nome(e.cliente)
    if not cli:
        return Decimal("0.00")

    mov_estorno = CreditoMovimento(
        cliente_id=cli.id,
        tipo="credito",
        valor=float(usado),
        referencia=f"Estorno Entrega #{e.id}",
    )
    db.session.add(mov_estorno)

    e.credito_usado = 0.0
    db.session.commit()

    # Recalcula saldo com base em TODOS os movimentos
    atualizar_saldo_credito_cliente(cli.id)

    return usado


def consumo_total_do_credito(credito_id: int) -> float:
    """
    Mantido por compatibilidade. Se você quiser, pode ignorar essa função
    e sempre olhar apenas os movimentos por cliente.
    """
    total = (
        db.session.query(func.sum(CreditoMovimento.valor))
        .filter(
            CreditoMovimento.credito_id == credito_id,
            CreditoMovimento.tipo == "debito",
        )
        .scalar()
        or 0.0
    )
    return float(total or 0.0)


# Constantes semânticas usadas nas rotas de crédito
TIPO_ENTRADA = 'ENTRADA'
TIPO_CONSUMO = 'CONSUMO'
TIPO_AJUSTE = 'AJUSTE'


def calc_valor_final(valor, desconto_tipo, desconto_valor):
    """Wrapper com nome antigo usado em algumas rotas."""
    return float(calcular_valor_final(valor, desconto_tipo, desconto_valor))


def atualizar_saldo_cliente(cliente_id, delta):
    """
    Função LEGADA. Hoje o saldo oficial é calculado por atualizar_saldo_credito_cliente.
    Se ainda tiver uso em algum lugar antigo, ela só ajusta o saldo_atual direto.
    """
    cli = Cliente.query.get(cliente_id)
    if not cli:
        return
    cli.saldo_atual = float(_as_decimal(cli.saldo_atual) + _as_decimal(delta))
    db.session.add(cli)


def registrar_movimento(cliente_id, tipo, valor,
                        referencia='',
                        credito_id=None,
                        entrega_id=None):
    """
    Também legado. Hoje o normal é:
      - criar movimentos diretamente nas funções novas
      - depois chamar atualizar_saldo_credito_cliente(cliente_id)

    Agora também aceita entrega_id para vincular o movimento a uma entrega.
    """
    tipo_up = (tipo or '').upper()
    if tipo_up in (TIPO_ENTRADA, TIPO_AJUSTE, 'CREDITO'):
        tm = 'credito'
    elif tipo_up in (TIPO_CONSUMO, 'DEBITO', 'DÉBITO'):
        tm = 'debito'
    else:
        tm = 'credito'

    mov = CreditoMovimento(
        cliente_id=cliente_id,
        tipo=tm,
        valor=float(_as_decimal(valor)),
        referencia=(referencia or '')[:120],
        credito_id=credito_id,
        entrega_id=entrega_id,
    )
    db.session.add(mov)
    return mov


def _delta_saldo_tipo_mov(tipo_raw, valor) -> float:
    t = (tipo_raw or '').upper()
    v = float(valor or 0)
    if t in (TIPO_ENTRADA, TIPO_AJUSTE, 'CREDITO'):
        return v
    if t in (TIPO_CONSUMO, 'DEBITO', 'DÉBITO'):
        return -v
    return 0.0


def br_date_ymd(dt_utc_naive: datetime) -> str:
    if not dt_utc_naive:
        return ''
    return to_brasilia(dt_utc_naive).date().isoformat()


# =========================================================
# FERIADOS / PERÍODO LEGÍVEL
# =========================================================
MUNICIPAIS_NATAL = {(11, 21): "Nossa Senhora da Apresentação (Municipal - Natal/RN)"}


def verifica_feriado(data_ref=None):
    if data_ref is None:
        data_ref = datetime.now(BRAZIL_TZ).date()
    feriados_nac = holidays.Brazil(years=data_ref.year)
    feriados_est = holidays.Brazil(state='RN', years=data_ref.year)
    nomes = []
    if data_ref in feriados_nac:
        nomes.append(f"Feriado Nacional – {feriados_nac.get(data_ref)}")
    if data_ref in feriados_est and feriados_est.get(data_ref) != feriados_nac.get(data_ref):
        nomes.append(f"Feriado Estadual (RN) – {feriados_est.get(data_ref)}")
    if (data_ref.month, data_ref.day) in MUNICIPAIS_NATAL:
        nomes.append(f"Feriado Municipal (Natal/RN) – {MUNICIPAIS_NATAL[(data_ref.month, data_ref.day)]}")
    return " | ".join(nomes) if nomes else None


def periodo_legivel_str(di_str, df_str):
    if di_str and df_str:
        di = datetime.strptime(di_str, "%Y-%m-%d").strftime("%d/%m/%Y")
        df = datetime.strptime(df_str, "%Y-%m-%d").strftime("%d/%m/%Y")
        return f"{di} a {df}"
    if di_str:
        di = datetime.strptime(di_str, "%Y-%m-%d").strftime("%d/%m/%Y")
        return f"desde {di}"
    if df_str:
        df = datetime.strptime(df_str, "%Y-%m-%d").strftime("%d/%m/%Y")
        return f"até {df}"
    return "todo o período"


def _now_brt():
    try:
        return datetime.now(BRAZIL_TZ)
    except Exception:
        return datetime.now()

# =========================================================
# CONFIG KV / PER_KM / PREÇO DE ROTAS
# =========================================================
ParametroSistemaCls = globals().get("ParametroSistema", None)

if ParametroSistemaCls is None:
    class ConfigKV(db.Model):
        __tablename__ = "config_kv"
        id = db.Column(db.Integer, primary_key=True)
        chave = db.Column(db.String(80), unique=True, nullable=False, index=True)
        valor = db.Column(db.String(255), nullable=True)

        def __repr__(self):
            return f"<ConfigKV {self.chave}={self.valor}>"

    def _get_param(chave: str, default=None):
        row = ConfigKV.query.filter_by(chave=chave).first()
        return row.valor if row and row.valor is not None else default

    def _set_param(chave: str, valor: str):
        row = ConfigKV.query.filter_by(chave=chave).first()
        if not row:
            row = ConfigKV(chave=chave, valor=valor)
            db.session.add(row)
        else:
            row.valor = valor
        db.session.commit()

else:
    def _get_param(chave: str, default=None):
        row = ParametroSistema.query.filter_by(chave=chave).first()
        return row.valor if row and row.valor is not None else default

    def _set_param(chave: str, valor: str):
        row = ParametroSistema.query.filter_by(chave=chave).first()
        if not row:
            row = ParametroSistema(chave=chave, valor=str(valor))
            db.session.add(row)
        else:
            row.valor = str(valor)
        db.session.commit()


def get_per_km():
    # ordem: DB -> ENV -> 3.00
    v = _get_param("per_km", None)
    if v is not None:
        try:
            return float(v)
        except Exception:
            pass
    try:
        return float(os.getenv("PER_KM", "3.00"))
    except Exception:
        return 3.00


def set_per_km(novo_valor: float):
    _set_param("per_km", f"{float(novo_valor):.2f}")
    return get_per_km()

def get_pix_chave():
    return _get_param("pix_chave", "") or ""


class PrecoRota(db.Model):
    __tablename__ = "preco_rota"
    id = db.Column(db.Integer, primary_key=True)
    origem = db.Column(db.String(120), nullable=False, index=True)
    destino = db.Column(db.String(120), nullable=False, index=True)
    valor = db.Column(db.Numeric(10, 2), nullable=False, default=0)
    criado_em = db.Column(db.DateTime, default=_now_brt)
    atualizado_em = db.Column(db.DateTime, default=_now_brt, onupdate=_now_brt)

    __table_args__ = (
        db.UniqueConstraint("origem", "destino", name="uq_preco_rota_pair"),
    )

    def to_dict(self):
        return {
            "id": self.id,
            "origem": self.origem,
            "destino": self.destino,
            "valor": float(self.valor),
        }

# =========================================================
# HELPERS GENÉRICOS / SEGURANÇA / REDIRECT
# =========================================================
def _norm(s: str) -> str:
    return (s or "").strip()


def _ci_equal(a: str, b: str) -> bool:
    return (_norm(a).casefold() == _norm(b).casefold())


@app.before_request
def remember_admin_filters():
    if request.endpoint == "admin" and request.method == "GET":
        keys = ["cooperado_id", "data_inicio", "data_fim", "status_pagamento", "cliente"]
        session["last_filters"] = {k: request.args.get(k) for k in keys if request.args.get(k)}


def _build_admin_url_from_referrer():
    ref = request.headers.get("Referer") or ""
    try:
        p = urlparse(ref)
        if not p.path.endswith("/admin"):
            return None
        qs = parse_qs(p.query)
        params = {k: v[0] for k, v in qs.items() if v}
        return url_for("admin", **params)
    except Exception:
        return None


def redirect_back_to_admin():
    next_url = request.args.get("next") or request.form.get("next")
    if next_url:
        return redirect(next_url)
    from_ref = _build_admin_url_from_referrer()
    if from_ref:
        return redirect(from_ref)
    params = session.get("last_filters") or {}
    return redirect(url_for("admin", **params))


def _assert_entrega_do_cooperado(entrega: 'Entrega'):
    uid = session.get('user_id')
    if uid is None or session.get('is_admin'):
        abort(403)
    if entrega.cooperado_id != uid:
        abort(403)


def master_required(view_func):
    @wraps(view_func)
    def _wrapped(*args, **kwargs):
        if not session.get('is_admin'):
            return redirect(url_for('login'))
        if not session.get('is_master'):
            flash('Acesso restrito ao admin master.')
            return redirect(url_for('admin'))
        return view_func(*args, **kwargs)
    return _wrapped


def render_or_string(template_name, fallback_html, **ctx):
    try:
        return render_template(template_name, **ctx)
    except TemplateNotFound:
        return render_template_string(fallback_html, **ctx)

# =========================================================
# ROTA INTRUSO (ARAPUCA)
# =========================================================
@app.route('/intruso')
def intruso():
    ip = request.headers.get('X-Forwarded-For', request.remote_addr)
    user_agent = request.headers.get('User-Agent', 'Desconhecido')
    username = request.args.get('u')

    agora_brasil = datetime.now(BRAZIL_TZ)
    acesso_data = agora_brasil.strftime('%d/%m/%Y %H:%M:%S')
    registro_id = agora_brasil.strftime('%Y%m%d%H%M%S')

    return render_template(
        'intruso.html',
        ip=ip,
        user_agent=user_agent,
        username=username,
        acesso_data=acesso_data,
        registro_id=registro_id
    )

# =========================================================
# LOGIN ADMIN / COOPERADO / CLIENTE
# =========================================================
@app.route('/', methods=['GET', 'POST'])
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        usuario = (request.form.get('usuario') or '').strip()
        senha = request.form.get('senha') or ''
        next_url = request.form.get('next') or ''
        user_lc = usuario.lower()

        # ARMADILHA: usuario=coopex / senha=05062721 -> manda pro /intruso
        if user_lc == 'coopex' and senha == '05062721':
            return redirect(url_for('intruso', u=usuario))

        # 1) Admin fixo
        if user_lc in ADMIN_CREDENTIALS:
            cred_map = ADMIN_CREDENTIALS[user_lc]
            if senha in cred_map:
                session.clear()
                session['user_id'] = 0
                session['user_nome'] = usuario
                session['is_admin'] = True
                session['is_master'] = bool(cred_map[senha].get('is_master'))
                return redirect(url_for('admin'))
            else:
                flash('Usuário ou senha incorretos.', 'error')
                try:
                    return render_template('login.html', now=lambda: datetime.now(BRAZIL_TZ))
                except TemplateNotFound:
                    pass

        # 2) Cooperado (login pelo nome)
        cooperado = Cooperado.query.filter(func.lower(Cooperado.nome) == user_lc).first()
        if cooperado and cooperado.check_senha(senha):
            if not getattr(cooperado, 'ativo', True):
                flash('Usuário inativo. Fale com o administrador.', 'error')
                try:
                    return render_template('login.html', now=lambda: datetime.now(BRAZIL_TZ))
                except TemplateNotFound:
                    pass

            session.clear()
            session['user_id'] = cooperado.id
            session['user_nome'] = cooperado.nome
            session['is_admin'] = False
            session['is_master'] = False
            session['tipo'] = 'cooperado'   # 👈 ESSENCIAL PARA /cooperado/atualizar_localizacao
            return redirect(url_for('painel_cooperado'))

        # 3) Cliente (login por username OU e-mail)
        cli = (
            Cliente.query.filter(func.lower(Cliente.username) == user_lc).first()
            or Cliente.query.filter(func.lower(Cliente.email) == user_lc).first()
        )
        if cli and cli.check_senha(senha):
            session.clear()
            session['cliente_id'] = cli.id
            session['cliente_username'] = cli.username
            session['cliente_nome'] = cli.nome
            session['is_cliente'] = True
            if next_url:
                return redirect(next_url)
            return redirect(url_for('meu_credito'))

        # nenhuma combinação deu certo
        flash('Usuário ou senha incorretos.', 'error')

    # GET ou erro: mostra tela de login bonita
    try:
        return render_template('login.html', now=lambda: datetime.now(BRAZIL_TZ))
    except TemplateNotFound:
        # fallback simples
        return render_template_string("""
        <h2>Login (Admin/Cooperado/Cliente)</h2>
        <form method="post">
          <div><label>Usuário ou e-mail</label><input name="usuario"></div>
          <div><label>Senha</label><input name="senha" type="password"></div>
          <button type="submit">Entrar</button>
        </form>
        """, now=lambda: datetime.now(BRAZIL_TZ))

@app.post('/api/mobile/login_cooperado')
def api_mobile_login_cooperado():
    """
    Login específico para o APP NATIVO do cooperado.

    Espera JSON:
    {
      "usuario": "nome do cooperado (mesmo do painel)",
      "senha": "1234"
    }

    Responde JSON:
    {
      "ok": true/false,
      "msg": "...",
      "cooperado": {...}  # se ok
    }
    """
    data = request.get_json(silent=True) or {}
    usuario = (data.get('usuario') or '').strip().lower()
    senha = data.get('senha') or ''

    # mesmo critério do login web: cooperado loga pelo NOME
    coop = Cooperado.query.filter(func.lower(Cooperado.nome) == usuario).first()
    if not coop or not coop.check_senha(senha):
        return jsonify(ok=False, msg='Usuário ou senha inválidos'), 401

    if not getattr(coop, 'ativo', True):
        return jsonify(ok=False, msg='Usuário inativo. Fale com a supervisão.'), 403

    # usa a mesma sessão do site (cookie), assim o app pode reaproveitar
    session.clear()
    session['user_id'] = coop.id
    session['user_nome'] = coop.nome
    session['is_admin'] = False
    session['is_master'] = False
    session['tipo'] = 'cooperado'

    return jsonify(
        ok=True,
        msg='Login efetuado com sucesso.',
        cooperado={
            "id": coop.id,
            "nome": coop.nome,
            "ativo": bool(coop.ativo),
        }
    )


@app.route('/logout')
def logout():
    # se for cooperado logado, marca offline
    uid = session.get('user_id')
    is_admin = session.get('is_admin')

    if uid and not is_admin:
        coop = Cooperado.query.get(uid)
        if coop:
            coop.online = False
            db.session.commit()

    session.clear()
    return redirect(url_for('login'))


# =========================================================
# CLIENTE: LOGIN / PRIMEIRO ACESSO / MEU CRÉDITO
# =========================================================
def _norm_phone(s: str) -> str:
    if s is None:
        return ""
    digits = re.sub(r'\D+', '', str(s))
    if digits.startswith('55'):
        digits = digits[2:]
    if len(digits) > 11:
        digits = digits[-11:]
    return digits


@app.route('/cliente/primeiro_acesso', methods=['GET', 'POST'])
def cliente_primeiro_acesso():
    # GET -> volta para tela de login já abrindo o painel de cadastro
    if request.method == 'GET':
        return redirect(url_for('login', signup=1))

    # POST (form do card de primeiro acesso)
    nome = (request.form.get('nome') or '').strip()
    username = (request.form.get('usuario') or '').strip()
    email = (request.form.get('email') or '').strip().lower()
    telefone = _norm_phone(request.form.get('telefone') or '')
    senha = request.form.get('senha') or ''
    senha_conf = request.form.get('senha_conf') or ''
    next_url = request.form.get('next') or url_for('meu_credito')

    # validações básicas
    if not nome or not username or not email or not telefone or not senha:
        flash('Preencha todos os campos obrigatórios.', 'error')
        return redirect(url_for('login', signup=1))

    if senha != senha_conf:
        flash('As senhas não conferem.', 'error')
        return redirect(url_for('login', signup=1))

    # usuário único
    if Cliente.query.filter(func.lower(Cliente.username) == username.lower()).first():
        flash('Nome de usuário já existe. Escolha outro.', 'error')
        return redirect(url_for('login', signup=1))

    # e-mail único
    if email and Cliente.query.filter(func.lower(Cliente.email) == email.lower()).first():
        flash('Já existe um cadastro com este e-mail.', 'error')
        return redirect(url_for('login', signup=1))

    # tentar reaproveitar cliente existente pelo telefone ou nome
    cli = None
    if telefone:
        cli = Cliente.query.filter(Cliente.telefone == telefone).first()
    if not cli and nome:
        cli = Cliente.query.filter(func.lower(Cliente.nome) == nome.lower()).first()

    if not cli:
        cli = Cliente(
            nome=nome,
            telefone=telefone,
            email=email,
            saldo_atual=0.0
        )
        db.session.add(cli)
        db.session.flush()
    else:
        cli.nome = nome or cli.nome
        cli.telefone = telefone or cli.telefone
        cli.email = email or cli.email

    cli.username = username
    cli.set_senha(senha)

    db.session.commit()

    # loga automaticamente
    session.clear()
    session['cliente_id'] = cli.id
    session['cliente_username'] = cli.username
    session['cliente_nome'] = cli.nome
    session['is_cliente'] = True

    flash('Conta criada com sucesso! Você já está logado.', 'ok')
    return redirect(next_url)

@app.route('/cliente/esqueci-senha', methods=['GET', 'POST'])
def cliente_esqueci_senha():
    if request.method == 'POST':
        usuario_email = (request.form.get('usuario_email') or '').strip()
        telefone_raw = request.form.get('telefone') or ''
        telefone = _norm_phone(telefone_raw)

        if not usuario_email and not telefone:
            flash('Informe usuário/e-mail ou telefone.', 'error')
            return redirect(url_for('cliente_esqueci_senha'))

        # tenta localizar cliente
        cli = None
        if usuario_email:
            u_lc = usuario_email.lower()
            cli = (Cliente.query.filter(func.lower(Cliente.username) == u_lc).first()
                   or Cliente.query.filter(func.lower(Cliente.email) == u_lc).first())
        if not cli and telefone:
            cli = Cliente.query.filter(Cliente.telefone == telefone).first()

        if not cli:
            flash('Nenhum cliente encontrado com esses dados.', 'error')
            return redirect(url_for('cliente_esqueci_senha'))

        # gera código de 6 dígitos
        code = f"{random.randint(0, 999999):06d}"
        cli.reset_code = code
        cli.reset_expires_at = datetime.utcnow() + timedelta(minutes=15)
        db.session.commit()

        # Aqui você integraria com e-mail/SMS real.
        # Por enquanto mostramos na tela (modo teste).
        flash(f'Enviamos um código de 6 dígitos para seu contato. (Código de teste: {code})', 'ok')
        return redirect(url_for('cliente_reset_senha', cliente_id=cli.id))

    # GET
    return render_or_string("cliente_esqueci_senha.html", """
    <!doctype html><html lang="pt-BR"><head>
    <meta charset="utf-8"><title>Esqueci minha senha</title>
    <meta name="viewport" content="width=device-width,initial-scale=1">
    </head><body style="font-family:system-ui;max-width:480px;margin:30px auto;">
      <h2>Esqueci minha senha (Cliente)</h2>
      {% with messages = get_flashed_messages(with_categories=true) %}
        {% if messages %}
          {% for cat, msg in messages %}
            <div style="margin:8px 0;padding:8px;border-radius:6px;
                  background:{{ '#ffe8ea' if cat=='error' else '#eafff2' }};
                  border:1px solid {{ '#ffccd2' if cat=='error' else '#c9f2da' }};">
              {{ msg }}
            </div>
          {% endfor %}
        {% endif %}
      {% endwith %}
      <p>Informe seu usuário/e-mail ou telefone cadastrado para receber um código de redefinição.</p>
      <form method="post">
        <div style="margin-bottom:8px">
          <label>Usuário ou e-mail</label><br>
          <input name="usuario_email" style="width:100%;padding:6px">
        </div>
        <div style="margin-bottom:8px">
          <label>Telefone (opcional)</label><br>
          <input name="telefone" style="width:100%;padding:6px">
        </div>
        <button type="submit" style="padding:8px 14px">Enviar código</button>
      </form>
      <p style="margin-top:10px">
        <a href="{{ url_for('login') }}">Voltar ao login</a>
      </p>
    </body></html>
    """)

@app.route('/cliente/reset-senha/<int:cliente_id>', methods=['GET', 'POST'])
def cliente_reset_senha(cliente_id):
    cli = Cliente.query.get_or_404(cliente_id)

    if request.method == 'POST':
        code = (request.form.get('codigo') or '').strip()
        nova = request.form.get('senha') or ''
        conf = request.form.get('senha_conf') or ''

        if not code or not nova:
            flash('Informe o código e a nova senha.', 'error')
            return redirect(url_for('cliente_reset_senha', cliente_id=cliente_id))

        if nova != conf:
            flash('As senhas não conferem.', 'error')
            return redirect(url_for('cliente_reset_senha', cliente_id=cliente_id))

        # valida código e validade
        agora = datetime.utcnow()
        if not cli.reset_code or cli.reset_code != code:
            flash('Código inválido.', 'error')
            return redirect(url_for('cliente_reset_senha', cliente_id=cliente_id))

        if cli.reset_expires_at and cli.reset_expires_at < agora:
            flash('Código expirado, faça uma nova solicitação.', 'error')
            cli.reset_code = None
            cli.reset_expires_at = None
            db.session.commit()
            return redirect(url_for('cliente_esqueci_senha'))

        # ok: troca senha
        cli.set_senha(nova)
        cli.reset_code = None
        cli.reset_expires_at = None
        db.session.commit()

        flash('Senha alterada com sucesso! Agora faça login novamente.', 'ok')
        return redirect(url_for('login'))

    # GET – formulário simples
    return render_or_string("cliente_reset_senha.html", """
    <!doctype html><html lang="pt-BR"><head>
    <meta charset="utf-8"><title>Redefinir senha</title>
    <meta name="viewport" content="width=device-width,initial-scale=1">
    </head><body style="font-family:system-ui;max-width:480px;margin:30px auto;">
      <h2>Redefinir senha — {{ cli.nome }}</h2>
      {% with messages = get_flashed_messages(with_categories=true) %}
        {% if messages %}
          {% for cat, msg in messages %}
            <div style="margin:8px 0;padding:8px;border-radius:6px;
                  background:{{ '#ffe8ea' if cat=='error' else '#eafff2' }};
                  border:1px solid {{ '#ffccd2' if cat=='error' else '#c9f2da' }};">
              {{ msg }}
            </div>
          {% endfor %}
        {% endif %}
      {% endwith %}
      <p>Digite o código recebido e crie uma nova senha.</p>
      <form method="post">
        <div style="margin-bottom:8px">
          <label>Código</label><br>
          <input name="codigo" style="width:100%;padding:6px">
        </div>
        <div style="margin-bottom:8px">
          <label>Nova senha</label><br>
          <input type="password" name="senha" style="width:100%;padding:6px">
        </div>
        <div style="margin-bottom:8px">
          <label>Confirmar senha</label><br>
          <input type="password" name="senha_conf" style="width:100%;padding:6px">
        </div>
        <button type="submit" style="padding:8px 14px">Salvar nova senha</button>
      </form>
      <p style="margin-top:10px">
        <a href="{{ url_for('login') }}">Voltar ao login</a>
      </p>
    </body></html>
    """, cli=cli)

@app.route('/cliente/login', methods=['GET', 'POST'])
def cliente_login():
    if request.method == 'POST':
        username = (request.form.get('username') or '').strip()
        senha = request.form.get('senha') or ''
        if not username or not senha:
            flash('Informe usuário e senha.')
            return redirect(url_for('cliente_login'))

        cli = Cliente.query.filter(func.lower(Cliente.username) == username.lower()).first()
        if not cli or not cli.check_senha(senha):
            flash('Usuário ou senha inválidos.')
            return redirect(url_for('cliente_login'))

        session['cliente_id'] = cli.id
        session['cliente_username'] = cli.username
        session['cliente_nome'] = cli.nome
        session['is_cliente'] = True
        return redirect(url_for('meu_credito'))

    return render_or_string("cliente_login.html", """
    <h2>Login do Cliente</h2>
    <form method="post">
      <div><label>Usuário</label><input name="username" required></div>
      <div><label>Senha</label><input type="password" name="senha" required></div>
      <button type="submit">Entrar</button>
    </form>
    <p>Novo por aqui? <a href="{{ url_for('cliente_primeiro_acesso') }}">Primeiro acesso</a></p>
    """)


@app.route('/cliente/logout')
def cliente_logout():
    for k in ['cliente_id', 'cliente_username', 'cliente_nome', 'is_cliente']:
        session.pop(k, None)
    flash('Você saiu da área do cliente.')
    # volta para o login principal (admin / cooperado / cliente)
    return redirect(url_for('login'))


def cliente_required(view_func):
    @wraps(view_func)
    def _wrap(*a, **kw):
        if not session.get('is_cliente') or not session.get('cliente_id'):
            return redirect(url_for('cliente_login'))
        return view_func(*a, **kw)
    return _wrap


@app.route('/meu-credito')
@cliente_required
def meu_credito():
    cid = session['cliente_id']
    cli = Cliente.query.get_or_404(cid)

    # Movimentações de crédito do cliente
    movs = (
        CreditoMovimento.query
        .filter(CreditoMovimento.cliente_id == cid)
        .order_by(CreditoMovimento.id.desc())
        .all()
    )

    # Últimas entregas do cliente (para "comprovante")
    entregas = (
        Entrega.query
        .filter(Entrega.cliente_id == cid)
        .order_by(Entrega.data_envio.desc())
        .limit(20)
        .all()
    )

    # Bairros disponíveis (a partir da tabela de preços)
    rotas = PrecoRota.query.all()
    bairros = sorted({
        _norm(r.origem) for r in rotas if _norm(r.origem)
    } | {
        _norm(r.destino) for r in rotas if _norm(r.destino)
    })

    pix_chave = get_pix_chave()

    return render_template(
        "meu_credito.html",
        cli=cli,
        movs=movs,
        entregas=entregas,
        bairros=bairros,
        pix_chave=pix_chave,
        to_brasilia=to_brasilia
    )


# =========================================================
# 6) APIS JSON PARA COTAÇÃO E PEDIDO DE ENTREGA DO CLIENTE
# =========================================================

def _cliente_atual():
    """Obtém o cliente logado a partir da sessão."""
    cid = session.get('cliente_id')
    if not cid:
        abort(401)
    cli = Cliente.query.get(cid)
    if not cli:
        abort(401)
    return cli


def _calcular_preco_bairros(bairro_origem, bairro_destino):
    """
    Calcula o preço da rota usando a tabela PrecoRota.
    Usa origem/destino normalizados (_norm).
    """
    if not bairro_origem or not bairro_destino:
        raise ValueError('Bairros de coleta e entrega são obrigatórios.')

    bo = _norm(bairro_origem)
    bd = _norm(bairro_destino)

    rota = (
        PrecoRota.query
        .filter(func.lower(PrecoRota.origem) == bo.lower(),
                func.lower(PrecoRota.destino) == bd.lower())
        .first()
    )

    if not rota:
        raise ValueError('Não existe preço configurado para essa rota.')

    # Tenta descobrir o campo de valor na tabela
    preco = None
    for campo in ('valor', 'preco', 'preco_total'):
        if hasattr(rota, campo):
            preco = getattr(rota, campo)
            break

    if preco is None:
        raise RuntimeError('Campo de preço não encontrado na tabela de preços.')

    return float(preco)


@app.route('/api/cliente/cotar-entrega', methods=['POST'])
@cliente_required
def api_cliente_cotar_entrega():
    """
    Recebe JSON com:
    {
      "coleta":  {"bairro": "...", "endereco": "...", ...},
      "entrega": {"bairro": "...", "endereco": "...", ...}
    }
    e devolve o preço e meios de pagamento disponíveis.
    """
    cli = _cliente_atual()
    data = request.get_json(silent=True) or {}

    coleta = data.get('coleta') or {}
    entrega = data.get('entrega') or {}

    bairro_coleta = coleta.get('bairro') or coleta.get('bairro_origem')
    bairro_entrega = entrega.get('bairro') or entrega.get('bairro_destino')

    try:
        preco = _calcular_preco_bairros(bairro_coleta, bairro_entrega)
    except Exception as e:
        return jsonify({'ok': False, 'erro': str(e)}), 400

    meios = []
    if cli.saldo_atual is not None and cli.saldo_atual >= preco:
        meios.append('CREDITO')  # prioridade se tiver saldo
    meios.extend(['PIX', 'DINHEIRO'])

    return jsonify({
        'ok': True,
        'preco': preco,
        'moeda': 'BRL',
        'cliente_saldo_atual': float(cli.saldo_atual or 0),
        'pode_usar_credito': 'CREDITO' in meios,
        'meios_pagamento': meios,
    })


@app.route('/api/cliente/solicitar-entrega', methods=['POST'])
@cliente_required
def api_cliente_solicitar_entrega():
    """
    Cliente faz o pedido de entrega.

    JSON esperado:
    {
      "coleta":  {...},
      "entrega": {...},
      "paradas": ["Rua X - Bairro Y", ...],     # opcional
      "meio_pagamento": "CREDITO" | "PIX" | "DINHEIRO",
      "apenas_simular": false
    }

    Regras:
    - Recalcula o preço pela tabela PrecoRota (coleta.bairro -> entrega.bairro).
    - Se apenas_simular = true, NÃO cria entrega, só devolve preço e formas de pgto.
    - Se meio_pagamento == CREDITO e saldo < preço => erro 400 (cliente escolhe outra forma).
    - Se meio_pagamento == CREDITO e saldo suficiente => cria entrega + consome crédito
      usando consumir_credito_em_entrega (vinculado à entrega).
    """
    cli = _cliente_atual()
    data = request.get_json(silent=True) or {}

    coleta = data.get('coleta') or {}
    entrega_dest = data.get('entrega') or {}
    paradas_lista = data.get('paradas') or []

    meio_pagamento = (data.get('meio_pagamento') or '').upper()
    apenas_simular = bool(data.get('apenas_simular'))

    bairro_coleta = coleta.get('bairro') or coleta.get('bairro_origem')
    bairro_entrega = entrega_dest.get('bairro') or entrega_dest.get('bairro_destino')

    # 1) Calcula preço
    try:
        preco = _calcular_preco_bairros(bairro_coleta, bairro_entrega)
    except Exception as e:
        return jsonify({'ok': False, 'erro': str(e)}), 400

    # 2) Simulação apenas (não cria nada no banco)
    if apenas_simular:
        meios = []
        if cli.saldo_atual is not None and cli.saldo_atual >= preco:
            meios.append('CREDITO')
        meios.extend(['PIX', 'DINHEIRO'])

        return jsonify({
            'ok': True,
            'simulacao': True,
            'preco': preco,
            'cliente_saldo_atual': float(cli.saldo_atual or 0),
            'meios_pagamento': meios,
        })

    # 3) Define meio de pagamento padrão
    if meio_pagamento not in ('CREDITO', 'PIX', 'DINHEIRO'):
        if cli.saldo_atual is not None and cli.saldo_atual >= preco:
            meio_pagamento = 'CREDITO'
        else:
            meio_pagamento = 'PIX'

    # 4) Se for CREDITO, exige saldo total
    if meio_pagamento == 'CREDITO':
        if cli.saldo_atual is None or cli.saldo_atual < preco:
            return jsonify({
                'ok': False,
                'erro': 'Crédito insuficiente para essa entrega. Escolha outra forma de pagamento.'
            }), 400

    try:
        # Sempre salva data_envio como UTC naive (padrão do sistema)
        data_envio_utc = datetime.utcnow()

        # JSONs com origem / destino / paradas
        origem_json_dict = {
            "endereco": coleta.get('endereco'),
            "bairro": bairro_coleta,
            "ref": coleta.get('referencia') or coleta.get('ref'),
            "lat": coleta.get('lat'),
            "lng": coleta.get('lng'),
        }
        destino_json_dict = {
            "endereco": entrega_dest.get('endereco'),
            "bairro": bairro_entrega,
            "ref": entrega_dest.get('referencia') or entrega_dest.get('ref'),
            "lat": entrega_dest.get('lat'),
            "lng": entrega_dest.get('lng'),
        }
        paradas_json_dict = {
            "stops": paradas_lista
        }

        # Campos da Entrega compatíveis com o seu model atual
        campos = {
            'cliente_id': cli.id,
            'cliente': cli.nome,          # texto para o admin enxergar
            'bairro': bairro_entrega,     # você só tem 1 campo de bairro na Entrega
            'valor': preco,
            'data_envio': data_envio_utc,
            'status': 'pendente',
            'status_pagamento': 'pago' if meio_pagamento == 'CREDITO' else 'pendente',
            'pagamento': meio_pagamento.capitalize(),  # "Credito", "Pix", "Dinheiro"
            'origem_json': json.dumps(origem_json_dict, ensure_ascii=False),
            'destino_json': json.dumps(destino_json_dict, ensure_ascii=False),
            'paradas_json': json.dumps(paradas_json_dict, ensure_ascii=False),
            # status_corrida fica com default 'pendente'
        }

        entrega_obj = Entrega(**campos)
        db.session.add(entrega_obj)
        db.session.flush()  # garante entrega_obj.id

        # 5) Se pagamento for CREDITO, consome o crédito de forma oficial
        if meio_pagamento == 'CREDITO':
            # Aqui usamos sua função nova, que:
            # - cria CreditoMovimento debito
            # - atualiza saldo do cliente via atualizar_saldo_credito_cliente
            # - preenche entrega.credito_usado, status_pagamento, pagamento, etc.
            valor_consumido = consumir_credito_em_entrega(entrega_obj.id, exigir_saldo_total=True)
            if valor_consumido <= 0:
                # Se por algum motivo não conseguiu consumir, aborta com erro
                db.session.rollback()
                return jsonify({
                    'ok': False,
                    'erro': 'Falha ao consumir crédito. Tente novamente ou escolha outra forma de pagamento.'
                }), 500

        db.session.commit()

    except Exception as e:
        db.session.rollback()
        current_app.logger.exception('Erro ao solicitar entrega')
        return jsonify({
            'ok': False,
            'erro': f'Erro ao solicitar entrega: {e.__class__.__name__}'
        }), 500

    return jsonify({
        'ok': True,
        'entrega_id': entrega_obj.id,
        'preco': preco,
        'meio_pagamento': meio_pagamento,
        'status_pagamento': entrega_obj.status_pagamento,
        'comprovante_url': url_for('cliente_comprovante', entrega_id=entrega_obj.id),
    })



# =========================================================
# 7) COMPROVANTE DA ENTREGA PARA O CLIENTE
# =========================================================

@app.route('/cliente/comprovante/<int:entrega_id>')
@cliente_required
def cliente_comprovante(entrega_id):
    """
    Mostra o comprovante de uma entrega específica para o cliente.
    Só deixa ver se a entrega for do próprio cliente.
    """
    cli = _cliente_atual()

    entrega = (
        Entrega.query
        .filter(Entrega.id == entrega_id, Entrega.cliente_id == cli.id)
        .first_or_404()
    )

    movs = (
        CreditoMovimento.query
        .filter(
            CreditoMovimento.cliente_id == cli.id,
            CreditoMovimento.entrega_id == entrega.id
        )
        .order_by(CreditoMovimento.id.desc())
        .all()
    )

    return render_template(
        'cliente_comprovante.html',
        cli=cli,
        entrega=entrega,
        movs=movs,
        to_brasilia=to_brasilia
    )

# =========================================================
# RASTREAMENTO (PÚBLICO / CLIENTE)
# =========================================================
@app.route('/rastreamento', methods=['GET'])
def rastreamento():
    """
    Tela simples para o cliente digitar o código da entrega (ID)
    e acompanhar o status.
    """
    codigo = (request.args.get('codigo') or '').strip()
    entrega = None
    eventos = []

    if codigo.isdigit():
        entrega = Entrega.query.get(int(codigo))
        if entrega:
            eventos = montar_eventos_rastreamento(entrega)
        else:
            flash('Nenhuma entrega encontrada com esse código.', 'error')
    elif codigo:
        flash('Código inválido. Use apenas números.', 'error')

    return render_or_string(
        "rastreamento.html",
        """
        <!doctype html>
        <html lang="pt-BR">
        <head>
          <meta charset="utf-8">
          <title>Rastreamento de Entrega - Coopex</title>
          <meta name="viewport" content="width=device-width, initial-scale=1">
          <style>
            body{
              font-family:system-ui,-apple-system,Segoe UI,Roboto,Arial,sans-serif;
              margin:0;
              background:#0f172a;
              color:#e5e7eb;
            }
            .wrap{
              max-width:640px;
              margin:0 auto;
              padding:24px 16px 32px;
            }
            .card{
              background:#020617;
              border-radius:18px;
              padding:18px 16px;
              border:1px solid #1f2937;
              box-shadow:0 20px 45px rgba(0,0,0,.55);
            }
            h1{
              font-size:1.4rem;
              margin:0 0 10px;
              color:#f9fafb;
            }
            .sub{
              font-size:.85rem;
              color:#9ca3af;
              margin-bottom:14px;
            }
            form{
              display:flex;
              gap:8px;
              margin-bottom:16px;
              flex-wrap:wrap;
            }
            input[type=text]{
              flex:1;
              min-width:130px;
              padding:10px 12px;
              border-radius:999px;
              border:1px solid #374151;
              background:#020617;
              color:#e5e7eb;
              font-size:.9rem;
              outline:none;
            }
            input[type=text]::placeholder{
              color:#6b7280;
            }
            button{
              padding:10px 16px;
              border-radius:999px;
              border:none;
              background:#2563eb;
              color:#eef2ff;
              font-weight:700;
              font-size:.9rem;
              cursor:pointer;
              white-space:nowrap;
            }
            button:hover{background:#1d4ed8;}
            .msg{
              margin:6px 0 10px;
              padding:8px 10px;
              border-radius:10px;
              font-size:.8rem;
            }
            .msg-error{
              background:#7f1d1d;
              color:#fee2e2;
            }
            .msg-ok{
              background:#064e3b;
              color:#bbf7d0;
            }
            .entrega-card{
              margin-top:6px;
              padding:10px 12px;
              border-radius:12px;
              background:#020617;
              border:1px solid #1f2937;
            }
            .entrega-head{
              display:flex;
              justify-content:space-between;
              gap:10px;
              align-items:center;
              margin-bottom:6px;
              font-size:.86rem;
            }
            .chip{
              display:inline-flex;
              align-items:center;
              padding:2px 10px;
              border-radius:999px;
              font-size:.7rem;
              font-weight:700;
            }
            .chip-status{
              background:#0f172a;
              color:#e5e7eb;
              border:1px solid #4b5563;
            }
            .chip-pago{
              background:#022c22;
              color:#6ee7b7;
              border:1px solid #059669;
            }
            .chip-pendente{
              background:#3b0764;
              color:#f9a8d4;
              border:1px solid #db2777;
            }
            .linha-tempo{
              margin-top:10px;
              padding-left:6px;
              border-left:2px solid #1f2937;
            }
            .evento{
              padding-left:12px;
              margin-bottom:10px;
              position:relative;
            }
            .evento::before{
              content:"";
              width:10px;
              height:10px;
              border-radius:999px;
              background:#2563eb;
              border:2px solid #0f172a;
              position:absolute;
              left:-7px;
              top:4px;
            }
            .evento-titulo{
              font-size:.86rem;
              font-weight:700;
              display:flex;
              align-items:center;
              gap:6px;
              margin-bottom:2px;
            }
            .evento-texto{
              font-size:.8rem;
              color:#d1d5db;
            }
            .evento-when{
              font-size:.75rem;
              color:#9ca3af;
              margin-top:2px;
            }
            footer{
              margin-top:14px;
              font-size:.75rem;
              color:#6b7280;
              text-align:center;
            }
          </style>
        </head>
        <body>
          <div class="wrap">
            <div class="card">
              <h1>Rastreamento de Entrega</h1>
              <div class="sub">
                Digite o <strong>código da entrega</strong> (número que a Coopex te informar, ex: 1234)
                para acompanhar o status em tempo real.
              </div>

              <form method="get">
                <input type="text" name="codigo"
                       placeholder="Ex: 1234"
                       value="{{ codigo or '' }}">
                <button type="submit">Rastrear</button>
              </form>

              {% with messages = get_flashed_messages(with_categories=true) %}
                {% if messages %}
                  {% for cat, msg in messages %}
                    <div class="msg {{ 'msg-error' if cat=='error' else 'msg-ok' }}">{{ msg }}</div>
                  {% endfor %}
                {% endif %}
              {% endwith %}

              {% if entrega %}
                <div class="entrega-card">
                  <div class="entrega-head">
                    <div>
                      <div style="font-size:.8rem;color:#9ca3af;">Entrega #{{ entrega.id }}</div>
                      <div style="font-size:.95rem;font-weight:600;">
                        {{ entrega.cliente or 'Cliente não informado' }}
                      </div>
                      <div style="font-size:.8rem;color:#9ca3af;">
                        Bairro destino: {{ entrega.bairro or '---' }}
                      </div>
                    </div>
                    <div style="text-align:right">
                      <div class="chip chip-status">
                        {{ (entrega.status or 'pendente')|capitalize }}
                      </div>
                      <div style="margin-top:4px">
                        {% set stpg = (entrega.status_pagamento or 'pendente')|lower %}
                        {% if stpg == 'pago' %}
                          <span class="chip chip-pago">Pagamento: Pago</span>
                        {% else %}
                          <span class="chip chip-pendente">Pagamento: Pendente</span>
                        {% endif %}
                      </div>
                    </div>
                  </div>

                  <div style="font-size:.8rem;color:#9ca3af;margin-bottom:6px;">
                    Registrado em:
                    {% if entrega.data_envio %}
                      {{ to_brasilia(entrega.data_envio).strftime('%d/%m/%Y %H:%M') }}
                    {% else %}
                      -
                    {% endif %}

                    {% if entrega.cooperado %}
                      <br>Cooperado: {{ entrega.cooperado.nome }}
                    {% endif %}
                  </div>

                  <div class="linha-tempo">
                    {% if eventos %}
                      {% for ev in eventos %}
                        <div class="evento">
                          <div class="evento-titulo">
                            {% if ev.icone %}<span>{{ ev.icone }}</span>{% endif %}
                            <span>{{ ev.titulo }}</span>
                          </div>
                          <div class="evento-texto">{{ ev.descricao }}</div>
                          <div class="evento-when">
                            {% if ev.quando %}
                              {{ ev.quando.strftime('%d/%m/%Y %H:%M') }}
                            {% else %}
                              Horário não registrado
                            {% endif %}
                          </div>
                        </div>
                      {% endfor %}
                    {% else %}
                      <div class="evento">
                        <div class="evento-texto">
                          Nenhum evento de rastreio disponível ainda para esta entrega.
                        </div>
                      </div>
                    {% endif %}
                  </div>
                </div>
              {% endif %}
            </div>

            <footer>
              Coopex Entregas — sistema de rastreio interno.  
              Em caso de dúvidas, fale com a supervisão.
            </footer>
          </div>
        </body></html>
        """,
        codigo=codigo,
        entrega=entrega,
        eventos=eventos,
        to_brasilia=to_brasilia
    )


@app.get('/api/rastreamento/<codigo>')
def api_rastreamento(codigo):
    """
    API JSON para apps externos / site do cliente.
    Usa o ID da entrega como código de rastreio.
    """
    if not codigo.isdigit():
        return jsonify(ok=False, erro="Código inválido. Use apenas números."), 400

    entrega = Entrega.query.get(int(codigo))
    if not entrega:
        return jsonify(ok=False, erro="Entrega não encontrada."), 404

    eventos = montar_eventos_rastreamento(entrega)

    def _dt(dt):
        return to_brasilia(dt).isoformat() if dt else None

    try:
        origem_extra = json.loads(entrega.origem_json) if entrega.origem_json else None
    except Exception:
        origem_extra = None

    try:
        destino_extra = json.loads(entrega.destino_json) if entrega.destino_json else None
    except Exception:
        destino_extra = None

    return jsonify({
        "ok": True,
        "entrega_id": entrega.id,
        "cliente": entrega.cliente,
        "bairro": entrega.bairro,
        "valor": float(entrega.valor or 0),
        "status": entrega.status,
        "status_pagamento": entrega.status_pagamento,
        "pagamento": entrega.pagamento,
        "cooperado": (entrega.cooperado.nome if entrega.cooperado else None),
        "data_envio": _dt(entrega.data_envio),
        "data_atribuida": _dt(entrega.data_atribuida),
        "origem_extra": origem_extra,
        "destino_extra": destino_extra,
        "eventos": [
            {
                "titulo": ev["titulo"],
                "descricao": ev["descricao"],
                "quando": ev["quando"].isoformat() if ev["quando"] else None,
                "icone": ev.get("icone")
            }
            for ev in eventos
        ]
    })


# =========================================================
# ADMIN: DASHBOARD PRINCIPAL
# =========================================================
@app.route('/admin')
def admin():
    if not session.get('is_admin'):
        return redirect(url_for('login'))

    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')
    cooperado_id = request.args.get('cooperado_id', 'todos')
    status_pagamento = request.args.get('status_pagamento', 'todos')
    cliente = (request.args.get('cliente') or '').strip()

    query = Entrega.query

    # padrão: dia de hoje
    if not data_inicio and not data_fim:
        hoje_brasil = datetime.now(BRAZIL_TZ).date()
        inicio_utc, fim_utc = local_date_window_to_utc_range(hoje_brasil)
        query = query.filter(Entrega.data_envio >= inicio_utc, Entrega.data_envio <= fim_utc)

    if cooperado_id and cooperado_id != 'todos':
        query = query.filter(Entrega.cooperado_id == int(cooperado_id))

    if data_inicio:
        di = datetime.strptime(data_inicio, "%Y-%m-%d").date()
        inicio_utc, _ = local_date_window_to_utc_range(di)
        query = query.filter(Entrega.data_envio >= inicio_utc)

    if data_fim:
        df_ = datetime.strptime(data_fim, "%Y-%m-%d").date()
        _, fim_utc = local_date_window_to_utc_range(df_)
        query = query.filter(Entrega.data_envio <= fim_utc)

    if status_pagamento and status_pagamento != 'todos':
        if status_pagamento == 'pago':
            query = query.filter(func.lower(Entrega.status_pagamento) == 'pago')
        elif status_pagamento == 'pendente':
            query = query.filter(
                (Entrega.status_pagamento == None) |
                (func.lower(Entrega.status_pagamento) == 'pendente')
            )

    if cliente:
        like = f"%{cliente.lower()}%"
        query = query.filter(func.lower(Entrega.cliente).like(like))

    entregas_all = (
        query.options(joinedload(Entrega.cooperado))
        .order_by(Entrega.data_envio.desc())
        .all()
    )
    nao_atribuidos = [e for e in entregas_all if not e.cooperado_id]
    atribuidos = [e for e in entregas_all if e.cooperado_id]
    entregas = nao_atribuidos + atribuidos

    # AQUI você já tinha isso:
    cooperados = Cooperado.query.order_by(Cooperado.nome).all()

    hoje = datetime.now(BRAZIL_TZ).date()
    inicio_dia_utc, fim_dia_utc = local_date_window_to_utc_range(hoje)

    total_dia = Entrega.query.filter(
        Entrega.data_envio >= inicio_dia_utc,
        Entrega.data_envio <= fim_dia_utc
    ).count()
    mes_ini_utc, mes_fim_utc = month_range_utc(hoje)
    total_mes = Entrega.query.filter(
        Entrega.data_envio >= mes_ini_utc,
        Entrega.data_envio <= mes_fim_utc
    ).count()
    ano_ini_utc, ano_fim_utc = year_range_utc(hoje)
    total_ano = Entrega.query.filter(
        Entrega.data_envio >= ano_ini_utc,
        Entrega.data_envio <= ano_fim_utc
    ).count()
    estatisticas = {"total_dia": total_dia, "total_mes": total_mes, "total_ano": total_ano}

    feriado_hoje = verifica_feriado(hoje)
    tem_pendente = Entrega.query.filter(
        Entrega.data_envio >= inicio_dia_utc,
        Entrega.data_envio <= fim_dia_utc,
        (Entrega.status_pagamento == None) |
        (func.lower(Entrega.status_pagamento) == 'pendente')
    ).count() > 0

    lista_espera = ListaEspera.query.order_by(ListaEspera.pos.asc(), ListaEspera.created_at.asc()).all()
    ids_em_fila = {it.cooperado_id for it in lista_espera if it.cooperado_id}
    cooperados_disponiveis = [c for c in cooperados if c.id not in ids_em_fila]

    # >>> NOVO: listas seguras para o JavaScript <<<
    cooperados_js = [
        {
            "id": c.id,
            "nome": c.nome
        }
        for c in cooperados
    ]

    motoboys_js = []
    for c in cooperados:
        if getattr(c, "last_lat", None) is not None and getattr(c, "last_lng", None) is not None:
            is_online, idle_s, status_str = calc_status_cooperado(c)

            motoboys_js.append({
                "id": c.id,
                "nome": c.nome,
                "lat": c.last_lat,
                "lng": c.last_lng,
                "online": bool(is_online),
                "status": status_str,
                "idle_seconds": idle_s,
                "velocidade": float(getattr(c, "last_speed_kmh", 0) or 0),
                "ultima_atualizacao": to_brasilia(c.last_ping).strftime('%d/%m %H:%M') if c.last_ping else ""
            })

    return render_template(
        'admin.html',
        entregas=entregas,
        cooperados=cooperados,
        estatisticas=estatisticas,
        data_inicio=data_inicio,
        data_fim=data_fim,
        to_brasilia=to_brasilia,
        request=request,
        now=lambda: datetime.now(BRAZIL_TZ),
        feriado_hoje=feriado_hoje,
        tem_pendente=tem_pendente,
        lista_espera=lista_espera,
        cooperados_disponiveis=cooperados_disponiveis,
        # >>> VARIÁVEIS NOVAS PARA O JS <<<
        cooperados_js=cooperados_js,
        motoboys_js=motoboys_js,
    )


@app.route("/admin_novo_socorro")
def admin_novo_socorro():
    """Rota que o admin consulta (polling) para saber se há socorros pendentes.
    Importante: **não** marca como lido aqui. Só marca quando o admin clicar no X.
    """
    if not session.get("is_admin") and not session.get("is_master"):
        abort(403)

    global SOCORRO_QUEUE

    pendentes = [s for s in (SOCORRO_QUEUE or []) if not s.get("lido")]
    if not pendentes:
        return jsonify({"novo": False, "count": 0}), 200

    ultimo = pendentes[-1]
    return jsonify({
        "novo": True,
        "count": len(pendentes),
        "id": ultimo.get("id"),
        "cooperado": ultimo.get("cooperado_nome"),
        "mensagem": ultimo.get("mensagem") or "",
        "momento": ultimo.get("momento"),
    }), 200



@app.post("/admin_socorro_marcar_lido")
def admin_socorro_marcar_lido():
    """Admin confirma que viu o socorro (clicou no X)."""
    if not session.get("is_admin") and not session.get("is_master"):
        abort(403)

    global SOCORRO_QUEUE

    data = request.get_json(silent=True) or {}
    sid = data.get("id")
    try:
        sid_int = int(sid)
    except Exception:
        return jsonify(ok=False, error="id inválido"), 400

    found = False
    for s in SOCORRO_QUEUE:
        if int(s.get("id") or 0) == sid_int:
            s["lido"] = True
            found = True
            break

    if not found:
        return jsonify(ok=False, error="socorro não encontrado"), 404

    pendentes = [s for s in (SOCORRO_QUEUE or []) if not s.get("lido")]
    return jsonify(ok=True, count=len(pendentes))

# =========================================================
# ADMIN — visualizar / baixar comprovante (foto) da entrega
# =========================================================
@app.get("/admin/entrega/<int:entrega_id>/comprovante")
def admin_ver_comprovante(entrega_id):
    if not session.get("is_admin") and not session.get("is_master"):
        abort(403)
    info = comprovante_info(entrega_id)
    if not info or not info.get("filename"):
        abort(404)
    fp = os.path.join(COMPROVANTE_DIR, info["filename"])
    if not os.path.exists(fp):
        abort(404)
    # envia inline (abre no navegador)
    return send_file(fp)

@app.get("/admin/entrega/<int:entrega_id>/comprovante/download")
def admin_baixar_comprovante(entrega_id):
    if not session.get("is_admin") and not session.get("is_master"):
        abort(403)
    info = comprovante_info(entrega_id)
    if not info or not info.get("filename"):
        abort(404)
    fp = os.path.join(COMPROVANTE_DIR, info["filename"])
    if not os.path.exists(fp):
        abort(404)
    return send_file(fp, as_attachment=True, download_name=info["filename"])


# ================================
# PAINEL DO COOPERADO (ESTILO UBER)
# ================================
@app.route('/painel_cooperado')
def painel_cooperado():
    # Cooperado logado = precisa ter user_id na sessão E NÃO ser admin
    if session.get('user_id') is None or session.get('is_admin'):
        return redirect(url_for('login'))

    user_id = session['user_id']

    inicio = request.args.get('inicio')
    fim = request.args.get('fim')
    status_pgto = (request.args.get('status_pgto') or 'todas').lower()
    todas_datas_flag = (request.args.get('todas_datas') or '') == '1'

    # Base de consultas: entregas desse cooperado
    base_q = Entrega.query.filter(Entrega.cooperado_id == user_id)

    # ========== CORRIDAS EM ABERTO / EM ANDAMENTO ==========
    # Qualquer corrida que ainda não esteja finalizada
    corridas_query = (
        base_q
        .filter(
            (Entrega.status_corrida == None) |
            (Entrega.status_corrida.in_(['pendente', 'aceita']))
        )
        .filter(
            (Entrega.status == None) |
            (~func.lower(Entrega.status).in_(['recebido', 'entregue']))
        )
        .order_by(Entrega.data_envio.desc())
    )

    corridas_raw = corridas_query.all()

    def _parse_json_field(raw):
        """Tenta fazer json.loads, se vier string; se der erro, devolve {}."""
        if not raw:
            return {}
        if isinstance(raw, dict):
            return raw
        try:
            return json.loads(raw)
        except Exception:
            return {}

    corridas = []
    for e in corridas_raw:
        origem = _parse_json_field(e.origem_json)
        destino = _parse_json_field(e.destino_json)
        paradas = _parse_json_field(e.paradas_json)

        origem_endereco = (
            origem.get('endereco')
            or origem.get('address')
            or (origem.get('rua') and f"{origem.get('rua')} {origem.get('numero', '')}".strip())
            or e.bairro
            or 'Origem não informada'
        )

        destino_endereco = (
            destino.get('endereco')
            or destino.get('address')
            or (destino.get('rua') and f"{destino.get('rua')} {destino.get('numero', '')}".strip())
            or 'Destino não informado'
        )

        origem_bairro = origem.get('bairro') or ''
        destino_bairro = destino.get('bairro') or ''

        # Lista simples de paradas intermediárias
        waypoints = paradas.get('stops') or paradas.get('paradas') or []

        corridas.append({
            "obj": e,
            "origem_endereco": origem_endereco,
            "destino_endereco": destino_endereco,
            "origem_bairro": origem_bairro,
            "destino_bairro": destino_bairro,
            "waypoints": waypoints,
        })

    # ========== HISTÓRICO (TABELA) ==========
    query = base_q

    # Filtro por status de pagamento
    if status_pgto == 'pago':
        query = query.filter(func.lower(Entrega.status_pagamento) == 'pago')
    elif status_pgto == 'pendente':
        query = query.filter(
            (Entrega.status_pagamento == None) |
            (func.lower(Entrega.status_pagamento) == 'pendente')
        )

    # Filtros de data
    if not todas_datas_flag:
        hoje_brasil = datetime.now(BRAZIL_TZ).date()

        # Nenhuma data informada -> dia atual
        if not inicio and not fim:
            inicio_utc, fim_utc = local_date_window_to_utc_range(hoje_brasil)
            query = query.filter(
                Entrega.data_envio >= inicio_utc,
                Entrega.data_envio <= fim_utc
            )

        # Data inicial
        if inicio:
            di = datetime.strptime(inicio, "%Y-%m-%d").date()
            inicio_utc, _ = local_date_window_to_utc_range(di)
            query = query.filter(Entrega.data_envio >= inicio_utc)

        # Data final
        if fim:
            df_ = datetime.strptime(fim, "%Y-%m-%d").date()
            _, fim_utc = local_date_window_to_utc_range(df_)
            query = query.filter(Entrega.data_envio <= fim_utc)

    # Ordenação e carregamento do cooperado (se precisar no template)
    entregas = (
        query
        .options(joinedload(Entrega.cooperado))
        .order_by(Entrega.data_envio.desc())
        .all()
    )

    # Totais
    total_geral = sum(float(e.valor or 0) for e in entregas)
    total_pago = sum(
        float(e.valor or 0)
        for e in entregas
        if (e.status_pagamento or '').lower() == 'pago'
    )
    total_pendente = max(0.0, total_geral - total_pago)

    return render_template(
        'painel_cooperado.html',
        entregas=entregas,
        corridas=corridas,
        total_geral=total_geral,
        total_pago=total_pago,
        total_pendente=total_pendente,
        request=request,
        to_brasilia=to_brasilia,
        status_pgto=status_pgto,
        ano_atual=datetime.now(BRAZIL_TZ).year,
        mes_atual=datetime.now(BRAZIL_TZ).month,
        meses_ano=[
            {'num':1,'nome':'Janeiro'},{'num':2,'nome':'Fevereiro'},{'num':3,'nome':'Março'},{'num':4,'nome':'Abril'},
            {'num':5,'nome':'Maio'},{'num':6,'nome':'Junho'},{'num':7,'nome':'Julho'},{'num':8,'nome':'Agosto'},
            {'num':9,'nome':'Setembro'},{'num':10,'nome':'Outubro'},{'num':11,'nome':'Novembro'},{'num':12,'nome':'Dezembro'},
        ],
    )

@app.route("/cooperado/verificar_nova_entrega")
def cooperado_verificar_nova_entrega():
    cooperado_id = session.get("user_id")
    if not cooperado_id or session.get('is_admin'):
        return jsonify({"tem_entrega": False})

    # Busca uma entrega ATRIBUÍDA para esse cooperado,
    # ainda não concluída e ainda "pendente" na visão da corrida.
    entrega = (
        Entrega.query
        .filter(
            Entrega.cooperado_id == cooperado_id,
            # ainda em aberto para o cooperado
            (Entrega.status_corrida == None) |
            (Entrega.status_corrida.in_(['pendente', 'aceita'])),
            # não concluída
            (Entrega.status == None) |
            (~func.lower(Entrega.status).in_(['recebido', 'entregue']))
        )
        .order_by(Entrega.data_atribuida.desc(), Entrega.data_envio.desc())
        .first()
    )

    if not entrega:
        return jsonify({"tem_entrega": False})

    # Usa os helpers do model Entrega para pegar origem/destino:
    origem = entrega.get_origem() or {}
    destino = entrega.get_destino() or {}

    origem_endereco = (
        origem.get('endereco')
        or origem.get('address')
        or origem.get('bairro')
        or None
    )
    origem_bairro = origem.get('bairro') or None

    destino_endereco = (
        destino.get('endereco')
        or destino.get('address')
        or destino.get('bairro')
        or entrega.bairro
    )
    destino_bairro = destino.get('bairro') or entrega.bairro

    payload = {
        "id": entrega.id,
        "cliente": entrega.cliente,
        "valor": float(entrega.valor or 0),

        "origem_endereco": origem_endereco,
        "origem_bairro": origem_bairro,

        "destino_endereco": destino_endereco,
        "destino_bairro": destino_bairro,

        "lat_origem": origem.get('lat'),
        "lng_origem": origem.get('lng'),
        "lat_destino": destino.get('lat'),
        "lng_destino": destino.get('lng'),

        "tempo_estimado": "aprox.",
        "distancia": 0,
        "status_pagamento": (entrega.status_pagamento or "").lower(),
        "data_entrega": entrega.data_envio.strftime("%Y-%m-%d") if entrega.data_envio else None,
        "recebida_por": entrega.recebido_por or "",
    }

    return jsonify({"tem_entrega": True, "entrega": payload})


# Aceitar entrega (via URL com <id>)
@app.route("/cooperado/aceitar_entrega", methods=["POST"])
def cooperado_aceitar_entrega():
    """
    Motoboy aceita uma entrega enviada pela supervisão.
    Deve SEMPRE devolver JSON.
    """

    # garante que é cooperado logado (mesma regra das outras rotas de cooperado)
    if session.get("user_id") is None or session.get("is_admin"):
        return jsonify({"status": "erro", "msg": "Não autorizado."}), 403

    cooperado_id = session["user_id"]

    # 🔴 AQUI é a parte que estava pegando só JSON
    # Agora tenta JSON OU form
    dados = request.get_json(silent=True) or {}
    entrega_id_raw = dados.get("entrega_id") or request.form.get("entrega_id")

    if not entrega_id_raw:
        return jsonify({"status": "erro", "msg": "id de entrega nao informado"}), 400

    # tenta converter pra int (segurança extra)
    try:
        entrega_id = int(entrega_id_raw)
    except ValueError:
        return jsonify({"status": "erro", "msg": "id de entrega inválido"}), 400

    entrega = Entrega.query.get(entrega_id)
    if not entrega:
        return jsonify({"status": "erro", "msg": "Entrega não encontrada."}), 404

    # Se já tiver cooperado diferente, não deixa "roubar"
    if entrega.cooperado_id and entrega.cooperado_id != cooperado_id:
        return jsonify({
            "status": "erro",
            "msg": "Essa entrega já foi aceita por outro motoboy."
        }), 409

    # Marca entrega como atribuída para esse cooperado
    entrega.cooperado_id = cooperado_id
    entrega.status = "em_andamento"   # usa o mesmo campo que você já usa no sistema
    entrega.status_corrida = "aceita"
    entrega.data_atribuida = datetime.utcnow()

    db.session.commit()

    entrega_json = {
        "id": entrega.id,
        "cliente": getattr(entrega, "cliente", None),
        "restaurante": getattr(entrega, "cliente", None),
        "valor": float(entrega.valor or 0),
        "origem_endereco": None,
        "origem_bairro": None,
        "destino_endereco": entrega.bairro,
        "destino_bairro": entrega.bairro,
        "lat_origem": None,
        "lng_origem": None,
        "lat_destino": None,
        "lng_destino": None,
        "distancia": None,
        "tempo_estimado": None,
        "status_pagamento": getattr(entrega, "status_pagamento", "pendente"),
        "recebida_por": getattr(entrega, "recebido_por", None),
        "data": (
            entrega.data_envio.date().isoformat()
            if getattr(entrega, "data_envio", None) else None
        ),
    }

    return jsonify({"status": "ok", "entrega": entrega_json}), 200


# Recusar entrega (via URL com <id>)
@app.route("/cooperado/recusar_entrega", methods=["POST"])
def cooperado_recusar_entrega():
    if session.get("user_id") is None or session.get("is_admin"):
        return jsonify(status="erro", msg="Não autorizado"), 401

    data = request.get_json() or {}
    entrega_id = data.get("entrega_id")

    if not entrega_id:
        return jsonify(status="erro", msg="ID de entrega não informado"), 400

    entrega = Entrega.query.get(entrega_id)
    if not entrega:
        return jsonify(status="erro", msg="Entrega não encontrada"), 404

    user_id = session["user_id"]
    if entrega.cooperado_id != user_id:
        # se quiser permitir recusa mesmo antes de atribuir, pode tirar esse if
        return jsonify(status="erro", msg="Entrega não pertence a este cooperado"), 403

    # volta pra fila do admin
    entrega.cooperado_id = None
    if hasattr(entrega, "status_entrega"):
        entrega.status_entrega = "pendente"
    if hasattr(entrega, "hora_atribuida"):
        entrega.hora_atribuida = None

    db.session.commit()
    return jsonify(status="ok")

@app.route("/cooperado/finalizar_entrega", methods=["POST"])
def cooperado_finalizar_entrega():
    if session.get("user_id") is None or session.get("is_admin"):
        return jsonify(status="erro", msg="Não autorizado"), 401

    data = request.get_json() or {}
    entrega_id = data.get("entrega_id")
    recebida_por = (data.get("recebida_por") or "").strip()

    if not entrega_id:
        return jsonify(status="erro", msg="ID de entrega não informado"), 400

    if not recebida_por:
        return jsonify(status="erro", msg="Nome de quem recebeu é obrigatório"), 400

    entrega = Entrega.query.get(entrega_id)
    if not entrega:
        return jsonify(status="erro", msg="Entrega não encontrada"), 404

    user_id = session["user_id"]
    if entrega.cooperado_id != user_id:
        return jsonify(status="erro", msg="Entrega não pertence a este cooperado"), 403

    # 👉 aqui NÃO tem checagem de localização, pode finalizar de qualquer lugar
    entrega.recebida_por = recebida_por

    if hasattr(entrega, "status_entrega"):
        entrega.status_entrega = "finalizada"  # isso vai aparecer como entregue no painel admin

    from datetime import datetime
    if hasattr(entrega, "hora_finalizada"):
        entrega.hora_finalizada = datetime.utcnow()

    # status_pagamento continua pendente, motoboy marca depois
    db.session.commit()

    entrega_dict = {
      "id": entrega.id,
      "cliente": getattr(entrega, "cliente", None),
      "restaurante": getattr(entrega, "restaurante", None),
      "origem_bairro": getattr(entrega, "origem_bairro", None),
      "destino_bairro": getattr(entrega, "destino_bairro", None),
      "origem_endereco": getattr(entrega, "origem_endereco", None),
      "destino_endereco": getattr(entrega, "destino_endereco", None),
      "valor": float(getattr(entrega, "valor", 0) or 0),
      "data": getattr(entrega, "data_entrega", None) or "",
      "recebida_por": entrega.recebida_por,
      "status_pagamento": getattr(entrega, "status_pagamento", "pendente"),
    }

    return jsonify(status="ok", entrega=entrega_dict)


# Aceitar via API (AJAX/Fetch com JSON)
@app.route('/cooperado/api/aceitar', methods=['POST'])
def cooperado_aceitar_corrida():
    if session.get('user_id') is None or session.get('is_admin'):
        return jsonify(ok=False, error='Não autorizado'), 401

    user_id = session['user_id']
    data = request.get_json() or {}
    entrega_id = data.get('entrega_id')

    if not entrega_id:
        return jsonify(ok=False, error='entrega_id obrigatório'), 400

    entrega = Entrega.query.get_or_404(entrega_id)

    if entrega.cooperado_id != user_id:
        return jsonify(ok=False, error='Entrega não pertence a este cooperado'), 403

    entrega.status_corrida = 'aceita'
    if not entrega.data_atribuida:
        entrega.data_atribuida = datetime.now(BRAZIL_TZ)

    db.session.commit()
    return jsonify(ok=True, status_corrida=entrega.status_corrida)

@app.route('/cooperado/atualizar_localizacao', methods=['POST'])
def cooperado_atualizar_localizacao():
    if session.get('user_id') is None or session.get('is_admin'):
        return jsonify({'status': 'erro', 'msg': 'Não autorizado'}), 403

    cooperado_id = session['user_id']
    cooperado = Cooperado.query.get(cooperado_id)
    if not cooperado:
        return jsonify({'status': 'erro', 'msg': 'Cooperado não encontrado'}), 404

    data = request.get_json(silent=True) or {}

    # lat/lng obrigatórios
    try:
        lat = float(data.get('lat'))
        lng = float(data.get('lng'))
    except (TypeError, ValueError):
        return jsonify({'status': 'erro', 'msg': 'Lat/Lng inválidos'}), 400

    # speed pode vir em m/s (Geolocation API) OU km/h (se você mandar assim)
    speed_mps = data.get('speed_mps', None)
    speed_kmh = data.get('velocidade', None)  # compatível com seu campo atual

    # heading/accuracy opcionais
    heading = data.get('heading', None)
    accuracy = data.get('accuracy', None)

    # normaliza velocidade
    v_kmh = None
    try:
        if speed_mps is not None:
            v_kmh = float(speed_mps) * 3.6
        elif speed_kmh is not None:
            v_kmh = float(speed_kmh)
    except (TypeError, ValueError):
        v_kmh = None

    # salva no banco
    cooperado.last_lat = lat
    cooperado.last_lng = lng
    cooperado.last_ping = datetime.utcnow()
    cooperado.online = True

    cooperado.last_speed_kmh = v_kmh
    try:
        cooperado.last_heading = float(heading) if heading is not None else None
    except (TypeError, ValueError):
        cooperado.last_heading = None
    try:
        cooperado.last_accuracy_m = float(accuracy) if accuracy is not None else None
    except (TypeError, ValueError):
        cooperado.last_accuracy_m = None

    # marca “último movimento”
    if v_kmh is not None and v_kmh >= MOVING_SPEED_KMH:
        cooperado.last_moving_at = datetime.utcnow()

    db.session.commit()

    # emite para o painel em tempo real (adicione campos no payload, item 4)
    emitir_posicao_motoboy(cooperado, lat, lng, v_kmh)

    return jsonify({'status': 'ok'})


# Recusar via API (AJAX/Fetch com JSON)
@app.route('/cooperado/api/recusar', methods=['POST'])
def cooperado_recusar_corrida():
    if session.get('user_id') is None or session.get('is_admin'):
        return jsonify(ok=False, error='Não autorizado'), 401

    user_id = session['user_id']
    data = request.get_json() or {}
    entrega_id = data.get('entrega_id')

    if not entrega_id:
        return jsonify(ok=False, error='entrega_id obrigatório'), 400

    entrega = Entrega.query.get_or_404(entrega_id)

    if entrega.cooperado_id != user_id:
        return jsonify(ok=False, error='Entrega não pertence a este cooperado'), 403

    entrega.status_corrida = 'recusada'
    db.session.commit()
    return jsonify(ok=True, status_corrida=entrega.status_corrida)


@app.route('/cooperado/api/novas', methods=['GET'])
def cooperado_novas_corridas():
    if session.get('user_id') is None or session.get('is_admin'):
        return jsonify(ok=False, error='Não autorizado'), 401

    user_id = session['user_id']

    q = (
        Entrega.query
        .filter(Entrega.cooperado_id == user_id)
        .filter(
            (Entrega.status_corrida == None) |
            (Entrega.status_corrida == 'pendente')
        )
        .filter(
            (Entrega.status == None) |
            (~func.lower(Entrega.status).in_(['recebido', 'entregue']))
        )
    )
    novas = q.count()
    return jsonify(ok=True, novas=novas)

@app.get('/api/mobile/cooperado/corridas')
def api_mobile_cooperado_corridas():
    """
    Lista as corridas em aberto / em andamento para o cooperado logado.
    Usado pela tela principal do app nativo.

    Responde JSON:
    {
      "ok": true,
      "corridas": [
        {
          "id": ...,
          "cliente": "...",
          "valor": 12.34,
          "origem_endereco": "...",
          "origem_bairro": "...",
          "destino_endereco": "...",
          "destino_bairro": "...",
          "status_corrida": "pendente"/"aceita",
          "status": "pendente"/"em_andamento"/"entregue",
          "status_pagamento": "pago"/"pendente"
        },
        ...
      ]
    }
    """
    if session.get('user_id') is None or session.get('is_admin'):
        return jsonify(ok=False, error='Não autorizado'), 401

    user_id = session['user_id']

    base_q = Entrega.query.filter(Entrega.cooperado_id == user_id)

    q = (
        base_q
        .filter(
            (Entrega.status_corrida == None) |
            (Entrega.status_corrida.in_(['pendente', 'aceita']))
        )
        .filter(
            (Entrega.status == None) |
            (~func.lower(Entrega.status).in_(['recebido', 'entregue']))
        )
        .order_by(Entrega.data_envio.desc())
    )

    def _parse_json_field(raw):
        if not raw:
            return {}
        if isinstance(raw, dict):
            return raw
        try:
            return json.loads(raw)
        except Exception:
            return {}

    corridas = []
    for e in q.all():
        origem = _parse_json_field(e.origem_json)
        destino = _parse_json_field(e.destino_json)

        origem_endereco = (
            origem.get('endereco')
            or origem.get('address')
            or origem.get('bairro')
            or e.bairro
            or 'Origem não informada'
        )
        destino_endereco = (
            destino.get('endereco')
            or destino.get('address')
            or destino.get('bairro')
            or e.bairro
            or 'Destino não informado'
        )

        corridas.append({
            "id": e.id,
            "cliente": e.cliente,
            "valor": float(e.valor or 0),
            "origem_endereco": origem_endereco,
            "origem_bairro": origem.get('bairro') or '',
            "destino_endereco": destino_endereco,
            "destino_bairro": destino.get('bairro') or e.bairro,
            "status_corrida": e.status_corrida,
            "status": e.status,
            "status_pagamento": e.status_pagamento,
        })

    return jsonify(ok=True, corridas=corridas)

# variável global bem simples pra sinalizar um novo socorro
SOCORRO_QUEUE = []
NEXT_SOCORRO_ID = 1

@app.route("/cooperado_socorro", methods=["POST"])
def cooperado_socorro():
    """Cooperado pede ajuda (socorro).
    Guarda em fila global simples para o admin visualizar até marcar como lido.
    """
    global SOCORRO_QUEUE, NEXT_SOCORRO_ID

    data = request.get_json(silent=True) or {}
    tipo = data.get("tipo")
    detalhes = (data.get("detalhes") or "").strip()

    if not tipo:
        return jsonify({"ok": False, "error": "Tipo de socorro não informado."}), 400

    cooperado_id = session.get("user_id")
    cooperado_nome = session.get("user_nome", "Cooperado")

    agora_brt = datetime.now(BRAZIL_TZ)

    item = {
        "id": int(NEXT_SOCORRO_ID),
        "cooperado_id": cooperado_id,
        "cooperado_nome": cooperado_nome,
        "mensagem": f"{tipo}: {detalhes}" if detalhes else str(tipo),
        "momento": agora_brt.strftime("%d/%m/%Y %H:%M"),
        "timestamp": datetime.utcnow().isoformat(),
        "lido": False,
    }
    NEXT_SOCORRO_ID += 1
    SOCORRO_QUEUE.append(item)

    # emite via socket (se o admin estiver conectado)
    try:
        socketio.emit("socorro_novo", item, broadcast=True)
    except Exception:
        pass

    return jsonify({"ok": True, "id": item["id"]})

# ================================
# CRUD de COOPERADO (mantidos)
# ================================
@app.route('/cooperados/cadastrar', methods=['GET', 'POST'])
def cadastrar_cooperado():
    if not session.get('is_admin'):
        return redirect(url_for('login'))

    if request.method == 'POST':
        nome = request.form.get('nome')
        senha = request.form.get('senha')
        if nome and senha:
            if Cooperado.query.filter_by(nome=nome).first():
                flash('Já existe um cooperado com esse nome!')
            else:
                novo = Cooperado(nome=nome)
                novo.set_senha(senha)
                db.session.add(novo)
                db.session.commit()
                flash('Cooperado cadastrado com sucesso!')
        else:
            flash('Preencha todos os campos.')
        return redirect(url_for('cadastrar_cooperado'))

    cooperados = Cooperado.query.order_by(Cooperado.nome).all()
    return render_template('cadastrar_cooperado.html', cooperados=cooperados)


@app.route('/cooperados/<int:coop_id>/atualizar', methods=['POST'])
def atualizar_cooperado(coop_id):
    if not session.get('is_admin'):
        return redirect(url_for('login'))

    cooperado = Cooperado.query.get_or_404(coop_id)
    novo_nome = request.form.get('novo_nome')
    nova_senha = request.form.get('nova_senha')

    if novo_nome and novo_nome != cooperado.nome:
        existe = Cooperado.query.filter_by(nome=novo_nome).first()
        if existe and existe.id != cooperado.id:
            flash('Já existe um cooperado com esse nome!')
            return redirect(url_for('cadastrar_cooperado'))
        cooperado.nome = novo_nome

    if nova_senha:
        cooperado.set_senha(nova_senha)

    db.session.commit()
    flash('Dados do cooperado atualizados!')
    return redirect(url_for('cadastrar_cooperado'))


@app.route('/cooperados/<int:coop_id>/excluir', methods=['POST'])
def excluir_cooperado(coop_id):
    if not session.get('is_admin'):
        return redirect(url_for('login'))

    cooperado = Cooperado.query.get_or_404(coop_id)
    db.session.delete(cooperado)
    db.session.commit()
    flash('Cooperado excluído com sucesso!')
    return redirect(url_for('cadastrar_cooperado'))


@app.route('/cooperados/<int:coop_id>/status', methods=['POST'])
def mudar_status_cooperado(coop_id):
    if not session.get('is_admin'):
        return redirect(url_for('login'))

    novo_status = request.form.get('novo_status')
    cooperado = Cooperado.query.get_or_404(coop_id)
    cooperado.ativo = (novo_status == "1")
    db.session.commit()
    flash(f"Status de {cooperado.nome} alterado para {'Ativo' if cooperado.ativo else 'Inativo'}!")
    return redirect(url_for('cadastrar_cooperado'))


# =========================================================
# CLIENTES (CRUD BÁSICO)
# =========================================================
@app.route('/clientes', methods=['GET', 'POST'])
def clientes():
    if not session.get('is_admin'):
        return redirect(url_for('login'))

    if request.method == 'POST':
        nome = (request.form.get('nome') or '').strip()
        telefone = _norm_phone(request.form.get('telefone') or '')
        bairro_origem = (request.form.get('bairro_origem') or '').strip()
        endereco = (request.form.get('endereco') or '').strip()
        if not nome:
            flash('Informe o nome do cliente.')
            return redirect(url_for('clientes'))

        existe = Cliente.query.filter(func.lower(Cliente.nome) == nome.lower()).first()
        if existe:
            flash('Já existe um cliente com esse nome.')
            return redirect(url_for('clientes'))

        cl = Cliente(
            nome=nome,
            telefone=telefone,
            bairro_origem=bairro_origem,
            endereco=endereco or None
        )
        db.session.add(cl)
        db.session.commit()
        flash('Cliente cadastrado!')
        return redirect(url_for('clientes'))

    # Métricas
    aggs = (
        db.session.query(
            Entrega.cliente.label('cli'),
            func.count(Entrega.id).label('qtd'),
            func.max(Entrega.data_envio).label('ultimo')
        )
        .group_by(Entrega.cliente)
        .all()
    )

    stats_by_full = defaultdict(lambda: {"qtd": 0, "ultimo": None})
    stats_by_first = defaultdict(lambda: {"qtd": 0, "ultimo": None})
    for row in aggs:
        raw = (row.cli or '').strip()
        key_full = normalize_letters_key(raw)
        key_first = normalize_first_token(raw)

        s = stats_by_full[key_full]
        s["qtd"] += int(row.qtd or 0)
        if row.ultimo and (s["ultimo"] is None or row.ultimo > s["ultimo"]):
            s["ultimo"] = row.ultimo

        f = stats_by_first[key_first]
        f["qtd"] += int(row.qtd or 0)
        if row.ultimo and (f["ultimo"] is None or row.ultimo > f["ultimo"]):
            f["ultimo"] = row.ultimo

    hoje_local = datetime.now(BRAZIL_TZ).date()
    lista = []
    for cl in Cliente.query.order_by(Cliente.nome).all():
        k_full = normalize_letters_key(cl.nome or '')
        k_first = normalize_first_token(cl.nome or '')

        tot, dt = 0, None
        if k_full in stats_by_full:
            tot = stats_by_full[k_full]["qtd"]
            dt = stats_by_full[k_full]["ultimo"]
        elif k_first in stats_by_first:
            tot = stats_by_first[k_first]["qtd"]
            dt = stats_by_first[k_first]["ultimo"]

        ultimo_ymd, ultimo_br, ultimo_days, row_class = None, None, None, ""
        if dt:
            loc_date = to_brasilia(dt).date()
            ultimo_ymd = loc_date.isoformat()
            ultimo_br = loc_date.strftime('%d/%m/%Y')
            ultimo_days = (hoje_local - loc_date).days
            if ultimo_days > 60:
                row_class = "st-gt60"
            elif ultimo_days > 30:
                row_class = "st-gt30"
            else:
                row_class = "st-lt30"

        lista.append({
            "id": cl.id,
            "nome": cl.nome,
            "telefone": cl.telefone,
            "bairro_origem": cl.bairro_origem,
            "endereco": getattr(cl, "endereco", None),
            "total_pedidos": int(tot or 0),
            "ultimo_ymd": ultimo_ymd,
            "ultimo_br": ultimo_br,
            "ultimo_days": ultimo_days,
            "row_class": row_class
        })

    total_clientes = len(lista)
    ativos = sum(1 for i in lista if i["ultimo_days"] is not None and i["ultimo_days"] <= 180)
    inativos = total_clientes - ativos

    return render_template(
        'clientes.html',
        clientes=lista,
        kpis={"total": total_clientes, "ativos": ativos, "inativos": inativos}
    )


@app.route('/clientes/<int:id>/editar', methods=['POST'])
def editar_cliente(id):
    if not session.get('is_admin'):
        return redirect(url_for('login'))

    cl = Cliente.query.get_or_404(id)
    nome = (request.form.get('nome') or '').strip()
    telefone = _norm_phone(request.form.get('telefone') or '')
    bairro_origem = (request.form.get('bairro_origem') or '').strip()
    endereco = (request.form.get('endereco') or '').strip()

    if not nome:
        if request.headers.get('X-Requested-With') == 'fetch' or request.args.get('format') == 'json' or (request.accept_mimetypes and request.accept_mimetypes.best == 'application/json'):
            return jsonify(ok=False, error='Informe o nome do cliente.'), 400
        flash('Informe o nome do cliente.')
        return redirect(url_for('clientes'))

    existe = Cliente.query.filter(
        func.lower(Cliente.nome) == nome.lower(),
        Cliente.id != id
    ).first()
    if existe:
        if request.headers.get('X-Requested-With') == 'fetch' or request.args.get('format') == 'json' or (request.accept_mimetypes and request.accept_mimetypes.best == 'application/json'):
            return jsonify(ok=False, error='Já existe outro cliente com esse nome.'), 400
        flash('Já existe outro cliente com esse nome.')
        return redirect(url_for('clientes'))

    cl.nome = nome
    cl.telefone = telefone
    cl.bairro_origem = bairro_origem
    cl.endereco = endereco or None
    db.session.commit()

    if request.headers.get('X-Requested-With') == 'fetch' or request.args.get('format') == 'json' or (request.accept_mimetypes and request.accept_mimetypes.best == 'application/json'):
        aggs = (
            db.session.query(
                Entrega.cliente.label('cli'),
                func.count(Entrega.id).label('qtd'),
                func.max(Entrega.data_envio).label('ultimo')
            )
            .group_by(Entrega.cliente)
            .all()
        )
        k_full = normalize_letters_key(cl.nome or '')
        k_first = normalize_first_token(cl.nome or '')

        tot, ultimo = 0, None
        for row in aggs:
            raw = (row.cli or '')
            if normalize_letters_key(raw) == k_full or normalize_first_token(raw) == k_first:
                tot += int(row.qtd or 0)
                if row.ultimo and (ultimo is None or row.ultimo > ultimo):
                    ultimo = row.ultimo

        return jsonify({
            "ok": True,
            "total_pedidos": int(tot or 0),
            "ultimo_uso": (br_date_ymd(ultimo) if ultimo else None)
        }), 200

    flash('Cliente atualizado!')
    return redirect(url_for('clientes'))


@app.route('/clientes/<int:id>/excluir', methods=['POST'])
def excluir_cliente(id):
    if not session.get('is_admin'):
        return redirect(url_for('login'))

    cl = Cliente.query.get_or_404(id)
    db.session.delete(cl)
    db.session.commit()

    if request.headers.get('X-Requested-With') == 'fetch' or request.args.get('format') == 'json' or (request.accept_mimetypes and request.accept_mimetypes.best == 'application/json'):
        return ("", 204)
    flash('Cliente excluído.')
    return redirect(url_for('clientes'))

# =========================================================
# TABELA DE PREÇOS & ROTAS
# =========================================================
@app.route('/precos-rotas', methods=['GET'], endpoint='precos_rotas')
def precos_rotas():
    if not session.get('is_admin'):
        return redirect(url_for('login'))

    try:
        bairros_rows = (
            Cliente.query
            .filter(Cliente.bairro_origem.isnot(None))
            .with_entities(Cliente.bairro_origem)
            .all()
        )
        bairros = sorted({(_norm(b[0])) for b in bairros_rows if _norm(b[0])})
    except Exception:
        bairros = []

    base_padrao = 12.0
    atualizado_em = _now_brt()
    per_km_val = get_per_km()

    return render_or_string(
        "precos_rotas.html",
        """
        <!doctype html><meta charset="utf-8">
        <h1>COOPEX — Tabela de Preços & Rotas</h1>
        <p>Base: R$ {{ '%.2f'|format(base_padrao) }}</p>
        <p>R$/km: <b>{{ '%.2f'|format(per_km) }}</b></p>
        <p>Atualizado em: {{ atualizado_em.strftime('%d/%m/%Y %H:%M') }}</p>
        """,
        base_padrao=base_padrao,
        atualizado_em=atualizado_em,
        bairros=bairros,
        per_km=per_km_val,
    )


@app.route("/api/precos", methods=["GET"], endpoint="api_list_precos")
def api_list_precos():
    if not session.get("is_admin"):
        return jsonify({"ok": False, "error": "unauthorized"}), 401

    q = request.args.get("q", "", type=str).strip()
    query = PrecoRota.query
    if q:
        like = f"%{q}%"
        query = query.filter(
            db.or_(
                PrecoRota.origem.ilike(like),
                PrecoRota.destino.ilike(like),
                func.cast(PrecoRota.valor, db.String).ilike(like),
            )
        )

    itens = [p.to_dict() for p in query.order_by(PrecoRota.origem.asc(), PrecoRota.destino.asc()).all()]

    try:
        bairros_rows = (
            Cliente.query
            .filter(Cliente.bairro_origem.isnot(None))
            .with_entities(Cliente.bairro_origem)
            .all()
        )
        bairros = sorted({(_norm(b[0])) for b in bairros_rows if _norm(b[0])})
    except Exception:
        bairros = sorted({p["origem"] for p in itens} | {p["destino"] for p in itens})

    return jsonify({
        "ok": True,
        "per_km": get_per_km(),
        "items": itens,
        "bairros": bairros,
    })


@app.route("/api/precos", methods=["POST"], endpoint="api_upsert_preco")
def api_upsert_preco():
    if not session.get("is_admin"):
        return jsonify({"ok": False, "error": "unauthorized"}), 401

    data = request.get_json(silent=True) or {}
    origem = _norm(data.get("origem"))
    destino = _norm(data.get("destino"))
    valor = data.get("valor", None)

    if not origem or not destino:
        return jsonify({"ok": False, "error": "Informe origem e destino."}), 400
    try:
        valor_f = float(valor)
    except Exception:
        return jsonify({"ok": False, "error": "Valor inválido."}), 400

    existente = (
        PrecoRota.query
        .filter(
            func.lower(PrecoRota.origem) == origem.lower(),
            func.lower(PrecoRota.destino) == destino.lower()
        )
        .first()
    )
    if existente:
        existente.origem = origem
        existente.destino = destino
        existente.valor = round(valor_f, 2)
        try:
            db.session.commit()
            return jsonify({"ok": True, "id": existente.id})
        except Exception as e:
            db.session.rollback()
            return jsonify({"ok": False, "error": f"Falha ao atualizar: {e}"}), 500
    else:
        novo = PrecoRota(origem=origem, destino=destino, valor=round(valor_f, 2))
        db.session.add(novo)
        try:
            db.session.commit()
            return jsonify({"ok": True, "id": novo.id})
        except IntegrityError:
            db.session.rollback()
            return jsonify({"ok": False, "error": "Par origem/destino já existe."}), 409
        except Exception as e:
            db.session.rollback()
            return jsonify({"ok": False, "error": f"Falha ao salvar: {e}"}), 500


@app.route("/api/precos/<int:item_id>", methods=["DELETE"], endpoint="api_delete_preco")
def api_delete_preco(item_id):
    if not session.get("is_admin"):
        return jsonify({"ok": False, "error": "unauthorized"}), 401

    it = PrecoRota.query.get(item_id)
    if not it:
        return jsonify({"ok": False, "error": "id não encontrado"}), 404
    db.session.delete(it)
    try:
        db.session.commit()
        return jsonify({"ok": True})
    except Exception as e:
        db.session.rollback()
        return jsonify({"ok": False, "error": f"Falha ao excluir: {e}"}), 500


@app.route("/api/precos/ajustes", methods=["PATCH"], endpoint="api_ajustes")
def api_ajustes():
    if not session.get("is_admin"):
        return jsonify({"ok": False, "error": "unauthorized"}), 401

    data = request.get_json(silent=True) or {}
    bairro = _norm(data.get("bairro", ""))
    delta = data.get("delta", None)
    global_delta = data.get("global_delta", None)

    changed = 0

    try:
        if bairro and delta is not None:
            try:
                dv = float(delta)
            except Exception:
                return jsonify({"ok": False, "error": "delta inválido."}), 400

            qs = PrecoRota.query.filter(
                db.or_(
                    func.lower(PrecoRota.origem) == bairro.lower(),
                    func.lower(PrecoRota.destino) == bairro.lower()
                )
            ).all()

            for it in qs:
                it.valor = round(float(it.valor) + dv, 2)
                changed += 1

            db.session.commit()
            return jsonify({"ok": True, "changed": changed})

        if global_delta is not None:
            try:
                gd = float(global_delta)
            except Exception:
                return jsonify({"ok": False, "error": "global_delta inválido."}), 400

            qs = PrecoRota.query.all()
            for it in qs:
                it.valor = round(float(it.valor) + gd, 2)
                changed += 1

            db.session.commit()
            return jsonify({"ok": True, "changed": changed})

        return jsonify({
            "ok": False,
            "error": "Nada a aplicar. Envie {bairro, delta} ou {global_delta}."
        }), 400

    except Exception as e:
        db.session.rollback()
        return jsonify({"ok": False, "error": f"Falha no ajuste: {e}"}), 500


@app.route("/api/perkm", methods=["POST"], endpoint="api_per_km")
def api_per_km():
    if not session.get("is_admin"):
        return jsonify({"ok": False, "error": "unauthorized"}), 401

    data = request.get_json(silent=True) or {}
    v = data.get("per_km", None)
    try:
        v = float(v)
    except Exception:
        return jsonify({"ok": False, "error": "per_km inválido."}), 400

    novo = set_per_km(v)
    return jsonify({"ok": True, "per_km": float(novo)})

# =========================================================
# TRAJETOS (HISTÓRICO POR COOPERADO / PERÍODO)
# =========================================================
@app.route('/trajetos')
def trajetos():
    if not session.get('is_admin'):
        return redirect(url_for('login'))

    # Limpa automaticamente trajetos com mais de 31 dias (sempre mantém último mês)
    try:
        limite_utc = datetime.utcnow() - timedelta(days=31)
        (
            Trajeto.query
            .filter(Trajeto.inicio < limite_utc)
            .delete(synchronize_session=False)
        )
        db.session.commit()
    except Exception:
        db.session.rollback()

    cooperados = Cooperado.query.order_by(Cooperado.nome).all()
    cooperado_id = request.args.get('cooperado_id', 'todos')
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')

    q = Trajeto.query.options(joinedload(Trajeto.cooperado))

    # Período padrão: últimos 30 dias em horário de Brasília
    hoje_brt = datetime.now(BRAZIL_TZ).date()
    if not data_inicio and not data_fim:
        di_default = hoje_brt - timedelta(days=29)
        di_utc, _ = local_date_window_to_utc_range(di_default)
        _, df_utc = local_date_window_to_utc_range(hoje_brt)
        q = q.filter(Trajeto.inicio >= di_utc, Trajeto.inicio <= df_utc)

        data_inicio = di_default.isoformat()
        data_fim = hoje_brt.isoformat()
    else:
        if data_inicio:
            di = datetime.strptime(data_inicio, "%Y-%m-%d").date()
            di_utc, _ = local_date_window_to_utc_range(di)
            q = q.filter(Trajeto.inicio >= di_utc)
        if data_fim:
            df = datetime.strptime(data_fim, "%Y-%m-%d").date()
            _, df_utc = local_date_window_to_utc_range(df)
            q = q.filter(Trajeto.inicio <= df_utc)

    if cooperado_id and cooperado_id != 'todos':
        try:
            q = q.filter(Trajeto.cooperado_id == int(cooperado_id))
        except ValueError:
            pass

    trajetos_list = q.order_by(Trajeto.inicio.desc()).limit(2000).all()

    # KPIs gerais
    total_km = sum((t.distancia_m or 0.0) for t in trajetos_list) / 1000.0
    total_horas = sum((t.duracao_s or 0) for t in trajetos_list) / 3600.0
    vel_media_geral = (total_km / total_horas) if total_horas > 0 else 0.0

    return render_template(
        'trajetos.html',
        trajetos=trajetos_list,
        cooperados=cooperados,
        cooperado_id=cooperado_id,
        data_inicio=data_inicio,
        data_fim=data_fim,
        total_km=total_km,
        total_horas=total_horas,
        vel_media_geral=vel_media_geral,
        to_brasilia=to_brasilia,
        now=lambda: datetime.now(BRAZIL_TZ),
    )

@app.route('/trajetos/exportar')
def trajetos_exportar():
    if not session.get('is_admin'):
        return redirect(url_for('login'))

    cooperado_id = request.args.get('cooperado_id', 'todos')
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')

    q = Trajeto.query.options(joinedload(Trajeto.cooperado))

    hoje_brt = datetime.now(BRAZIL_TZ).date()
    if not data_inicio and not data_fim:
        di_default = hoje_brt - timedelta(days=29)
        di_utc, _ = local_date_window_to_utc_range(di_default)
        _, df_utc = local_date_window_to_utc_range(hoje_brt)
        q = q.filter(Trajeto.inicio >= di_utc, Trajeto.inicio <= df_utc)
    else:
        if data_inicio:
            di = datetime.strptime(data_inicio, "%Y-%m-%d").date()
            di_utc, _ = local_date_window_to_utc_range(di)
            q = q.filter(Trajeto.inicio >= di_utc)
        if data_fim:
            df = datetime.strptime(data_fim, "%Y-%m-%d").date()
            _, df_utc = local_date_window_to_utc_range(df)
            q = q.filter(Trajeto.inicio <= df_utc)

    if cooperado_id and cooperado_id != 'todos':
        try:
            q = q.filter(Trajeto.cooperado_id == int(cooperado_id))
        except ValueError:
            pass

    trajetos_list = q.order_by(Trajeto.inicio.asc()).all()

    rows = []
    for t in trajetos_list:
        ini_local = to_brasilia(t.inicio) if t.inicio else None
        fim_local = to_brasilia(t.fim) if t.fim else None
        rows.append({
            'Cooperado': t.cooperado.nome if t.cooperado else '',
            'Início (Brasília)': ini_local.strftime('%d/%m/%Y %H:%M:%S') if ini_local else '',
            'Fim (Brasília)': fim_local.strftime('%d/%m/%Y %H:%M:%S') if fim_local else '',
            'Duração (min)': round((t.duracao_s or 0) / 60.0, 1),
            'Distância (km)': round((t.distancia_m or 0.0) / 1000.0, 3),
            'Velocidade média (km/h)': round(t.velocidade_media_kmh or 0.0, 1),
            'Origem (lat,lng)': (
                f"{t.origem_lat:.6f},{t.origem_lng:.6f}"
                if t.origem_lat is not None and t.origem_lng is not None
                else ''
            ),
            'Destino (lat,lng)': (
                f"{t.destino_lat:.6f},{t.destino_lng:.6f}"
                if t.destino_lat is not None and t.destino_lng is not None
                else ''
            ),
        })

    df_out = pd.DataFrame(rows)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        sheet = 'Trajetos'
        df_out.to_excel(writer, index=False, sheet_name=sheet)
        ws = writer.sheets[sheet]

        widths = [26, 22, 22, 14, 16, 22, 20, 20]
        for i, w in enumerate(widths[:len(df_out.columns)]):
            ws.set_column(i, i, w)

        money_fmt = writer.book.add_format({'num_format': '#,##0.000'})
        vel_fmt = writer.book.add_format({'num_format': '#,##0.0'})
        cols = list(df_out.columns)
        if 'Distância (km)' in cols:
            idx = cols.index('Distância (km)')
            ws.set_column(idx, idx, 16, money_fmt)
        if 'Velocidade média (km/h)' in cols:
            idx = cols.index('Velocidade média (km/h)')
            ws.set_column(idx, idx, 22, vel_fmt)

    output.seek(0)
    return send_file(output, download_name='trajetos.xlsx', as_attachment=True)

from flask import request, jsonify

@app.route('/mapa_motoboys')
def mapa_motoboys():
    if not session.get('is_admin'):
        return redirect(url_for('login'))

    cooperados = Cooperado.query.order_by(Cooperado.nome).all()
    motoboys_js = []

    for c in cooperados:
        if c.last_lat is not None and c.last_lng is not None:
            # ---- MONTA STATUS PRO MAPA ----
            # ajusta esses campos conforme o teu modelo:
            #   c.online -> app ligado?
            #   c.em_corrida ou c.ocupado -> está em entrega?
            online = bool(getattr(c, "online", False))
            em_corrida = bool(getattr(c, "em_corrida", False) or getattr(c, "ocupado", False))

            if not online:
                status = "offline"
            elif em_corrida:
                status = "em_corrida"
            else:
                status = "livre"

            is_online, idle_s, status_str = calc_status_cooperado(c)

            motoboys_js.append({
                "id": c.id,
                "nome": c.nome,
                "lat": float(c.last_lat),
                "lng": float(c.last_lng),
                "online": bool(is_online),
                "status": status_str,
                "idle_seconds": idle_s,
                "velocidade": float(getattr(c, "last_speed_kmh", 0) or 0),
                "ultima_atualizacao": (to_brasilia(c.last_ping).strftime('%d/%m %H:%M') if c.last_ping else ""),
                "endereco": getattr(c, "zona", None) or getattr(c, "bairro", None) or "",
                "observacao": getattr(c, "observacao", "") or ""
            })

    # 👇 Se for chamada via fetch (admin embutido) → JSON
    if request.headers.get('X-Requested-With') == 'fetch' or request.args.get('format') == 'json' or (request.accept_mimetypes and request.accept_mimetypes.best == 'application/json'):
        resp = jsonify(motoboys_js)
        resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        return resp

    # 👇 Se for acesso normal (mapa em tela cheia) → HTML usando esse mesmo motoboys_js
    return render_template('mapa_motoboys.html', motoboys_js=motoboys_js)


# =========================================================
# ENTREGAS: CADASTRAR / AGENDAR / EDITAR / EXCLUIR
# =========================================================
def _wants_json():
    """
    Decide se a resposta deve ser JSON (para AJAX / fetch).
    - ?format=json
    - request.is_json
    - Accept: application/json
    """
    try:
        if request.args.get('format') == 'json':
            return True
    except RuntimeError:
        pass

    try:
        if request.is_json:
            return True
        best = request.accept_mimetypes.best
        return best == 'application/json'
    except Exception:
        return False

def _parse_money_to_float(v) -> float:
    """
    Aceita:
      12.34
      "12,34"
      "R$ 12,34"
      "  12,34  "
    """
    if v is None:
        raise ValueError("valor ausente")

    if isinstance(v, (int, float)):
        return float(v)

    s = str(v).strip()
    if not s:
        raise ValueError("valor vazio")

    # remove R$, espaços e tudo que não for número, vírgula, ponto ou menos
    s = re.sub(r"[^\d,.\-]", "", s)

    # pt-BR: vírgula decimal
    # se vier "1.234,56" -> remove milhares e troca decimal
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    else:
        # se vier só vírgula: troca por ponto
        if "," in s:
            s = s.replace(",", ".")

    val = float(s)
    return val


@app.route("/api/entregas/<int:entrega_id>/valor", methods=["PATCH"])
def api_update_entrega_valor(entrega_id):
    if not session.get("is_admin"):
        return jsonify({"ok": False, "error": "unauthorized"}), 401

    e = Entrega.query.get_or_404(entrega_id)
    data = request.get_json(silent=True) or {}

    try:
        novo_valor = _parse_money_to_float(data.get("valor"))
    except Exception:
        return jsonify({"ok": False, "error": "Valor inválido."}), 400

    if novo_valor < 0:
        return jsonify({"ok": False, "error": "Valor não pode ser negativo."}), 400

    e.valor = float(novo_valor)
    db.session.commit()

    # Atualiza painéis em tempo real (se você usa isso)
    emitir_atualizacao_entrega(e, "editada")

    return jsonify({"ok": True, "id": e.id, "valor": float(e.valor)}), 200


@app.patch("/api/entregas/<int:entrega_id>/inline")
def api_update_entrega_inline(entrega_id):
    """Atualização inline (admin) para edição rápida na tabela.
    Aceita JSON com qualquer combinação:
      - valor (string/number)
      - cooperado_id (int ou '' para remover)
      - status (string)
      - status_pagamento (string)
    """
    if not session.get("is_admin") and not session.get("is_master"):
        return jsonify(ok=False, error="unauthorized"), 401

    e = Entrega.query.get_or_404(entrega_id)
    data = request.get_json(silent=True) or {}

    changed = False

    if "valor" in data:
        try:
            novo_valor = _parse_money_to_float(data.get("valor"))
            if novo_valor is not None:
                e.valor = float(novo_valor)
                changed = True
        except Exception:
            return jsonify(ok=False, error="valor inválido"), 400

    if "cooperado_id" in data:
        cid = data.get("cooperado_id")
        if cid in (None, "", 0, "0"):
            e.cooperado_id = None
            changed = True
        else:
            try:
                cid_int = int(cid)
            except Exception:
                return jsonify(ok=False, error="cooperado_id inválido"), 400
            coop = Cooperado.query.get(cid_int)
            if not coop:
                return jsonify(ok=False, error="cooperado não encontrado"), 404
            e.cooperado_id = cid_int
            changed = True

    if "status" in data:
        st = (data.get("status") or "").strip().lower()
        if st:
            e.status = st
            changed = True

    if "status_pagamento" in data:
        sp = (data.get("status_pagamento") or "").strip().lower()
        if sp:
            e.status_pagamento = sp
            changed = True

    if changed:
        db.session.commit()

    return jsonify(
        ok=True,
        entrega_id=e.id,
        valor=float(e.valor or 0),
        cooperado_id=e.cooperado_id,
        cooperado_nome=(e.cooperado.nome if getattr(e, "cooperado", None) else None),
        status=e.status,
        status_pagamento=e.status_pagamento,
    )


@app.route('/clonar_entrega/<int:id>', methods=['POST'])
def clonar_entrega(id):
    if not session.get('is_admin'):
        return redirect(url_for('login'))

    e = Entrega.query.get_or_404(id)
    nova = Entrega(
        cliente=e.cliente,
        bairro=e.bairro,
        valor=e.valor,
        data_envio=datetime.utcnow(),
        data_atribuida=None,
        cooperado_id=None,
        status='pendente',
        status_pagamento='pendente',
        pagamento=e.pagamento,
        recebido_por=None
    )
    db.session.add(nova)
    db.session.commit()

    msg = f'Entrega #{e.id} clonada em #{nova.id}. Edite para atribuir um cooperado.'
    flash(msg)

    if _wants_json():
        return jsonify(
            ok=True,
            message=msg,
            entrega={
                'id': nova.id,
                'origem_id': e.id,
                'cliente': nova.cliente,
                'bairro': nova.bairro,
                'valor': float(nova.valor or 0),
                'status': nova.status,
                'status_pagamento': nova.status_pagamento,
            }
        )

    return redirect_back_to_admin()


@app.route('/cadastrar_entrega', methods=['GET', 'POST'])
def cadastrar_entrega():
    if not session.get('is_admin'):
        return redirect(url_for('login'))

    cooperados = Cooperado.query.order_by(Cooperado.nome).all()
    clientes_lista = Cliente.query.order_by(Cliente.nome).all()

    if request.method == 'POST':
        cliente_nome = (request.form.get('cliente') or '').strip()
        bairro = request.form.get('bairro')
        valor = float(request.form.get('valor') or 0)
        cooperado_id = request.form.get('cooperado_id')
        pagamento = (request.form.get('pagamento') or '').strip()

        cliente_id_form = request.form.get('cliente_id', type=int)
        cli = None
        if cliente_id_form:
            cli = Cliente.query.get(cliente_id_form)
        if not cli and cliente_nome:
            cli = _find_cliente_by_nome(cliente_nome)

        entrega = Entrega(
            cliente=cliente_nome,
            bairro=bairro,
            valor=valor,
            data_envio=datetime.utcnow(),
            status_pagamento='pendente',
            status='pendente',
            pagamento=pagamento
        )

        if cli:
            entrega.cliente_id = cli.id

        if cooperado_id:
            entrega.cooperado_id = int(cooperado_id)
            entrega.data_atribuida = datetime.utcnow()

        db.session.add(entrega)

        if cooperado_id:
            ListaEspera.query.filter_by(cooperado_id=int(cooperado_id)).delete()

        db.session.commit()

        # DEBUG AQUI
        print("DEBUG_PAGAMENTO_ENTREGA", entrega.id, repr(entrega.pagamento))

        credito_consumido = 0.0
        erro_credito = False
        msg = 'Entrega cadastrada!'
        msg_category = 'info'

        # Tenta consumir crédito e mostra o resultado
        try:
            if pagamento_usa_credito(entrega.pagamento):
                valor_consumido = consumir_credito_em_entrega(entrega.id)
                credito_consumido = float(valor_consumido or 0.0)
                if credito_consumido > 0:
                    msg = (
                        f'Entrega cadastrada! Consumiu R$ {credito_consumido:.2f} '
                        f'de crédito do cliente.'
                    )
                    msg_category = 'success'
                else:
                    msg = (
                        'Entrega cadastrada! (nenhum crédito foi consumido para '
                        'este cliente).'
                    )
                    msg_category = 'info'
            else:
                msg = (
                    'Entrega cadastrada! (nenhum crédito foi consumido para '
                    'este cliente).'
                )
                msg_category = 'info'
        except Exception as ex:
            app.logger.exception(
                "Falha ao consumir crédito na entrega %s: %s", entrega.id, ex
            )
            erro_credito = True
            msg = (
                'Entrega cadastrada, mas houve erro ao tentar consumir crédito '
                'automaticamente.'
            )
            msg_category = 'warning'

        flash(msg, msg_category)

         # 🔴 EMITE PARA O PAINEL EM TEMPO REAL
        emitir_atualizacao_entrega(entrega, 'criada')

        if _wants_json():
            return jsonify(
                ok=True,
                message=msg,
                erro_credito=erro_credito,
                credito_consumido=credito_consumido,
                entrega_id=entrega.id,
                status=entrega.status,
                status_pagamento=entrega.status_pagamento,
                cooperado_id=entrega.cooperado_id,
            )

        return redirect_back_to_admin()

    return render_template('cadastrar_entrega.html', cooperados=cooperados, clientes=clientes_lista)


@app.route('/agendar_entrega', methods=['GET', 'POST'])
def agendar_entrega():
    if not session.get('is_admin'):
        return redirect(url_for('login'))

    cooperados = Cooperado.query.order_by(Cooperado.nome).all()
    clientes_lista = Cliente.query.order_by(Cliente.nome).all()

    if request.method == 'POST':
        cliente_nome = (request.form.get('cliente') or '').strip()
        bairro = request.form.get('bairro')
        valor = float(request.form.get('valor') or 0)
        data_str = request.form.get('data')  # 'YYYY-MM-DDTHH:MM'
        status_entrega = request.form.get('status_entrega')
        status_pagamento = request.form.get('status_pagamento')
        cooperado_id = request.form.get('cooperado_id')
        pagamento = (request.form.get('pagamento') or '').strip()

        data_envio = parse_local_datetime_to_utc_naive(data_str)

        cliente_id_form = request.form.get('cliente_id', type=int)
        cli = None
        if cliente_id_form:
            cli = Cliente.query.get(cliente_id_form)
        if not cli and cliente_nome:
            cli = _find_cliente_by_nome(cliente_nome)

        entrega = Entrega(
            cliente=cliente_nome,
            bairro=bairro,
            valor=valor,
            data_envio=data_envio,
            cooperado_id=int(cooperado_id) if cooperado_id else None,
            status=(status_entrega or 'pendente'),
            status_pagamento=(status_pagamento or 'pendente').lower(),
            pagamento=pagamento
        )

        if cli:
            entrega.cliente_id = cli.id

        db.session.add(entrega)

        if cooperado_id:
            ListaEspera.query.filter_by(cooperado_id=int(cooperado_id)).delete()

        db.session.commit()

        credito_consumido = 0.0
        erro_credito = False
        msg = 'Entrega agendada!'
        msg_category = 'info'

        # Tenta consumir crédito e mostra o resultado
        try:
            if pagamento_usa_credito(entrega.pagamento):
                valor_consumido = consumir_credito_em_entrega(entrega.id)
                credito_consumido = float(valor_consumido or 0.0)
                if credito_consumido > 0:
                    msg = (
                        f'Entrega agendada! Consumiu R$ {credito_consumido:.2f} '
                        f'de crédito do cliente.'
                    )
                    msg_category = 'success'
                else:
                    msg = (
                        'Entrega agendada! (nenhum crédito foi consumido para '
                        'este cliente).'
                    )
                    msg_category = 'info'
            else:
                msg = (
                    'Entrega agendada! (nenhum crédito foi consumido para '
                    'este cliente).'
                )
                msg_category = 'info'
        except Exception as ex:
            app.logger.exception(
                "Falha ao consumir crédito (agendada) na entrega %s: %s",
                entrega.id, ex
            )
            erro_credito = True
            msg = (
                'Entrega agendada, mas houve erro ao tentar consumir crédito '
                'automaticamente.'
            )
            msg_category = 'warning'

        flash(msg, msg_category)

        # 🔴 EMITE PARA O PAINEL EM TEMPO REAL (entrega agendada)
        emitir_atualizacao_entrega(entrega, 'criada')

        if _wants_json():
            return jsonify(
                ok=True,
                message=msg,
                erro_credito=erro_credito,
                credito_consumido=credito_consumido,
                entrega_id=entrega.id,
                status=entrega.status,
                status_pagamento=entrega.status_pagamento,
                cooperado_id=entrega.cooperado_id,
            )

        return redirect_back_to_admin()

    return render_template('agendar_entrega.html', cooperados=cooperados, clientes=clientes_lista)


@app.route('/editar_entrega/<int:id>', methods=['GET', 'POST'])
def editar_entrega(id):
    entrega = Entrega.query.get_or_404(id)
    cooperados = Cooperado.query.order_by(Cooperado.nome).all()
    is_admin = session.get('is_admin')

    if not is_admin and entrega.cooperado_id != session.get('user_id'):
        flash("Acesso não permitido.")
        return redirect(url_for('painel_cooperado'))

    if request.method == 'POST':
        if is_admin:
            novo_cliente_nome = (request.form.get('cliente') or '').strip()
            entrega.cliente = novo_cliente_nome
            entrega.bairro = request.form.get('bairro')

            try:
                entrega.valor = float(request.form.get('valor') or entrega.valor or 0)
            except Exception:
                entrega.valor = 0.0

            cliente_id_form = request.form.get('cliente_id', type=int)
            cli = None
            if cliente_id_form:
                cli = Cliente.query.get(cliente_id_form)
            if not cli and novo_cliente_nome:
                cli = _find_cliente_by_nome(novo_cliente_nome)
            entrega.cliente_id = cli.id if cli else None

            novo_coop_id = request.form.get('cooperado_id')
            if novo_coop_id:
                novo_coop_id = int(novo_coop_id)
                if entrega.cooperado_id != novo_coop_id:
                    entrega.cooperado_id = novo_coop_id
                    entrega.data_atribuida = datetime.utcnow()
                    ListaEspera.query.filter_by(cooperado_id=novo_coop_id).delete()
            else:
                entrega.cooperado_id = None

            entrega.status_pagamento = (
                request.form.get('status_pagamento')
                or entrega.status_pagamento
                or 'pendente'
            ).lower()
            entrega.status = request.form.get('status') or entrega.status
            entrega.recebido_por = request.form.get('recebido_por')
            entrega.pagamento = (
                request.form.get('pagamento') or entrega.pagamento or ''
            ).strip()

            db.session.commit()

            try:
                if pagamento_usa_credito(entrega.pagamento):
                    desfazer_consumo_credito_da_entrega(entrega.id)
                    consumir_credito_em_entrega(entrega.id)
                else:
                    if (entrega.credito_usado or 0) > 0:
                        desfazer_consumo_credito_da_entrega(entrega.id)

            except Exception as ex:
                app.logger.exception(
                    "Falha ao recalcular crédito na entrega %s: %s",
                    entrega.id, ex
                )

             # 🔴 EMITE PARA O PAINEL EM TEMPO REAL (edição)
            emitir_atualizacao_entrega(entrega, 'editada')

            flash('Entrega atualizada!')

            if _wants_json():
                return jsonify(
                    ok=True,
                    message='Entrega atualizada!',
                    entrega_id=entrega.id,
                    status=entrega.status,
                    status_pagamento=entrega.status_pagamento,
                    cooperado_id=entrega.cooperado_id,
                    cliente=entrega.cliente,
                    bairro=entrega.bairro,
                    valor=float(entrega.valor or 0),
                )

            return redirect_back_to_admin()

        else:
            entrega.status_pagamento = (
                request.form.get('status_pagamento')
                or entrega.status_pagamento
                or 'pendente'
            ).lower()
            entrega.status = request.form.get('status') or entrega.status
            entrega.recebido_por = request.form.get('recebido_por')
            db.session.commit()
            flash('Entrega atualizada!')

            if _wants_json():
                return jsonify(
                    ok=True,
                    message='Entrega atualizada!',
                    entrega_id=entrega.id,
                    status=entrega.status,
                    status_pagamento=entrega.status_pagamento,
                    recebido_por=entrega.recebido_por,
                )

            return redirect(url_for('painel_cooperado'))

    if is_admin:
        return render_template('editar_entrega.html', entrega=entrega, cooperados=cooperados)
    else:
        return render_template('editar_entrega_cooperado.html', entrega=entrega)


@app.post('/atribuir_cooperado/<int:id>')
def atribuir_cooperado(id):
    if not session.get('is_admin'):
        return redirect(url_for('login'))

    entrega = Entrega.query.get_or_404(id)
    coop_id = (request.form.get('cooperado_id') or '').strip()

    try:
        if coop_id:
            coop = Cooperado.query.get_or_404(int(coop_id))
            entrega.cooperado_id = coop.id
            entrega.data_atribuida = datetime.utcnow()

            # chave do fluxo: entrega atribuída, aguardando aceite do cooperado
            entrega.status_corrida = 'pendente'
        else:
            entrega.cooperado_id = None
            entrega.data_atribuida = None
            entrega.status_corrida = None

        db.session.commit()
        msg = 'Entrega atribuída com sucesso!'
        flash(msg, 'success')

        if _wants_json():
            return jsonify(
                ok=True,
                message=msg,
                entrega_id=entrega.id,
                cooperado_id=entrega.cooperado_id,
                status_corrida=entrega.status_corrida,
            )

    except Exception as e:
        db.session.rollback()
        msg = 'Erro ao atribuir entrega'
        flash(msg, 'danger')

        if _wants_json():
            return jsonify(ok=False, message=msg), 500

    return redirect(request.referrer or url_for('admin'))


@app.route('/excluir_entrega/<int:id>', methods=['POST'])
def excluir_entrega(id):
    if not session.get('is_admin'):
        return redirect(url_for('login'))

    entrega = Entrega.query.get_or_404(id)

    try:
        desfazer_consumo_credito_da_entrega(entrega.id)
    except Exception as ex:
        current_app.logger.exception(
            "Falha ao estornar crédito da entrega %s: %s", entrega.id, ex
        )

    try:
        db.session.execute(
            text("UPDATE credito_movimento SET entrega_id = NULL WHERE entrega_id = :eid"),
            {"eid": id}
        )
        db.session.execute(
            text("UPDATE entrega SET credito_mov_id = NULL WHERE id = :eid"),
            {"eid": id}
        )

        db.session.delete(entrega)
        db.session.commit()
        msg = 'Entrega excluída com sucesso.'
        flash(msg, 'success')

        if _wants_json():
            return jsonify(ok=True, message=msg, entrega_id=id)

    except IntegrityError:
        db.session.rollback()
        msg = 'Não foi possível excluir: há vínculos de crédito ativos.'
        flash(msg, 'danger')
        current_app.logger.exception("IntegrityError ao excluir entrega %s", id)

        if _wants_json():
            return jsonify(
                ok=False,
                message=msg,
                motivo='integrity'
            ), 400

    except Exception as e:
        db.session.rollback()
        msg = f'Erro ao excluir entrega: {e.__class__.__name__}'
        flash(msg, 'danger')
        current_app.logger.exception("Erro ao excluir entrega %s", id)

        if _wants_json():
            return jsonify(ok=False, message=msg), 500

    return redirect_back_to_admin()


@app.post('/entregas/<int:id>/marcar-pagamento')
def marcar_pagamento(id):
    if not session.get('is_admin'):
        return redirect(url_for('login'))

    e = Entrega.query.get_or_404(id)
    e.status_pagamento = "pago"
    db.session.commit()

    if _wants_json():
        return jsonify(
            ok=True,
            entrega_id=e.id,
            status_pagamento=e.status_pagamento,
        )

    return redirect_back_to_admin()


@app.post('/entregas/<int:id>/marcar-entregue')
def marcar_entregue(id):
    if not session.get('is_admin'):
        return redirect(url_for('login'))

    e = Entrega.query.get_or_404(id)
    e.status = "entregue"
    db.session.commit()

    if _wants_json():
        return jsonify(
            ok=True,
            entrega_id=e.id,
            status=e.status,
        )

    return redirect_back_to_admin()

# ================================
# INLINE_EDIT_VALOR_ENTREGA_ROUTE
# ================================
@app.post('/api/entregas/<int:id>/valor')
def api_atualizar_valor_entrega(id):
    if not session.get('is_admin'):
        return jsonify(ok=False, error="unauthorized"), 401

    e = Entrega.query.get_or_404(id)

    data = request.get_json(silent=True) or {}
    novo_valor_raw = data.get("valor", None)

    try:
        novo_valor = float(str(novo_valor_raw).replace(",", "."))
        if novo_valor < 0:
            return jsonify(ok=False, error="Valor não pode ser negativo."), 400
    except Exception:
        return jsonify(ok=False, error="Valor inválido."), 400

    # arredonda para 2 casas para ficar consistente
    novo_valor = round(novo_valor, 2)

    # se não mudou, só devolve ok
    atual = round(float(e.valor or 0), 2)
    if novo_valor == atual:
        return jsonify(ok=True, entrega_id=e.id, valor=atual, changed=False)

    e.valor = novo_valor
    db.session.add(e)
    db.session.commit()

    # Recalcula crédito se necessário (mesma lógica do editar_entrega)
    try:
        if pagamento_usa_credito(e.pagamento):
            # zera consumo antigo e tenta consumir de novo no novo valor
            desfazer_consumo_credito_da_entrega(e.id)
            consumir_credito_em_entrega(e.id)
        else:
            # se não usa crédito e tinha crédito usado, estorna
            if (e.credito_usado or 0) > 0:
                desfazer_consumo_credito_da_entrega(e.id)
    except Exception as ex:
        current_app.logger.exception("Falha ao recalcular crédito na entrega %s: %s", e.id, ex)
        # não bloqueia o update do valor, mas avisa no retorno
        # (você pode escolher retornar 500 se preferir)

    # Emite atualização em tempo real
    try:
        emitir_atualizacao_entrega(e, 'editada')
    except Exception:
        pass

    return jsonify(
        ok=True,
        entrega_id=e.id,
        valor=round(float(e.valor or 0), 2),
        status_pagamento=(e.status_pagamento or "").lower(),
        changed=True
    )

@app.get('/api/cliente/saldo')
@cliente_required
def api_cliente_saldo():
    cli = _cliente_atual()

    # garante que o saldo esteja correto (opcional mas recomendado)
    try:
        atualizar_saldo_credito_cliente(cli.id)
        cli = Cliente.query.get(cli.id)
    except Exception:
        pass

    return jsonify({
        "ok": True,
        "saldo": float(cli.saldo_atual or 0.0),
        "cliente": {
            "id": cli.id,
            "nome": cli.nome,
            "username": getattr(cli, "username", None),
            "telefone": getattr(cli, "telefone", None),
            "email": getattr(cli, "email", None),
        }
    })


# =========================================================
# CRÉDITOS (SUPERVISOR)
# =========================================================
@app.route('/creditos')
def creditos():
    if not session.get('is_admin'):
        return redirect(url_for('login'))

    # usado só pra pré-selecionar no select
    cliente_id = request.args.get('cliente_id', type=int)

    # todos os clientes para o formulário de lançamento
    clientes_form = Cliente.query.order_by(Cliente.nome.asc()).all()

    movimentos_por_cliente = {}
    saldos_por_cliente = {}
    creditos_originais_por_cliente = {}
    consumos_por_cliente = {}

    clientes_lista = []  # apenas os que terão histórico no acordeão

    for cli in clientes_form:
        movs = (
            CreditoMovimento.query
            .filter(CreditoMovimento.cliente_id == cli.id)
            .order_by(CreditoMovimento.data.asc(), CreditoMovimento.id.asc())
            .all()
        )

        # se não tem movimento e saldo_atual é zero/nulo, não aparece no histórico
        if not movs and not (cli.saldo_atual or 0):
            continue

        saldo = 0.0
        rows = []
        total_creditos_originais = 0.0  # só créditos "Crédito #...", sem estorno

        for mov in movs:
            valor = float(mov.valor or 0.0)
            ref = (mov.referencia or '').lower()

            if mov.tipo == 'credito':
                delta = valor
                # conta como "crédito lançado" só se NÃO for estorno
                if 'estorno' not in ref:
                    total_creditos_originais += valor
            elif mov.tipo == 'debito':
                delta = -valor
            else:
                delta = 0.0

            saldo_antes = saldo
            saldo_depois = saldo_antes + delta

            rows.append({
                "mov": mov,
                "saldo_antes": saldo_antes,
                "saldo_depois": saldo_depois,
            })

            saldo = saldo_depois

        movimentos_por_cliente[cli.id] = rows
        saldos_por_cliente[cli.id] = saldo
        creditos_originais_por_cliente[cli.id] = total_creditos_originais
        consumos_por_cliente[cli.id] = total_creditos_originais - saldo

        clientes_lista.append(cli)

    # totais globais (apenas clientes que aparecem no histórico)
    total_saldo = sum(saldos_por_cliente.values()) if saldos_por_cliente else 0.0
    total_creditos = sum(creditos_originais_por_cliente.values()) if creditos_originais_por_cliente else 0.0
    total_consumos = total_creditos - total_saldo

    if _wants_json():
        return jsonify(
            ok=True,
            total_saldo=total_saldo,
            total_creditos=total_creditos,
            total_consumos=total_consumos,
            clientes=[
                {
                    'id': cli.id,
                    'nome': cli.nome,
                    'saldo': float(saldos_por_cliente.get(cli.id, 0.0)),
                    'total_creditos': float(creditos_originais_por_cliente.get(cli.id, 0.0)),
                    'total_consumos': float(consumos_por_cliente.get(cli.id, 0.0)),
                }
                for cli in clientes_lista
            ]
        )

    return render_template(
        'creditos.html',
        # formulário de lançamento
        clientes_form=clientes_form,
        # clientes que aparecem no histórico
        clientes_lista=clientes_lista,
        cliente_id=cliente_id,
        movimentos_por_cliente=movimentos_por_cliente,
        saldos_por_cliente=saldos_por_cliente,
        creditos_originais_por_cliente=creditos_originais_por_cliente,
        consumos_por_cliente=consumos_por_cliente,
        total_saldo=total_saldo,
        total_creditos=total_creditos,
        total_consumos=total_consumos,
        request=request
    )


@app.route('/creditos/<int:cliente_id>/limpar', methods=['POST'])
def creditos_limpar_cliente(cliente_id):
    """
    Apaga TODOS os créditos e movimentos de um cliente
    e zera o saldo (mesmo que hoje esteja 0 ou diferente de 0).
    """
    if not session.get('is_admin'):
        return redirect(url_for('login'))

    cli = Cliente.query.get_or_404(cliente_id)

    # apaga todos os movimentos e créditos do cliente
    CreditoMovimento.query.filter_by(cliente_id=cliente_id).delete()
    Credito.query.filter_by(cliente_id=cliente_id).delete()

    cli.saldo_atual = 0.0
    db.session.add(cli)
    db.session.commit()

    msg = 'Histórico de créditos deste cliente foi totalmente limpo e saldo zerado.'
    flash(msg, 'success')

    if _wants_json():
        return jsonify(ok=True, message=msg, cliente_id=cliente_id)

    return redirect(url_for('creditos'))


@app.route('/creditos/novo', methods=['GET', 'POST'])
def creditos_novo():
    if not session.get('is_admin'):
        return redirect(url_for('login'))

    if request.method == 'POST':
        cliente_id = request.form.get('cliente_id', type=int)
        valor_bruto = request.form.get('valor', type=float)
        desconto_tipo = request.form.get('desconto_tipo', default='nenhum')
        desconto_valor = request.form.get('desconto_valor', type=float, default=0.0)
        motivo = request.form.get('motivo', default='')
        criado_por = session.get('user_nome', 'Supervisor')

        try:
            registrar_credito(
                cliente_id,
                valor_bruto,
                desconto_tipo,
                desconto_valor,
                motivo,
                criado_por
            )
            msg = 'Crédito criado com sucesso.'
            flash(msg, 'success')

            if _wants_json():
                return jsonify(ok=True, message=msg, cliente_id=cliente_id)

            return redirect(url_for('creditos', cliente_id=cliente_id))
        except Exception as e:
            db.session.rollback()
            current_app.logger.exception('Erro ao criar crédito')
            msg = f'Erro ao criar crédito: {e.__class__.__name__}'
            flash(msg, 'danger')

            if _wants_json():
                return jsonify(ok=False, message=msg), 500

    return render_template('credito_form.html')


@app.route('/creditos/<int:credito_id>/editar', methods=['GET', 'POST'])
def creditos_editar(credito_id):
    if not session.get('is_admin'):
        return redirect(url_for('login'))

    cred = Credito.query.get_or_404(credito_id)

    if request.method == 'POST':
        valor_bruto = request.form.get('valor', type=float, default=cred.valor_bruto)
        desconto_tipo = request.form.get('desconto_tipo', default=cred.desconto_tipo or 'nenhum')
        desconto_valor = request.form.get('desconto_valor', type=float, default=cred.desconto_valor or 0.0)
        motivo = request.form.get('motivo', default=cred.motivo or '')

        try:
            editar_credito(
                credito_id=credito_id,
                valor_bruto=valor_bruto,
                desconto_tipo=desconto_tipo,
                desconto_valor=desconto_valor,
                motivo=motivo,
            )
            msg = 'Crédito atualizado.'
            flash(msg, 'success')

            if _wants_json():
                cred_atual = Credito.query.get(credito_id)
                return jsonify(
                    ok=True,
                    message=msg,
                    credito_id=cred_atual.id,
                    cliente_id=cred_atual.cliente_id,
                    valor_final=float(cred_atual.valor_final or 0.0),
                )

            return redirect(url_for('creditos', cliente_id=cred.cliente_id))

        except Exception as e:
            db.session.rollback()
            current_app.logger.exception('Erro ao editar crédito')
            msg = f'Erro ao editar crédito: {e.__class__.__name__}'
            flash(msg, 'danger')

            if _wants_json():
                return jsonify(ok=False, message=msg), 500

    return render_template('credito_form.html', credito=cred)


@app.route('/creditos/<int:id>/excluir', methods=['POST'])
def creditos_excluir(id):
    if not session.get('is_admin'):
        return redirect(url_for('login'))

    c = Credito.query.get_or_404(id)
    cliente_id = c.cliente_id
    # remove movimentos ligados a este crédito
    CreditoMovimento.query.filter_by(credito_id=c.id).delete()
    db.session.delete(c)
    db.session.commit()

    atualizar_saldo_credito_cliente(cliente_id)
    msg = 'Crédito excluído e saldo recalculado.'
    flash(msg, 'success')

    if _wants_json():
        return jsonify(ok=True, message=msg, credito_id=id, cliente_id=cliente_id)

    return redirect(url_for('creditos', cliente_id=cliente_id))


@app.route('/creditos/exportar')
def creditos_exportar():
    if not session.get('is_admin'):
        return redirect(url_for('login'))

    cliente_id = request.args.get('cliente_id', type=int)
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')

    q = (db.session.query(
            Credito.id.label('id'),
            Cliente.nome.label('cliente'),
            Credito.valor_bruto.label('valor_bruto'),
            Credito.desconto_tipo.label('desconto_tipo'),
            Credito.desconto_valor.label('desconto_valor'),
            Credito.valor_final.label('valor_final'),
            Credito.motivo.label('motivo'),
            Credito.saldo_antes.label('saldo_antes'),
            Credito.saldo_depois.label('saldo_depois'),
            Credito.criado_por.label('criado_por'),
            Credito.criado_em.label('criado_em'),
        )
        .join(Cliente, Cliente.id == Credito.cliente_id)
    )

    if cliente_id:
        q = q.filter(Credito.cliente_id == cliente_id)
    if data_inicio:
        di = datetime.strptime(data_inicio, "%Y-%m-%d")
        q = q.filter(Credito.criado_em >= di)
    if data_fim:
        df = datetime.strptime(data_fim, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
        q = q.filter(Credito.criado_em <= df)

    q = q.order_by(Credito.criado_em.asc())

    rows = []
    for r in q.all():
        dt_local = to_brasilia(r.criado_em)
        rows.append({
            'Data': dt_local.strftime('%d/%m/%Y %H:%M') if dt_local else '',
            'Cliente': r.cliente,
            'Valor Bruto': float(r.valor_bruto or 0),
            'Desconto Tipo': r.desconto_tipo or 'nenhum',
            'Desconto Valor': float(r.desconto_valor or 0),
            'Valor Final': float(r.valor_final or 0),
            'Motivo': r.motivo or '',
            'Saldo Antes': float(r.saldo_antes or 0),
            'Saldo Depois': float(r.saldo_depois or 0),
            'Criado Por': r.criado_por or '',
            'ID Crédito': int(r.id),
        })

    df_out = pd.DataFrame(rows)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        sheet = 'Créditos'
        df_out.to_excel(writer, index=False, sheet_name=sheet)
        ws = writer.sheets[sheet]

        widths = [20, 28, 14, 16, 16, 14, 30, 14, 14, 16, 12]
        for i, w in enumerate(widths[:len(df_out.columns)]):
            ws.set_column(i, i, w)

        money_fmt = writer.book.add_format({'num_format': '#,##0.00'})
        for col_name in ['Valor Bruto', 'Desconto Valor', 'Valor Final', 'Saldo Antes', 'Saldo Depois']:
            if col_name in df_out.columns:
                idx = list(df_out.columns).index(col_name)
                ws.set_column(idx, idx, None, money_fmt)

    output.seek(0)
    return send_file(output, download_name='creditos.xlsx', as_attachment=True)


@app.route('/creditos/cadastrar', methods=['POST'])
def creditos_cadastrar():
    """
    Atalho mais simples para lançar um crédito via POST.

    Espera no formulário (ou JSON):
      - cliente_id
      - valor
      - desconto_tipo  (opcional, default 'nenhum')
      - desconto_valor (opcional, default 0)
      - motivo         (opcional)

    Usa a mesma lógica de registrar_credito() e depois redireciona
    para a tela /creditos já focada no cliente.
    """
    if not session.get('is_admin'):
        return redirect(url_for('login'))

    # Pode vir via form ou JSON
    data = request.form or (request.get_json(silent=True) or {})

    cliente_id = data.get('cliente_id', type=int) if hasattr(data, 'get') else int(data.get('cliente_id') or 0)
    valor_bruto = data.get('valor') or data.get('valor_bruto') or 0
    desconto_tipo = (data.get('desconto_tipo') or 'nenhum').strip()
    desconto_valor = data.get('desconto_valor') or 0
    motivo = (data.get('motivo') or '').strip()
    criado_por = session.get('user_nome', 'Supervisor')

    try:
        valor_bruto = float(valor_bruto)
    except Exception:
        valor_bruto = 0.0

    try:
        desconto_valor = float(desconto_valor)
    except Exception:
        desconto_valor = 0.0

    if not cliente_id or valor_bruto <= 0:
        msg = 'Informe cliente e um valor de crédito maior que zero.'
        if _wants_json():
            return jsonify(ok=False, message=msg), 400
        flash(msg, 'danger')
        return redirect(url_for('creditos'))

    try:
        registrar_credito(
            cliente_id=cliente_id,
            valor_bruto=valor_bruto,
            desconto_tipo=desconto_tipo,
            desconto_valor=desconto_valor,
            motivo=motivo,
            criado_por=criado_por
        )
        msg = 'Crédito cadastrado com sucesso.'
        if _wants_json():
            return jsonify(ok=True, message=msg, cliente_id=cliente_id)

        flash(msg, 'success')
        return redirect(url_for('creditos', cliente_id=cliente_id))

    except Exception as e:
        db.session.rollback()
        current_app.logger.exception('Erro ao cadastrar crédito')
        msg = f'Erro ao cadastrar crédito: {e.__class__.__name__}'
        if _wants_json():
            return jsonify(ok=False, message=msg), 500

        flash(msg, 'danger')
        return redirect(url_for('creditos'))
        

@app.route('/cliente/<int:cliente_id>/credito')
def cliente_credito(cliente_id):
    if not session.get('is_admin'):
        return redirect(url_for('login'))

    cli = Cliente.query.get_or_404(cliente_id)
    movs = (
        CreditoMovimento.query
        .filter(CreditoMovimento.cliente_id == cliente_id)
        .order_by(CreditoMovimento.criado_em.desc())
        .all()
    )

    total_creditos = sum(float(m.valor or 0) for m in movs if m.tipo == 'credito')
    total_debitos = sum(float(m.valor or 0) for m in movs if m.tipo == 'debito')
    saldo_atual = float(cli.saldo_atual or 0)

    if _wants_json():
        return jsonify(
            ok=True,
            cliente={
                'id': cli.id,
                'nome': cli.nome,
                'telefone': cli.telefone,
            },
            saldo_atual=saldo_atual,
            total_creditos=total_creditos,
            total_debitos=total_debitos,
            movimentos=[
                {
                    'id': m.id,
                    'tipo': m.tipo,
                    'valor': float(m.valor or 0.0),
                    'referencia': m.referencia,
                    'entrega_id': m.entrega_id,
                    'credito_id': m.credito_id,
                    'criado_em': to_brasilia(m.criado_em).isoformat()
                    if m.criado_em else None,
                }
                for m in movs
            ]
        )

    return render_or_string("credito_cliente.html", """
<!doctype html>
<html lang="pt-BR">
<head>
  <meta charset="utf-8">
  <title>Extrato de Crédito — {{ cliente.nome }}</title>
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <style>
    body{font-family:system-ui,-apple-system,Segoe UI,Roboto,Arial,sans-serif;margin:0;background:#f3f4ff;color:#0f172a}
    .wrap{max-width:1100px;margin:0 auto;padding:18px}
    .card{background:#ffffff;border:1px solid #d0ddff;border-radius:14px;padding:14px 16px;box-shadow:0 6px 20px rgba(15,23,42,.08)}
    h1{margin:0 0 6px;font-size:1.4rem}
    .sub{font-size:.9rem;color:#64748b;margin-bottom:10px}
    .chips{display:flex;flex-wrap:wrap;gap:8px;margin:8px 0 14px}
    .chip{border-radius:999px;padding:6px 10px;font-size:.8rem;font-weight:800;border:1px solid #d0ddff;background:#e5edff;color:#1e3a8a}
    .chip.good{background:#dcfce7;border-color:#bbf7d0;color:#166534}
    .chip.bad{background:#fee2e2;border-color:#fecaca;color:#b91c1c}
    .table-wrap{overflow:auto;border-radius:12px;border:1px solid #d0ddff;max-height:520px;background:#fff}
    table{width:100%;border-collapse:collapse;font-size:13.5px}
    th,td{padding:8px;border-bottom:1px solid #e2e8f0}
    th{position:sticky;top:0;background:#1e3a8a;color:#e5edff;text-align:left;z-index:1}
    tbody tr:nth-child(even) td{background:#f8fafc}
    .money{font-weight:900}
    .tag-credito{color:#16a34a;font-weight:700}
    .tag-debito{color:#dc2626;font-weight:700}
  </style>
</head>
<body>
  <div class="wrap">
    <div class="card">
      <h1>Extrato de crédito — {{ cliente.nome }}</h1>
      <div class="sub">
        Telefone:
        {% if cliente.telefone %}
          {{ cliente.telefone }}
        {% else %}
          <span style="opacity:.6">não informado</span>
        {% endif %}
      </div>

      <div class="chips">
        <span class="chip">
          Saldo atual:
          <span class="money" style="margin-left:6px">
            R$ {{ '%.2f'|format(saldo_atual)|replace('.', ',') }}
          </span>
        </span>
        <span class="chip good">
          Total créditos:
          R$ {{ '%.2f'|format(total_creditos)|replace('.', ',') }}
        </span>
        <span class="chip bad">
          Total débitos:
          R$ {{ '%.2f'|format(total_debitos)|replace('.', ',') }}
        </span>
        <span class="chip">
          Movimentos: {{ movs|length }}
        </span>
      </div>

      <div class="table-wrap">
        <table>
          <thead>
            <tr>
              <th>Data</th>
              <th>Tipo</th>
              <th>Descrição</th>
              <th>Valor</th>
              <th>Ref.</th>
            </tr>
          </thead>
          <tbody>
            {% for m in movs %}
              <tr>
                <td>
                  {% if m.criado_em %}
                    {{ to_brasilia(m.criado_em).strftime('%d/%m/%Y %H:%M') }}
                  {% else %}
                    -
                  {% endif %}
                </td>
                <td>
                  {% if m.tipo == 'credito' %}
                    <span class="tag-credito">CRÉDITO</span>
                  {% elif m.tipo == 'debito' %}
                    <span class="tag-debito">DÉBITO</span>
                  {% else %}
                    {{ m.tipo }}
                  {% endif %}
                </td>
                <td>{{ m.referencia or '-' }}</td>
                <td>
                  R$ {{ '%.2f'|format(m.valor or 0.0)|replace('.', ',') }}
                </td>
                <td>
                  {% if m.entrega_id %}
                    Entrega #{{ m.entrega_id }}
                  {% elif m.credito_id %}
                    Crédito #{{ m.credito_id }}
                  {% else %}
                    -
                  {% endif %}
                </td>
              </tr>
            {% else %}
              <tr>
                <td colspan="5" style="text-align:center;padding:12px;color:#6b7280;">
                  Nenhuma movimentação encontrada.
                </td>
              </tr>
            {% endfor %}
          </tbody>
        </table>
      </div>

      <div style="margin-top:10px;font-size:.8rem;color:#6b7280;">
        <a href="{{ url_for('creditos', cliente_id=cliente.id) }}">&larr; Voltar à tela de créditos</a>
      </div>
    </div>
  </div>
</body>
</html>
""", cliente=cli, movs=movs,
           saldo_atual=saldo_atual,
           total_creditos=total_creditos,
           total_debitos=total_debitos,
           to_brasilia=to_brasilia)

@app.route('/creditos/movimento/novo', methods=['POST'])
def credmov_novo():
    if not session.get('is_admin'):
        return redirect(url_for('login'))

    cliente_id = request.form.get('cliente_id', type=int)
    credito_id = request.form.get('credito_id', type=int)
    entrega_id = request.form.get('entrega_id', type=int)
    tipo_raw = request.form.get('tipo', default=TIPO_AJUSTE)
    valor = abs(request.form.get('valor', type=float, default=0.0))
    referencia = request.form.get('referencia', default='')

    try:
        # Aplica o efeito no saldo com base no tipo informado
        delta = _delta_saldo_tipo_mov(tipo_raw, valor)
        if abs(delta) > 1e-7:
            atualizar_saldo_cliente(cliente_id, delta)

        # Registra o movimento com o mesmo "tipo" recebido
        registrar_movimento(
            cliente_id, tipo_raw, valor,
            referencia=referencia,
            credito_id=credito_id,
            entrega_id=entrega_id
        )
        db.session.commit()
        msg = 'Movimento registrado.'
        flash(msg, 'success')

        if _wants_json():
            return jsonify(
                ok=True,
                message=msg,
                cliente_id=cliente_id,
            )

    except Exception as e:
        db.session.rollback()
        current_app.logger.exception('Erro ao criar movimento')
        msg = f'Erro ao criar movimento: {e.__class__.__name__}'
        flash(msg, 'danger')

        if _wants_json():
            return jsonify(ok=False, message=msg), 500

    return redirect(url_for('creditos', cliente_id=cliente_id))


@app.route('/creditos/movimento/<int:mov_id>/editar', methods=['GET', 'POST'])
def credmov_editar(mov_id):
    if not session.get('is_admin'):
        return redirect(url_for('login'))

    mov = CreditoMovimento.query.get_or_404(mov_id)

    if request.method == 'POST':
        # Pode vir 'ENTRADA'/'CONSUMO'/'AJUSTE' do formulário
        # ou 'credito'/'debito' (valor já salvo)
        novo_tipo_raw = (request.form.get('tipo') or mov.tipo)
        novo_valor = abs(request.form.get('valor', type=float, default=mov.valor))
        nova_ref = request.form.get('referencia', default=mov.referencia)

        try:
            # 1) Remove o efeito antigo do saldo
            delta_antigo = _delta_saldo_tipo_mov(mov.tipo, mov.valor)
            if abs(delta_antigo) > 1e-7:
                atualizar_saldo_cliente(mov.cliente_id, -delta_antigo)

            # 2) Aplica o efeito novo
            delta_novo = _delta_saldo_tipo_mov(novo_tipo_raw, novo_valor)
            if abs(delta_novo) > 1e-7:
                atualizar_saldo_cliente(mov.cliente_id, delta_novo)

            # 3) Normaliza e grava o tipo em 'credito' / 'debito' na tabela
            tipo_up = (novo_tipo_raw or '').upper()
            if tipo_up in (TIPO_ENTRADA, TIPO_AJUSTE, 'CREDITO'):
                mov.tipo = 'credito'
            elif tipo_up in (TIPO_CONSUMO, 'DEBITO', 'DÉBITO'):
                mov.tipo = 'debito'
            else:
                mov.tipo = 'credito'

            mov.valor = novo_valor
            mov.referencia = (nova_ref or '')[:120]
            db.session.add(mov)
            db.session.commit()
            msg = 'Movimento atualizado.'
            flash(msg, 'success')

            if _wants_json():
                return jsonify(
                    ok=True,
                    message=msg,
                    movimento_id=mov.id,
                    cliente_id=mov.cliente_id,
                )

        except Exception as e:
            db.session.rollback()
            current_app.logger.exception('Erro ao editar movimento')
            msg = f'Erro ao editar movimento: {e.__class__.__name__}'
            flash(msg, 'danger')

            if _wants_json():
                return jsonify(ok=False, message=msg), 500

        return redirect(url_for('creditos', cliente_id=mov.cliente_id))

    return render_template('credmov_form.html', movimento=mov)


@app.route('/creditos/movimento/<int:mov_id>/excluir', methods=['POST'])
def credmov_excluir(mov_id):
    if not session.get('is_admin'):
        return redirect(url_for('login'))

    mov = CreditoMovimento.query.get_or_404(mov_id)
    try:
        # Estorna o efeito desse movimento no saldo
        delta = _delta_saldo_tipo_mov(mov.tipo, mov.valor)
        if abs(delta) > 1e-7:
            atualizar_saldo_cliente(mov.cliente_id, -delta)

        # Se estiver vinculado a entrega, limpa o vínculo
        if mov.entrega_id:
            try:
                db.session.execute(
                    text("UPDATE entrega SET credito_mov_id = NULL WHERE id = :eid"),
                    {"eid": mov.entrega_id}
                )
            except Exception:
                pass

        cliente_id = mov.cliente_id
        db.session.delete(mov)
        db.session.commit()
        msg = 'Movimento excluído.'
        flash(msg, 'success')

        if _wants_json():
            return jsonify(ok=True, message=msg, cliente_id=cliente_id)

    except IntegrityError:
        db.session.rollback()
        msg = 'Não é possível excluir o movimento (vínculos).'
        flash(msg, 'danger')

        if _wants_json():
            return jsonify(ok=False, message=msg, motivo='integrity'), 400

    except Exception as e:
        db.session.rollback()
        current_app.logger.exception('Erro ao excluir movimento')
        msg = f'Erro ao excluir movimento: {e.__class__.__name__}'
        flash(msg, 'danger')

        if _wants_json():
            return jsonify(ok=False, message=msg), 500

    return redirect(url_for('creditos', cliente_id=mov.cliente_id))

# =========================================================
# JSON DO PAINEL DO COOPERADO
# =========================================================
@app.post('/cooperado/toggle_pagamento/<int:id>')
def toggle_pagamento(id):
    e = Entrega.query.get_or_404(id)
    _assert_entrega_do_cooperado(e)
    atual = (e.status_pagamento or 'pendente').lower()
    novo = 'pago' if atual != 'pago' else 'pendente'
    e.status_pagamento = novo
    db.session.commit()
    return jsonify(ok=True, status_pagamento=novo)


@app.get('/cooperado/api/ganhos')
def api_ganhos():
    if session.get('user_id') is None or session.get('is_admin'):
        return jsonify(ok=False, error='unauthorized'), 401

    cooperado_id = int(session.get('user_id'))
    hoje_local = datetime.now(BRAZIL_TZ).date()
    ano = request.args.get('ano', type=int) or hoje_local.year
    mes = request.args.get('mes', type=int) or hoje_local.month

    # janela do mês (em BRT) -> UTC range
    first = date(ano, mes, 1)
    # último dia do mês
    if mes == 12:
        last = date(ano + 1, 1, 1) - timedelta(days=1)
    else:
        last = date(ano, mes + 1, 1) - timedelta(days=1)

    ini_utc, _ = local_date_window_to_utc_range(first)
    _, fim_utc = local_date_window_to_utc_range(last)

    q = Entrega.query.filter(
        Entrega.cooperado_id == cooperado_id,
        Entrega.data_envio >= ini_utc,
        Entrega.data_envio <= fim_utc,
    )

    entregas = q.all()
    total_mes = sum(float(e.valor or 0) for e in entregas)
    total_pago_mes = sum(float(e.valor or 0) for e in entregas if (e.status_pagamento or '').lower() == 'pago')
    total_pendente_mes = max(0.0, total_mes - total_pago_mes)

    # ano atual
    first_y = date(ano, 1, 1)
    last_y = date(ano, 12, 31)
    ini_y, _ = local_date_window_to_utc_range(first_y)
    _, fim_y = local_date_window_to_utc_range(last_y)

    qy = Entrega.query.filter(
        Entrega.cooperado_id == cooperado_id,
        Entrega.data_envio >= ini_y,
        Entrega.data_envio <= fim_y,
    )
    ent_ano = qy.all()
    total_ano = sum(float(e.valor or 0) for e in ent_ano)
    total_pago_ano = sum(float(e.valor or 0) for e in ent_ano if (e.status_pagamento or '').lower() == 'pago')
    total_pendente_ano = max(0.0, total_ano - total_pago_ano)

    return jsonify(ok=True,
                   ano=ano, mes=mes,
                   total_mes=round(total_mes, 2),
                   pago_mes=round(total_pago_mes, 2),
                   pendente_mes=round(total_pendente_mes, 2),
                   total_ano=round(total_ano, 2),
                   pago_ano=round(total_pago_ano, 2),
                   pendente_ano=round(total_pendente_ano, 2),
                   qtd_mes=len(entregas),
                   qtd_ano=len(ent_ano))



@app.post('/cooperado/marcar_entregue/<int:id>')
def cooperado_marcar_entregue(id):
    """Marca entrega como recebida/entregue.
    Agora aceita:
      - JSON: {recebido_por: "..."} (compatível com o que já existia)
      - multipart/form-data: recebido_por (opcional) + foto (opcional)
    Regra: precisa ter **nome** OU **foto**.
    """
    e = Entrega.query.get_or_404(id)
    _assert_entrega_do_cooperado(e)

    recebido_por = ''
    foto_fs = None

    # 1) Se veio multipart (FormData), pega do form/files
    # OBS: mesmo sem arquivo, o browser envia multipart e request.files pode vir vazio,
    # então usamos mimetype pra decidir.
    if (request.mimetype or '').startswith('multipart/form-data'):
        recebido_por = (request.form.get('recebido_por') or '').strip()
        foto_fs = request.files.get('foto')
    else:
        # 2) Compatibilidade com JSON antigo
        payload = request.get_json(silent=True) or {}
        recebido_por = (payload.get('recebido_por') or '').strip()

    if not recebido_por and not foto_fs:
        return jsonify(ok=False, error='Informe o nome de quem recebeu OU envie uma foto.'), 400

    # salva foto (se veio)
    if foto_fs and getattr(foto_fs, "filename", ""):
        try:
            _salvar_comprovante(e.id, foto_fs)
        except Exception:
            return jsonify(ok=False, error='Não foi possível salvar a foto agora.'), 500

    e.status = 'recebido'
    e.recebido_por = recebido_por or (e.recebido_por or None)
    db.session.commit()
    return jsonify(ok=True, tem_foto=comprovante_existe(e.id))


@app.get('/cooperado/api/entrega_atribuida')
def api_entrega_atribuida():
    if session.get('user_id') is None or session.get('is_admin'):
        return jsonify({'tem': False}), 401

    cooperado_id = session.get('user_id')

    entrega = (
        Entrega.query
        .filter(
            Entrega.cooperado_id == cooperado_id,
            # só entregas que ainda não foram concluídas
            (Entrega.status == None) |
            (~func.lower(Entrega.status).in_(['recebido', 'entregue'])),
            # e que ainda estão pendentes ou recém aceitas
            (Entrega.status_corrida == None) |
            (Entrega.status_corrida.in_(['pendente', 'aceita']))
        )
        .order_by(Entrega.data_atribuida.desc(), Entrega.data_envio.desc())
        .first()
    )

    if not entrega:
        return jsonify({'tem': False})

    return jsonify({
        'tem': True,
        'id': entrega.id,
        'cliente': entrega.cliente,
        'valor': float(entrega.valor or 0),
        'status_corrida': entrega.status_corrida,
    })


# =========================================================
# ESTATÍSTICAS (ADMIN MASTER)
# =========================================================
@app.route('/estatisticas_cooperado')
@master_required
def estatisticas_cooperado():
    cooperados = Cooperado.query.order_by(Cooperado.nome).all()
    cooperado_id = request.args.get('cooperado_id', 'todos')
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')
    status_pagamento = request.args.get('status_pagamento', 'todos')
    cliente = (request.args.get('cliente') or '').strip()

    query = Entrega.query
    if cooperado_id != 'todos':
        query = query.filter(Entrega.cooperado_id == int(cooperado_id))
    if data_inicio:
        di = datetime.strptime(data_inicio, "%Y-%m-%d").date()
        inicio_utc, _ = local_date_window_to_utc_range(di)
        query = query.filter(Entrega.data_envio >= inicio_utc)
    if data_fim:
        df_ = datetime.strptime(data_fim, "%Y-%m-%d").date()
        _, fim_utc = local_date_window_to_utc_range(df_)
        query = query.filter(Entrega.data_envio <= fim_utc)
    if status_pagamento and status_pagamento != 'todos':
        if status_pagamento == 'pago':
            query = query.filter(func.lower(Entrega.status_pagamento) == 'pago')
        elif status_pagamento == 'pendente':
            query = query.filter(
                (Entrega.status_pagamento == None) |
                (func.lower(Entrega.status_pagamento) == 'pendente')
            )
    if cliente:
        like = f"%{cliente.lower()}%"
        query = query.filter(func.lower(Entrega.cliente).like(like))

    entregas = query.options(joinedload(Entrega.cooperado)).order_by(Entrega.data_envio.asc()).all()

    total = len(entregas)
    pagas = len([e for e in entregas if (e.status_pagamento or '').lower() == 'pago'])
    pendentes = total - pagas
    total_valor = sum(float(e.valor or 0) for e in entregas)
    ticket_medio = (total_valor / total) if total > 0 else 0.0

    cont_dias = Counter()
    for e in entregas:
        dt_local = to_brasilia(e.data_envio)
        if dt_local:
            cont_dias[dt_local.date()] += 1

    dia_top = {"data": None, "qtd": 0, "nome": "-"}
    if cont_dias:
        d, qtd = cont_dias.most_common(1)[0]
        dia_top = {
            "data": d.strftime('%Y-%m-%d'),
            "qtd": qtd,
            "nome": f"{d.strftime('%d/%m/%Y')} ({qtd})"
        }

    cont_horas = Counter()
    for e in entregas:
        dt_local = to_brasilia(e.data_envio)
        if dt_local:
            cont_horas[dt_local.strftime('%H:00')] += 1
    hora_pico = cont_horas.most_common(1)[0][0] if cont_horas else "-"
    horas_pico_top3 = [f"{h} ({q})" for h, q in cont_horas.most_common(3)]

    cont_pgto = Counter([e.pagamento for e in entregas if e.pagamento])
    pgto_top = cont_pgto.most_common(1)[0][0] if cont_pgto else "-"

    mapa_coop = defaultdict(lambda: {"qtd": 0, "total": 0.0})
    total_geral_periodo = 0.0
    for e in entregas:
        nm = e.cooperado.nome if e.cooperado else "Sem Cooperado"
        mapa_coop[nm]["qtd"] += 1
        mapa_coop[nm]["total"] += float(e.valor or 0)
        total_geral_periodo += float(e.valor or 0)

    ranking_cooperados = []
    for nome, dct in mapa_coop.items():
        percent = (dct["total"] / total_geral_periodo * 100.0) if total_geral_periodo > 0 else 0.0
        ranking_cooperados.append({
            "nome": nome,
            "qtd": dct["qtd"],
            "total_valor": round(dct["total"], 2),
            "percent": percent
        })
    ranking_cooperados.sort(key=lambda x: x["total_valor"], reverse=True)

    cont_bairros = Counter([e.bairro for e in entregas if e.bairro])
    ranking_bairros = [{"bairro": b, "qtd": q} for b, q in cont_bairros.most_common()]

    nomes_clientes = {e.cliente for e in entregas if e.cliente}
    if nomes_clientes:
        clientes_cadastrados = Cliente.query.filter(Cliente.nome.in_(list(nomes_clientes))).all()
    else:
        clientes_cadastrados = []
    mapa_cliente = {c.nome: c for c in clientes_cadastrados}

    cont_bairros_origem = Counter()
    for e in entregas:
        if not e.cliente:
            continue
        cl = mapa_cliente.get(e.cliente)
        if cl and cl.bairro_origem:
            cont_bairros_origem[(cl.bairro_origem or '').strip()] += 1

    ranking_bairros_origem = [
        {"bairro": (b or 'Não informado'), "qtd": q}
        for b, q in cont_bairros_origem.most_common()
    ]

    ranking_pgto = [{"forma": f, "qtd": q} for f, q in cont_pgto.most_common()]

    soma_por_cliente = defaultdict(lambda: {"qtd": 0, "total": 0.0})
    for e in entregas:
        if e.cliente:
            soma_por_cliente[e.cliente]["qtd"] += 1
            soma_por_cliente[e.cliente]["total"] += float(e.valor or 0)
    ranking_clientes = [
        {"cliente": c, "qtd": d["qtd"], "total": round(d["total"], 2)}
        for c, d in sorted(soma_por_cliente.items(), key=lambda kv: kv[1]["total"], reverse=True)
    ]

    dias_ordenados = sorted(list(cont_dias.keys()))
    chart_entregas_labels = [d.strftime("%d/%m") for d in dias_ordenados]
    chart_entregas_values = [cont_dias[d] for d in dias_ordenados]

    chart_faturamento_labels = [r["nome"] for r in ranking_cooperados]
    chart_faturamento_values = [r["total_valor"] for r in ranking_cooperados]

    periodo_legivel = periodo_legivel_str(data_inicio, data_fim)

    estatisticas = {
        "total": total,
        "pagas": pagas,
        "pendentes": pendentes,
        "total_valor": total_valor,
        "ticket_medio": ticket_medio,
        "dia_top": dia_top,
        "hora_pico": hora_pico,
        "pgto_top": pgto_top
    }

    por_ano_total = defaultdict(float)
    por_ano_qtd = defaultdict(int)
    for e in entregas:
        dt_local = to_brasilia(e.data_envio)
        if not dt_local:
            continue
        ano_local = dt_local.year
        if ano_local < 2025:
            continue
        por_ano_qtd[ano_local] += 1
        por_ano_total[ano_local] += float(e.valor or 0)

    if por_ano_total:
        ultimo_ano = max(set(por_ano_total.keys()) | set(por_ano_qtd.keys()))
    else:
        ultimo_ano = max(2025, datetime.now(BRAZIL_TZ).year)

    chart_ano_labels = list(range(2025, ultimo_ano + 1))
    chart_ano_totais = []
    chart_ano_qtd = []
    chart_ano_ticket = []
    for y in chart_ano_labels:
        tot = float(por_ano_total.get(y, 0.0))
        qtd = int(por_ano_qtd.get(y, 0))
        tkt = (tot / qtd) if qtd else 0.0
        chart_ano_totais.append(round(tot, 2))
        chart_ano_qtd.append(qtd)
        chart_ano_ticket.append(round(tkt, 2))

    return render_template(
        'estatisticas_cooperado.html',
        cooperados=cooperados,
        cooperado_id=cooperado_id,
        data_inicio=data_inicio,
        data_fim=data_fim,
        status_pagamento=status_pagamento,
        cliente=cliente,
        estatisticas=estatisticas,
        ranking_cooperados=ranking_cooperados,
        ranking_bairros=ranking_bairros,
        ranking_bairros_origem=ranking_bairros_origem,
        ranking_pgto=ranking_pgto,
        ranking_clientes=ranking_clientes,
        horas_pico_top3=horas_pico_top3,
        chart_entregas_labels=chart_entregas_labels,
        chart_entregas_values=chart_entregas_values,
        chart_faturamento_labels=chart_faturamento_labels,
        chart_faturamento_values=chart_faturamento_values,
        periodo_legivel=periodo_legivel,
        chart_ano_labels=chart_ano_labels,
        chart_ano_totais=chart_ano_totais,
        chart_ano_qtd=chart_ano_qtd,
        chart_ano_ticket=chart_ano_ticket,
    )



# =========================================================
# EXPORTAÇÃO ESTATÍSTICAS (MASTER)
# =========================================================
@app.route('/estatisticas_cooperado_exportar_xlsx')
@master_required
def estatisticas_cooperado_exportar_xlsx():
    cooperado_id = request.args.get('cooperado_id', 'todos')
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')
    status_pagamento = request.args.get('status_pagamento', 'todos')
    cliente = (request.args.get('cliente') or '').strip()

    query = Entrega.query
    if cooperado_id != 'todos':
        query = query.filter(Entrega.cooperado_id == int(cooperado_id))
    if data_inicio:
        di = datetime.strptime(data_inicio, "%Y-%m-%d").date()
        inicio_utc, _ = local_date_window_to_utc_range(di)
        query = query.filter(Entrega.data_envio >= inicio_utc)
    if data_fim:
        df_ = datetime.strptime(data_fim, "%Y-%m-%d").date()
        _, fim_utc = local_date_window_to_utc_range(df_)
        query = query.filter(Entrega.data_envio <= fim_utc)
    if status_pagamento and status_pagamento != 'todos':
        if status_pagamento == 'pago':
            query = query.filter(func.lower(Entrega.status_pagamento) == 'pago')
        elif status_pagamento == 'pendente':
            query = query.filter(
                (Entrega.status_pagamento == None) |
                (func.lower(Entrega.status_pagamento) == 'pendente')
            )
    if cliente:
        like = f"%{cliente.lower()}%"
        query = query.filter(func.lower(Entrega.cliente).like(like))

    entregas = query.all()

    soma_por_coop = defaultdict(lambda: {"qtd": 0, "total": 0.0})
    total_geral = 0.0
    for e in entregas:
        nm = e.cooperado.nome if e.cooperado else "Sem Cooperado"
        soma_por_coop[nm]["qtd"] += 1
        soma_por_coop[nm]["total"] += float(e.valor or 0)
        total_geral += float(e.valor or 0)

    linhas = []
    for nome, d in soma_por_coop.items():
        percent = (d["total"] / total_geral * 100.0) if total_geral > 0 else 0.0
        linhas.append({
            "Cooperado": nome,
            "Qtd Entregas": d["qtd"],
            "Valor Total (R$)": round(d["total"], 2),
            "% do Total": round(percent, 1)
        })
    linhas.sort(key=lambda r: r["Valor Total (R$)"], reverse=True)

    df_out = pd.DataFrame(linhas)
    titulo = f"Faturamento dos cooperados do período ({periodo_legivel_str(data_inicio, data_fim)})"

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        sheet = 'Resumo'
        start_row = 1
        df_out.to_excel(writer, index=False, sheet_name=sheet, startrow=start_row)
        ws = writer.sheets[sheet]

        last_col = len(df_out.columns) - 1
        ws.merge_range(
            0, 0, 0, last_col, titulo,
            writer.book.add_format({
                'bold': True, 'font_size': 14, 'align': 'center', 'valign': 'vcenter',
                'font_color': '#003399'
            })
        )

        widths = [28, 14, 18, 12]
        for i, w in enumerate(widths[:len(df_out.columns)]):
            ws.set_column(i, i, w)

        money_fmt = writer.book.add_format({'num_format': '#,##0.00'})
        pct_fmt = writer.book.add_format({'num_format': '0.0"%"'})
        cols = list(df_out.columns)
        if "Valor Total (R$)" in cols:
            idx = cols.index("Valor Total (R$)")
            ws.set_column(idx, idx, 18, money_fmt)
        if "% do Total" in cols:
            idx = cols.index("% do Total")
            ws.set_column(idx, idx, 12, pct_fmt)

    output.seek(0)
    return send_file(output, download_name="faturamento_cooperados.xlsx", as_attachment=True)


@app.route('/exportar_xlsx')
def exportar_xlsx():
    if not session.get('is_admin'):
        return redirect(url_for('login'))

    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')
    cooperado_id = request.args.get('cooperado_id', 'todos')
    cliente = (request.args.get('cliente') or '').strip()

    query = Entrega.query
    if cooperado_id != 'todos':
        query = query.filter(Entrega.cooperado_id == int(cooperado_id))
    if cliente:
        like = f"%{cliente.lower()}%"
        query = query.filter(func.lower(Entrega.cliente).like(like))
    if data_inicio:
        di = datetime.strptime(data_inicio, "%Y-%m-%d").date()
        inicio_utc, _ = local_date_window_to_utc_range(di)
        query = query.filter(Entrega.data_envio >= inicio_utc)
    if data_fim:
        df_ = datetime.strptime(data_fim, "%Y-%m-%d").date()
        _, fim_utc = local_date_window_to_utc_range(df_)
        query = query.filter(Entrega.data_envio <= fim_utc)

    entregas = query.order_by(Entrega.data_envio.asc()).all()

    rows = []
    for e in entregas:
        dt_local = to_brasilia(e.data_envio)
        rows.append({
            'Data': dt_local.strftime('%d/%m/%Y') if dt_local else '',
            'Cliente': e.cliente,
            'Bairro': e.bairro,
            'Valor': e.valor,
            'Status Pagamento': e.status_pagamento,
            'Status Entrega': e.status,
            'Forma Pagamento': e.pagamento,
            'Cooperado': (e.cooperado.nome if e.cooperado else 'Sem Cooperado'),
            'Recebido Por': e.recebido_por or ''
        })

    df_out = pd.DataFrame(rows)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        sheet = 'Entregas'
        df_out.to_excel(writer, index=False, sheet_name=sheet)
        ws = writer.sheets[sheet]
        col_widths = [12, 28, 18, 10, 18, 16, 16, 22, 18]
        for i, w in enumerate(col_widths[:len(df_out.columns)]):
            ws.set_column(i, i, w)
    output.seek(0)
    return send_file(output, download_name="entregas.xlsx", as_attachment=True)

# =========================================================
# EXPORTAR / IMPORTAR CLIENTES
# =========================================================
@app.route('/clientes/exportar')
def exportar_clientes():
    if not session.get('is_admin'):
        return redirect(url_for('login'))

    aggs = (
        db.session.query(
            Entrega.cliente.label('cli'),
            func.count(Entrega.id).label('qtd'),
            func.max(Entrega.data_envio).label('ultimo')
        )
        .group_by(Entrega.cliente)
        .all()
    )
    stats = defaultdict(lambda: {"qtd": 0, "ultimo": None})
    for row in aggs:
        key = normalize_letters_key(row.cli or '')
        s = stats[key]
        s["qtd"] += int(row.qtd or 0)
        if row.ultimo and (s["ultimo"] is None or row.ultimo > s["ultimo"]):
            s["ultimo"] = row.ultimo

    rows = []
    for c in Cliente.query.order_by(Cliente.nome).all():
        key = normalize_letters_key(c.nome or '')
        s = stats.get(key, {})
        rows.append([
            c.id, c.nome, c.telefone, c.bairro_origem, c.endereco,
            br_date_ymd(s.get("ultimo")) if s else "", int((s or {}).get("qtd") or 0)
        ])

    df_out = pd.DataFrame(rows, columns=[
        "ID", "Nome", "Telefone", "Bairro", "Endereco", "UltimoUso", "TotalPedidos"
    ])

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        sheet = 'Clientes'
        df_out.to_excel(writer, index=False, sheet_name=sheet)
        ws = writer.sheets[sheet]
        widths = [8, 28, 18, 18, 32, 12, 14]
        for i, w in enumerate(widths[:len(df_out.columns)]):
            ws.set_column(i, i, w)
    output.seek(0)
    return send_file(output, download_name="clientes.xlsx", as_attachment=True)


@app.route('/clientes/importar', methods=['POST'])
def importar_clientes():
    if not session.get('is_admin'):
        return redirect(url_for('login'))

    f = request.files.get('arquivo')
    if not f or not f.filename:
        if request.headers.get('X-Requested-With') == 'fetch' or request.args.get('format') == 'json' or (request.accept_mimetypes and request.accept_mimetypes.best == 'application/json'):
            return jsonify(ok=False, error="Envie um arquivo (.xlsx ou .csv)."), 400
        flash("Envie um arquivo (.xlsx ou .csv).")
        return redirect(url_for('clientes'))

    filename = f.filename.lower()

    try:
        raw = f.read()
        if not raw:
            raise ValueError("Arquivo vazio.")
    except Exception as e:
        msg = f"Falha ao ler upload: {e}"
        if request.headers.get('X-Requested-With') == 'fetch' or request.args.get('format') == 'json' or (request.accept_mimetypes and request.accept_mimetypes.best == 'application/json'):
            return jsonify(ok=False, error=msg), 400
        flash(msg)
        return redirect(url_for('clientes'))

    df_in = None
    load_errors = []

    if filename.endswith('.xlsx'):
        try:
            df_in = pd.read_excel(io.BytesIO(raw), engine='openpyxl', dtype=str)
        except Exception as e:
            load_errors.append(f"Pandas/openpyxl: {e}")

        if df_in is None:
            try:
                from openpyxl import load_workbook
                wb = load_workbook(io.BytesIO(raw), data_only=True)
                ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.active
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    raise ValueError("Planilha vazia.")
                header = [str(x).strip() if x is not None else '' for x in rows[0]]
                data = [[("" if c is None else str(c)) for c in r] for r in rows[1:]]
                df_in = pd.DataFrame(data, columns=header)
            except Exception as e:
                load_errors.append(f"openpyxl: {e}")

    if df_in is None and (filename.endswith('.csv') or filename.endswith('.txt')):
        try:
            df_in = pd.read_csv(io.BytesIO(raw), sep=None, engine='python', dtype=str, encoding='utf-8')
        except Exception:
            try:
                df_in = pd.read_csv(io.BytesIO(raw), sep=None, engine='python', dtype=str, encoding='latin-1')
            except Exception as e:
                load_errors.append(f"CSV: {e}")

    if df_in is None:
        msg = "Não consegui ler o arquivo. " + (" | ".join(load_errors) if load_errors else "")
        if request.headers.get('X-Requested-With') == 'fetch' or request.args.get('format') == 'json' or (request.accept_mimetypes and request.accept_mimetypes.best == 'application/json'):
            return jsonify(ok=False, error=msg), 400
        flash(msg)
        return redirect(url_for('clientes'))

    cols_map = {str(c).lower().strip(): c for c in df_in.columns}

    def colget(*ops, opt=False):
        for k in ops:
            if k in cols_map:
                return cols_map[k]
        return None if opt else None

    col_id = colget('id')
    col_nome = colget('nome', 'name')
    col_tel = colget('telefone', 'phone', 'numero', 'número', 'mobile', 'celular')
    col_bairro = colget('bairro', 'bairro_origem')
    col_end = colget('endereco', 'endereço', 'address')

    missing = []
    if not col_nome:
        missing.append("Nome")
    if not col_tel:
        missing.append("Telefone/Número")
    if missing:
        msg = f"Cabeçalho ausente: {', '.join(missing)}. Colunas recebidas: {list(df_in.columns)}"
        if request.headers.get('X-Requested-With') == 'fetch' or request.args.get('format') == 'json' or (request.accept_mimetypes and request.accept_mimetypes.best == 'application/json'):
            return jsonify(ok=False, error=msg), 400
        flash(msg)
        return redirect(url_for('clientes'))

    def norm_phone(s: str) -> str:
        if s is None:
            return ""
        digits = re.sub(r'\D+', '', str(s))
        if digits.startswith('55'):
            digits = digits[2:]
        if len(digits) > 11:
            digits = digits[-11:]
        return digits

    adicionados = 0
    atualizados = 0
    erros = 0
    detalhes = []

    for i, row in df_in.iterrows():
        try:
            rid = None
            if col_id and not pd.isna(row.get(col_id)):
                try:
                    rid = int(str(row[col_id]).strip())
                except Exception:
                    rid = None

            nome = str(row.get(col_nome) or '').strip()
            tel = norm_phone(row.get(col_tel))
            bairro = str(row.get(col_bairro) or '').strip() if col_bairro else None
            ender = str(row.get(col_end) or '').strip() if col_end else None

            if not nome and not tel:
                continue

            if tel and len(tel) not in (10, 11):
                erros += 1
                detalhes.append(
                    f"Linha {i+2}: telefone inválido '{tel}' (esperado 10 ou 11 dígitos)."
                )
                continue

            if rid:
                cl = Cliente.query.get(rid)
                if not cl:
                    cl = Cliente.query.filter(func.lower(Cliente.nome) == nome.lower()).first()
            else:
                cl = Cliente.query.filter(func.lower(Cliente.nome) == nome.lower()).first()

            if cl:
                cl.nome = nome
                cl.telefone = tel or None
                cl.bairro_origem = bairro or None
                cl.endereco = ender or None
                atualizados += 1
            else:
                novo = Cliente(
                    nome=nome,
                    telefone=tel or None,
                    bairro_origem=bairro or None,
                    endereco=ender or None
                )
                db.session.add(novo)
                adicionados += 1

        except Exception as e:
            erros += 1
            detalhes.append(f"Linha {i+2}: erro inesperado ({e}).")

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        app.logger.exception("Erro ao salvar no banco durante importação")
        msg = "Erro ao salvar no banco."
        if app.debug:
            msg += f" Detalhes: {e}"
        if request.headers.get('X-Requested-With') == 'fetch' or request.args.get('format') == 'json' or (request.accept_mimetypes and request.accept_mimetypes.best == 'application/json'):
            return jsonify(ok=False, error=msg), 500
        flash(msg)
        return redirect(url_for('clientes'))

    if request.headers.get('X-Requested-With') == 'fetch' or request.args.get('format') == 'json' or (request.accept_mimetypes and request.accept_mimetypes.best == 'application/json'):
        return jsonify(
            ok=True,
            adicionados=adicionados,
            atualizados=atualizados,
            erros=erros,
            detalhes=detalhes
        )
    else:
        flash(
            f'Importação concluída: {adicionados} adicionados, '
            f'{atualizados} atualizados, {erros} erros.'
        )
        return redirect(url_for('clientes'))

# =========================================================
# FILA DE ESPERA
# =========================================================
@app.route('/lista_espera/add', methods=['POST'])
def lista_espera_add():
    if not session.get('is_admin'):
        return redirect(url_for('login'))

    cooperado_id = request.form.get('cooperado_id')
    nome_form = (request.form.get('nome') or '').strip()

    if not cooperado_id and not nome_form:
        flash('Selecione um cooperado ou informe um nome.')
        return redirect_back_to_admin()

    if cooperado_id:
        coop = Cooperado.query.get(int(cooperado_id))
        if not coop:
            flash('Cooperado inválido.')
            return redirect_back_to_admin()

        if ListaEspera.query.filter_by(cooperado_id=coop.id).first():
            flash('Este cooperado já está na fila de espera.')
            return redirect_back_to_admin()

        max_pos = db.session.query(func.max(ListaEspera.pos)).scalar() or 0
        item = ListaEspera(
            cooperado_id=coop.id,
            nome=coop.nome,
            pos=max_pos + 1,
            created_at=datetime.utcnow()
        )
        db.session.add(item)
        db.session.commit()
        flash('Cooperado adicionado à lista de espera.')
        return redirect_back_to_admin()

    if ListaEspera.query.filter(func.lower(ListaEspera.nome) == nome_form.lower()).first():
        flash('Este nome já está na fila de espera.')
        return redirect_back_to_admin()

    max_pos = db.session.query(func.max(ListaEspera.pos)).scalar() or 0
    item = ListaEspera(
        nome=nome_form,
        cooperado_id=None,
        pos=max_pos + 1,
        created_at=datetime.utcnow()
    )
    db.session.add(item)
    db.session.commit()
    flash('Nome adicionado à lista de espera.')
    return redirect_back_to_admin()


@app.route('/lista_espera/remove/<int:id>', methods=['POST'])
def lista_espera_remove(id):
    if not session.get('is_admin'):
        return redirect(url_for('login'))

    item = ListaEspera.query.get_or_404(id)
    db.session.delete(item)
    db.session.commit()
    flash('Removido da lista de espera.')
    return redirect_back_to_admin()


@app.route('/lista_espera/reordenar', methods=['POST'])
def lista_espera_reordenar():
    if not session.get('is_admin'):
        return ("", 403)
    data = request.get_json(silent=True) or {}
    ordem = data.get('ordem') or []
    try:
        for i, sid in enumerate(ordem, start=1):
            try:
                _id = int(sid)
            except Exception:
                continue
            db.session.query(ListaEspera).filter_by(id=_id).update({"pos": i})
        db.session.commit()
        return ("", 204)
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Reordenar fila falhou: {e}")
        return ("", 500)

# =========================================================
# RELATÓRIO TÉRMICO
# =========================================================
@app.route('/relatorio_termico')
def relatorio_termico():
    if not session.get('is_admin'):
        return redirect(url_for('login'))

    cooperado_id = request.args.get('cooperado_id', 'todos')
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')
    status_pagamento = request.args.get('status_pagamento', 'todos')
    cliente = (request.args.get('cliente') or '').strip()

    if data_inicio:
        di_date = datetime.strptime(data_inicio, "%Y-%m-%d").date()
        inicio_utc, _ = local_date_window_to_utc_range(di_date)
    else:
        hoje = datetime.now(BRAZIL_TZ).date()
        inicio_utc, _ = local_date_window_to_utc_range(hoje)

    if data_fim:
        df_date = datetime.strptime(data_fim, "%Y-%m-%d").date()
        _, fim_utc = local_date_window_to_utc_range(df_date)
    else:
        base = datetime.strptime(data_inicio, "%Y-%m-%d").date() if data_inicio else datetime.now(BRAZIL_TZ).date()
        _, fim_utc = local_date_window_to_utc_range(base)

    q = Entrega.query

    if cooperado_id and cooperado_id != 'todos':
        try:
            q = q.filter(Entrega.cooperado_id == int(cooperado_id))
        except Exception:
            pass

    if status_pagamento and status_pagamento != 'todos':
        if status_pagamento == 'pago':
            q = q.filter(func.lower(Entrega.status_pagamento) == 'pago')
        elif status_pagamento == 'pendente':
            q = q.filter(
                (Entrega.status_pagamento == None) |
                (func.lower(Entrega.status_pagamento) == 'pendente')
            )

    if cliente:
        like = f"%{cliente.lower()}%"
        q = q.filter(func.lower(Entrega.cliente).like(like))

    coalesce_dt = func.coalesce(Entrega.data_atribuida, Entrega.data_envio)
    q = q.filter(coalesce_dt >= inicio_utc, coalesce_dt <= fim_utc).order_by(
        coalesce_dt.asc(),
        Entrega.cliente.asc()
    )

    entregas = q.options(joinedload(Entrega.cooperado)).all()

    periodo_txt = periodo_legivel_str(data_inicio, data_fim)

    coop_nome = "Todos"
    if cooperado_id and cooperado_id != "todos":
        coop = Cooperado.query.get(int(cooperado_id))
        if coop:
            coop_nome = coop.nome

    total_relatorio = sum(float(e.valor or 0) for e in entregas)
    agora = datetime.now(BRAZIL_TZ)

    return render_template(
        'relatorio_termico.html',
        entregas=entregas,
        periodo_txt=periodo_txt,
        coop_nome=coop_nome,
        agora=agora,
        to_brasilia=to_brasilia,
        total_relatorio=total_relatorio
    )

# =========================================================
# BOOTSTRAP BANCO / DDL / ÍNDICES / BACKFILL
# =========================================================
def criar_bd():
    with app.app_context():
        db.create_all()

        try:
            db.session.execute(text("PRAGMA foreign_keys = ON"))
        except Exception:
            pass

        ddl_cmds = [
            "ALTER TABLE lista_espera ADD COLUMN IF NOT EXISTS cooperado_id INTEGER",
            "ALTER TABLE lista_espera ADD COLUMN IF NOT EXISTS pos INTEGER",
            "ALTER TABLE lista_espera ADD COLUMN IF NOT EXISTS created_at TIMESTAMP",

            "ALTER TABLE cliente ADD COLUMN IF NOT EXISTS endereco VARCHAR(255)",
            "ALTER TABLE cliente ADD COLUMN IF NOT EXISTS saldo_atual REAL DEFAULT 0",
            "ALTER TABLE cliente ADD COLUMN IF NOT EXISTS username VARCHAR(80)",
            "ALTER TABLE cliente ADD COLUMN IF NOT EXISTS senha_hash VARCHAR(128)",

            "ALTER TABLE cliente ADD COLUMN IF NOT EXISTS email VARCHAR(120)",
            "ALTER TABLE cliente ADD COLUMN IF NOT EXISTS reset_code VARCHAR(10)",
            "ALTER TABLE cliente ADD COLUMN IF NOT EXISTS reset_expires_at TIMESTAMP",

            "ALTER TABLE entrega ADD COLUMN IF NOT EXISTS credito_usado REAL DEFAULT 0",
            "ALTER TABLE entrega ADD COLUMN IF NOT EXISTS credito_mov_id INTEGER",
            "ALTER TABLE entrega ADD COLUMN IF NOT EXISTS cliente_id INTEGER",

            "ALTER TABLE entrega ADD COLUMN IF NOT EXISTS origem_json TEXT",
            "ALTER TABLE entrega ADD COLUMN IF NOT EXISTS destino_json TEXT",
            
            "ALTER TABLE credito ADD COLUMN IF NOT EXISTS desconto_tipo VARCHAR(20) DEFAULT 'nenhum'",
            "ALTER TABLE credito ADD COLUMN IF NOT EXISTS desconto_valor REAL DEFAULT 0",
            "ALTER TABLE credito ADD COLUMN IF NOT EXISTS valor_final REAL",
            "ALTER TABLE credito ADD COLUMN IF NOT EXISTS motivo VARCHAR(180)",
            "ALTER TABLE credito ADD COLUMN IF NOT EXISTS saldo_antes REAL DEFAULT 0",
            "ALTER TABLE credito ADD COLUMN IF NOT EXISTS saldo_depois REAL DEFAULT 0",
            "ALTER TABLE credito ADD COLUMN IF NOT EXISTS criado_por VARCHAR(80)",
            "ALTER TABLE credito ADD COLUMN IF NOT EXISTS criado_em TIMESTAMP",

            "ALTER TABLE credito_movimento ADD COLUMN IF NOT EXISTS cliente_id INTEGER",
            "ALTER TABLE credito_movimento ADD COLUMN IF NOT EXISTS tipo VARCHAR(10)",
            "ALTER TABLE credito_movimento ADD COLUMN IF NOT EXISTS valor REAL",
            "ALTER TABLE credito_movimento ADD COLUMN IF NOT EXISTS referencia VARCHAR(120)",
            "ALTER TABLE credito_movimento ADD COLUMN IF NOT EXISTS criado_em TIMESTAMP",
            "ALTER TABLE credito_movimento ADD COLUMN IF NOT EXISTS credito_id INTEGER",
            "ALTER TABLE credito_movimento ADD COLUMN IF NOT EXISTS entrega_id INTEGER",
        ]
        for s in ddl_cmds:
            try:
                db.session.execute(text(s))
            except Exception:
                pass

        fk_cmds_create_if_missing = [
            (
                "DO $$ BEGIN "
                "IF NOT EXISTS (SELECT 1 FROM information_schema.table_constraints "
                "WHERE constraint_name='lista_espera_cooperado_id_fkey') THEN "
                "ALTER TABLE lista_espera ADD CONSTRAINT lista_espera_cooperado_id_fkey "
                "FOREIGN KEY (cooperado_id) REFERENCES cooperado(id) ON DELETE SET NULL; "
                "END IF; END $$;"
            ),
            (
                "DO $$ BEGIN "
                "IF NOT EXISTS (SELECT 1 FROM information_schema.table_constraints "
                "WHERE constraint_name='entrega_cooperado_id_fkey') THEN "
                "ALTER TABLE entrega ADD CONSTRAINT entrega_cooperado_id_fkey "
                "FOREIGN KEY (cooperado_id) REFERENCES cooperado(id) ON DELETE SET NULL; "
                "END IF; END $$;"
            ),
            (
                "DO $$ BEGIN "
                "IF NOT EXISTS (SELECT 1 FROM information_schema.table_constraints "
                "WHERE constraint_name='entrega_cliente_id_fkey') THEN "
                "ALTER TABLE entrega ADD CONSTRAINT entrega_cliente_id_fkey "
                "FOREIGN KEY (cliente_id) REFERENCES cliente(id) ON DELETE SET NULL; "
                "END IF; END $$;"
            ),
            (
                "DO $$ BEGIN "
                "IF NOT EXISTS (SELECT 1 FROM information_schema.table_constraints "
                "WHERE constraint_name='credito_cliente_id_fkey') THEN "
                "ALTER TABLE credito ADD CONSTRAINT credito_cliente_id_fkey "
                "FOREIGN KEY (cliente_id) REFERENCES cliente(id) ON DELETE CASCADE; "
                "END IF; END $$;"
            ),
            (
                "DO $$ BEGIN "
                "IF NOT EXISTS (SELECT 1 FROM information_schema.table_constraints "
                "WHERE constraint_name='credito_movimento_cliente_id_fkey') THEN "
                "ALTER TABLE credito_movimento ADD CONSTRAINT credito_movimento_cliente_id_fkey "
                "FOREIGN KEY (cliente_id) REFERENCES cliente(id) ON DELETE CASCADE; "
                "END IF; END $$;"
            ),
            (
                "DO $$ BEGIN "
                "IF NOT EXISTS (SELECT 1 FROM information_schema.table_constraints "
                "WHERE constraint_name='credito_movimento_credito_id_fkey') THEN "
                "ALTER TABLE credito_movimento ADD CONSTRAINT credito_movimento_credito_id_fkey "
                "FOREIGN KEY (credito_id) REFERENCES credito(id) ON DELETE SET NULL; "
                "END IF; END $$;"
            ),
            (
                "DO $$ BEGIN "
                "IF NOT EXISTS (SELECT 1 FROM information_schema.table_constraints "
                "WHERE constraint_name='credito_movimento_entrega_id_fkey') THEN "
                "ALTER TABLE credito_movimento ADD CONSTRAINT credito_movimento_entrega_id_fkey "
                "FOREIGN KEY (entrega_id) REFERENCES entrega(id) ON DELETE SET NULL; "
                "END IF; END $$;"
            ),
        ]
        for s in fk_cmds_create_if_missing:
            try:
                db.session.execute(text(s))
            except Exception:
                pass

        fix_fk_cmd = (
            "DO $$ DECLARE del CHAR; BEGIN "
            "IF EXISTS (SELECT 1 FROM pg_constraint WHERE conname='credito_movimento_entrega_id_fkey') THEN "
            "  SELECT c.confdeltype INTO del FROM pg_constraint c WHERE c.conname='credito_movimento_entrega_id_fkey'; "
            "  IF del IS DISTINCT FROM 'n' THEN "
            "    ALTER TABLE credito_movimento DROP CONSTRAINT credito_movimento_entrega_id_fkey; "
            "    ALTER TABLE credito_movimento ADD CONSTRAINT credito_movimento_entrega_id_fkey "
            "    FOREIGN KEY (entrega_id) REFERENCES entrega(id) ON DELETE SET NULL; "
            "  END IF; "
            "END IF; "
            "END $$;"
        )
        try:
            db.session.execute(text(fix_fk_cmd))
        except Exception:
            pass

        idx_cmds = [
            "CREATE INDEX IF NOT EXISTS idx_entrega_data_envio ON entrega (data_envio DESC)",
            "CREATE INDEX IF NOT EXISTS idx_entrega_cooperado_id ON entrega (cooperado_id)",
            "CREATE INDEX IF NOT EXISTS idx_entrega_cliente_id ON entrega (cliente_id)",
            "CREATE INDEX IF NOT EXISTS idx_entrega_status_pagamento_lower ON entrega ((lower(status_pagamento)))",
            "CREATE INDEX IF NOT EXISTS idx_entrega_cliente_lower ON entrega ((lower(cliente)))",

            "CREATE INDEX IF NOT EXISTS idx_lista_espera_pos ON lista_espera (pos ASC)",

            "CREATE INDEX IF NOT EXISTS idx_cliente_nome_lower ON cliente ((lower(nome)))",
            "CREATE UNIQUE INDEX IF NOT EXISTS idx_cliente_username ON cliente (username)",
            "CREATE UNIQUE INDEX IF NOT EXISTS idx_cliente_email ON cliente (email)",

            "CREATE INDEX IF NOT EXISTS idx_credito_cliente_id ON credito (cliente_id)",
            "CREATE INDEX IF NOT EXISTS idx_credito_criado_em ON credito (criado_em DESC)",

            "CREATE INDEX IF NOT EXISTS idx_credmov_cliente_id ON credito_movimento (cliente_id)",
            "CREATE INDEX IF NOT EXISTS idx_credmov_entrega_id ON credito_movimento (entrega_id)",
            "CREATE INDEX IF NOT EXISTS idx_credmov_criado_em ON credito_movimento (criado_em DESC)",
            "CREATE INDEX IF NOT EXISTS idx_credmov_tipo ON credito_movimento (tipo)",

            "CREATE INDEX IF NOT EXISTS idx_trajeto_cooperado_id ON trajeto (cooperado_id)",
            "CREATE INDEX IF NOT EXISTS idx_trajeto_inicio ON trajeto (inicio DESC)",
        ]
        for s in idx_cmds:
            try:
                db.session.execute(text(s))
            except Exception:
                pass

        try:
            pend = (
                Entrega.query
                .filter((Entrega.cliente_id == None) | (Entrega.cliente_id.is_(None)))
                .limit(5000)
                .all()
            )
            if pend:
                nomes = {(e.cliente or '').strip().lower() for e in pend if (e.cliente or '').strip()}
                if nomes:
                    mapa = {
                        c.nome.strip().lower(): c.id
                        for c in Cliente.query.filter(
                            func.lower(Cliente.nome).in_(list(nomes))
                        ).all()
                        if (c.nome or '').strip()
                    }
                    mudou = 0
                    for e in pend:
                        cid = mapa.get((e.cliente or '').strip().lower())
                        if cid:
                            e.cliente_id = cid
                            mudou += 1
                    if mudou:
                        db.session.commit()
        except Exception:
            db.session.rollback()

        db.session.commit()

criar_bd()

# =========================================================
# EVENTOS SOCKET.IO (TEMPO REAL)
# =========================================================

from flask import request

# Conexão do cliente
@socketio.on("connect")
def handle_connect(auth=None):
    try:
        if current_user.is_authenticated and getattr(current_user, "tipo", "") == "admin":
            join_room("admins")
    except Exception:
        pass


# Desconexão do cliente
@socketio.on("disconnect")
def handle_disconnect(reason=None):
    # O Socket.IO passa 1 argumento (normalmente o 'reason'), por isso reason=None
    print(f"Cliente desconectado do Socket.IO: sid={request.sid}, reason={reason}")


@socketio.on("entrar_sala")
def handle_entrar_sala(data):
    """
    data esperado:
    {
        "sala": "entrega_123" ou "chat_456",
        "usuario_id": 123  # opcional, se você quiser identificar quem entrou
    }
    """
    sala = data.get("sala")
    if not sala:
        return

    usuario_id = data.get("usuario_id")

    join_room(sala)

    # avisa todo mundo da sala que alguém entrou
    emit(
        "status",
        {
            "tipo": "entrada",
            "sala": sala,
            "usuario": usuario_id,
        },
        room=sala,
    )


@socketio.on("sair_sala")
def handle_sair_sala(data):
    """
    data esperado:
    {
        "sala": "entrega_123" ou "chat_456",
        "usuario_id": 123  # opcional
    }
    """
    sala = data.get("sala")
    if not sala:
        return

    usuario_id = data.get("usuario_id")

    leave_room(sala)

    emit(
        "status",
        {
            "tipo": "saida",
            "sala": sala,
            "usuario": usuario_id,
        },
        room=sala,
    )


@socketio.on("nova_mensagem")
def handle_nova_mensagem(data):
    """
    data esperado (exemplo):
    {
        "sala": "entrega_123",
        "remetente_id": 1,                    # id de quem mandou
        "remetente_tipo": "cliente"/"admin"/"motoboy",
        "texto": "Olá, estou a caminho",
        "extra": {...}   # opcional
    }
    """
    sala = data.get("sala")
    texto = data.get("texto")

    if not sala or not texto:
        return

    payload = {
        "sala": sala,
        "texto": texto,
        "remetente_id": data.get("remetente_id"),
        "remetente_tipo": data.get("remetente_tipo"),
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "extra": data.get("extra") or {},
    }

    # Envia a mensagem para todo mundo que está na sala (cliente, motoboy, admin)
    emit("mensagem_recebida", payload, room=sala)


@socketio.on("atualizar_entrega")
def handle_atualizar_entrega(data):
    """
    data esperado (exemplo):
    {
        "entrega_id": 123,
        "campos": {
            "status_entrega": "em_andamento",
            "status_pagamento": "pago"
        }
    }

    Aqui NÃO estamos mexendo no banco,
    só avisando em tempo real pros navegadores atualizarem a tela.
    """
    entrega_id = data.get("entrega_id")
    if not entrega_id:
        return

    emit(
        "entrega_atualizada",
        {
            "entrega_id": entrega_id,
            "campos": data.get("campos") or {},
        },
        room=f"entrega_{entrega_id}",
    )



# =========================================================
# LINK DE RASTREIO (por entrega) — desativa ao concluir
# =========================================================
@app.get("/rastreio/<token>")
def rastreio_publico(token):
    try:
        data = ler_token_rastreio(token)
        entrega_id = int(data.get("entrega_id"))
    except Exception:
        return "<h2>Link inválido.</h2>", 400

    e = Entrega.query.get(entrega_id)
    if not e:
        return "<h2>Entrega não encontrada.</h2>", 404

    st = (e.status or "").lower()
    if st in ["recebido", "entregue", "concluido", "concluída", "concluida"]:
        return "<h2>Rastreio encerrado: entrega concluída.</h2>", 410

    coop_nome = (e.cooperado.nome if getattr(e, "cooperado", None) else "Cooperado")
    html = f"""<!doctype html>
<html lang="pt-br"><head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Rastreio — Entrega #{entrega_id}</title>
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css">
<style>
  body{{margin:0;font-family:system-ui,-apple-system,Segoe UI,Roboto,Arial;background:#0b1220;color:#fff}}
  header{{padding:10px 12px;background:linear-gradient(90deg,#0b2cc2,#1a47ff);font-weight:800}}
  #map{{height: calc(100vh - 54px); width:100%}}
  .small{{opacity:.9;font-weight:700}}
</style>
</head><body>
<header>Rastreio em tempo real — Entrega #{entrega_id} <span class="small">({coop_nome})</span></header>
<div id="map"></div>
<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<script>
  const token = {json.dumps(token)};
  const map = L.map('map', {{ zoomControl:true }}).setView([-5.7945,-35.2110], 13);
  L.tileLayer('https://{{s}}.tile.openstreetmap.org/{{z}}/{{x}}/{{y}}.png', {{ maxZoom: 19 }}).addTo(map);
  let marker = null;

  async function pull(){{
    try{{
      const r = await fetch('/api/rastreio_pos/'+encodeURIComponent(token), {{cache:'no-store'}});
      if(r.status === 410){{
        document.body.innerHTML = '<h2 style="padding:16px">Rastreio encerrado: entrega concluída.</h2>';
        return;
      }}
      const data = await r.json();
      if(!data.ok) return;

      const lat = data.lat, lng = data.lng;
      if(typeof lat !== 'number' || typeof lng !== 'number') return;

      const txt = (data.cooperado || '') + ' • ' + (data.quando_local || '');
      if(!marker){{
        marker = L.circleMarker([lat,lng], {{
          radius: 7,
          weight: 2,
          fillOpacity: 0.8
        }}).addTo(map);
        marker.bindTooltip(txt, {{direction:'top', sticky:true}});
        map.setView([lat,lng], 15);
      }} else {{
        marker.setLatLng([lat,lng]);
        marker.setTooltipContent(txt);
      }}
    }}catch(e){{}}
  }}
  pull();
  setInterval(pull, 5000);
</script>
</body></html>"""
    return html

@app.get("/api/rastreio_pos/<token>")
def api_rastreio_pos(token):
    try:
        data = ler_token_rastreio(token)
        entrega_id = int(data.get("entrega_id"))
    except Exception:
        return jsonify(ok=False, error="invalid_token"), 400

    e = Entrega.query.get(entrega_id)
    if not e:
        return jsonify(ok=False, error="not_found"), 404

    st = (e.status or "").lower()
    if st in ["recebido", "entregue", "concluido", "concluída", "concluida"]:
        return jsonify(ok=False, error="ended"), 410

    coop = getattr(e, "cooperado", None)
    if not coop or coop.last_lat is None or coop.last_lng is None:
        return jsonify(ok=True, lat=None, lng=None, cooperado=(coop.nome if coop else None), quando_local=None)

    when_local = None
    try:
        if coop.last_ping:
            when_local = to_brasilia(coop.last_ping).strftime("%d/%m/%Y %H:%M:%S")
    except Exception:
        when_local = None

    return jsonify(ok=True,
                   lat=float(coop.last_lat),
                   lng=float(coop.last_lng),
                   cooperado=coop.nome,
                   quando_local=when_local)


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    # importante rodar pelo socketio, não pelo app.run
    socketio.run(app, host='0.0.0.0', port=port)
